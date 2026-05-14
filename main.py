import logging
import os
import uuid
import asyncio
import json
import html
import hashlib
import tempfile
import base64
import requests
import re
import time
import random
import resend
from datetime import datetime, timezone
from zoneinfo import ZoneInfo
from typing import Optional
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from PIL import Image
import io
import docx
from duckduckgo_search import DDGS

# --- Kriptografiya kutubxonalari ---
from cryptography.fernet import Fernet
from cryptography.hazmat.primitives.asymmetric import ed25519
from cryptography.hazmat.primitives import serialization

from telegram import (
    Update,
    InlineKeyboardMarkup,
    InlineKeyboardButton,
    WebAppInfo,
    ReplyKeyboardMarkup,
    KeyboardButton,
    ReplyKeyboardRemove,
    LabeledPrice
)
from telegram.constants import ChatMemberStatus, ParseMode
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    ChatMemberHandler,
    ContextTypes,
    TypeHandler,
    ApplicationHandlerStop,
    PreCheckoutQueryHandler,
    filters,
)

from db import DB

TZ = ZoneInfo("Asia/Tashkent")

# ==========================================
# 1. KONFIGURATSIYA VA SOZLAMALAR
# ==========================================
BOT_TOKEN = os.getenv("BOT_TOKEN", "").strip()
if not BOT_TOKEN:
    raise RuntimeError("ENV BOT_TOKEN yo'q")

WEB_BASE_URL = os.getenv("WEB_BASE_URL", "http://localhost:8000")

SUPERADMINS = {int(x) for x in (os.getenv("SUPERADMINS", "") or "").split(",") if x.strip().isdigit()}
LOWER_ADMINS = {int(x) for x in (os.getenv("LOWER_ADMINS", "") or "").split(",") if x.strip().isdigit()}

ADMIN_CARD = os.getenv("ADMIN_CARD", "0000 0000 0000 0000 (Ism Familiya)")
REQUIRED_CHANNEL = os.getenv("REQUIRED_CHANNEL", "@your_channel")
PUBLIC_TEST_CHANNEL = os.getenv("PUBLIC_TEST_CHANNEL", "@your_public_channel")

try:
    MAIN_ADMIN_ID = int(os.getenv("MAIN_ADMIN_ID", "0"))
except ValueError:
    MAIN_ADMIN_ID = 0

GROQ_API_KEY = os.getenv("GROQ_API_KEY", "").strip()
GROQ_MODEL = os.getenv("GROQ_MODEL", "llama-3.3-70b-versatile")
VISION_MODEL = "llama-3.2-90b-vision-preview"
GROQ_URL = "https://api.groq.com/openai/v1/chat/completions"

RESEND_API_KEY = os.getenv("RESEND_API_KEY", "")
resend.api_key = RESEND_API_KEY

db = DB()

# ==========================================
# POCHTAGA KOD YUBORISH
# ==========================================
def send_reset_code_email(to_email, code):
    try:
        resend.Emails.send({
            "from": os.getenv("RESEND_FROM_EMAIL", "onboarding@resend.dev"),
            "to": to_email,
            "subject": "Geo Ustoz - PIN kodni tiklash",
            "html": f"""
            <div style="font-family: sans-serif; padding: 20px;">
                <h2>Assalomu alaykum!</h2>
                <p>Geo Ustoz platformasida PIN kodingizni tiklash uchun so'rov yubordingiz.</p>
                <p>Sizning tasdiqlash kodingiz: <b style="font-size: 24px; color: #3b82f6;">{code}</b></p>
                <p><i>Agar bu so'rovni siz yubormagan bo'lsangiz, ushbu xatni e'tiborsiz qoldiring.</i></p>
                <p><i>Bu kodni hech kimga bermang hattoki @geo_ustoz_bot danmiz deyishsa ham!!!<i><p>
            </div>
            """
        })
        return True
    except Exception as e:
        logging.error(f"Email yuborishda xato: {e}")
        return False

# ==========================================
# 🔐 KRIPTOGRAFIYA MENEJERI
# ==========================================
class CryptoManager:
    def __init__(self, secret_token):
        hasher = hashlib.sha256(secret_token.encode())
        self.fernet = Fernet(base64.urlsafe_b64encode(hasher.digest()))

    def generate_wallet(self):
        private_key = ed25519.Ed25519PrivateKey.generate()
        public_key = private_key.public_key()

        priv_bytes = private_key.private_bytes(
            encoding=serialization.Encoding.Raw,
            format=serialization.PrivateFormat.Raw,
            encryption_algorithm=serialization.NoEncryption()
        )
        pub_bytes = public_key.public_bytes(
            encoding=serialization.Encoding.Raw,
            format=serialization.PublicFormat.Raw
        )

        enc_priv = self.fernet.encrypt(priv_bytes).decode('utf-8')
        pub_hex = pub_bytes.hex()
        return pub_hex, enc_priv

    def sign_transaction(self, enc_priv, message_bytes):
        priv_bytes = self.fernet.decrypt(enc_priv.encode('utf-8'))
        private_key = ed25519.Ed25519PrivateKey.from_private_bytes(priv_bytes)
        signature = private_key.sign(message_bytes)
        return signature.hex()

crypto_mgr = CryptoManager(BOT_TOKEN)

K = {
    "mode": "mode",
    "create": "create",
    "broadcast": "broadcast",
    "convert": "convert",
}

RUNTIME = {
    "pending_ready": "pending_ready",
    "running_session": "running_session",
}

# ==========================================
# 2. TILLAR LUG'ATI
# ==========================================
BOT_LANGS = {
    "uz": {
        "welcome": "👋 <b>Salom, {name}!</b>\n\nGeo Ustoz botiga xush kelibsiz.\n\nPastdagi menyu orqali kerakli bo'limni tanlang 👇",
        "account": "👤 <b>Sizning hisobingiz:</b>\n\n🆔 ID: <code>{user_id}</code>\n💼 Maqom: <b>{status}</b>{premium_text}\n🌐 Til: 🇺🇿 O'zbek (Lotin)\n\n<i>Premium maqomi orqali barcha cheklovlardan xalos bo'lasiz!</i>",
        "change_lang_btn": "🌐 Tilni o'zgartirish",
        "choose_lang": "Iltimos, o'zingizga qulay tilni tanlang:",
        "lang_saved": "✅ Til muvaffaqiyatli o'zgartirildi!",
        "btn_cabinet": "🖥 Shaxsiy kabinet",
        "btn_account": "👤 Hisobim",
        "btn_wallet": "👛 Mening Hamyonim",
        "btn_referral": "🔗 Do'stlarni taklif qilish",
        "btn_top": "🏆 Reyting va O'rnim",
        "wallet_info": "👛 <b>GEO WALLET (GWT)</b>\n\n🔑 Hamyon manzili:\n<code>{address}</code>\n\n💰 Balans: <b>{balance} GWT</b>\n\n<i>Quyidagi tugmalar orqali token o'tkazishingiz yoki qabul qilishingiz mumkin.</i>",
        "btn_transfer": "💸 Token o'tkazish",
        "btn_receive": "📥 Qabul qilish",
        "btn_ai": "🤖 AI bilan gaplashish",
        "btn_search": "🔍 Qidiruv",
        "btn_create_manual": "🧩 qo'lda test yaratish",
        "btn_create_word": "📝 Worddan test yaratish",
        "btn_results": "📊 Natijalar / Excel",
        "btn_check_chats": "✅ Chatlarni tekshirish",
        "btn_premium": "💎 Premium sotib olish",
        "btn_add_bot": "➕ Botni guruh/kanalga qo'shish",
        "btn_lock": "🔐 Qulflash",
        "btn_back": "⬅️ Orqaga",
        "btn_back_main": "⬅️ Asosiy menyu",
        "btn_cancel": "❌ Bekor qilish",
        "btn_yes": "✅ Ha",
        "btn_no": "❌ Yo'q",
        "btn_solve_site": "🌐 Saytda ishlash (Qulay)",
        "btn_solve_bot": "💬 Botda ishlash",
        "btn_finish_test_session": "🛑 Testni tugatish",
        "btn_export_excel": "📤 Excel yuklab olish",
        "btn_close_test": "⛔ Testni yopish",
        "btn_publish_test": "🌐 Public qilish (@nom)",
        "btn_delete_test": "🗑 Testni o'chirish",
        "btn_pay_card": "💳 Karta (Chek yuborish)",
        "btn_pay_stars": "⭐️ Telegram Stars (Avtomatik)",
        "btn_pay_token": "🪙 GWT Token (Blokcheyn)",
        "btn_create_private": "🔒 Private test (faqat o'zim uchun)",
        "month_1": "1 Oy", "month_3": "3 Oy", "month_6": "6 Oy", "month_12": "1 Yil",
        "captcha_msg": "🔒 <b>Xavfsizlik tekshiruvi!</b>\n\nGeo Ustoz botidan foydalanish uchun, iltimos, pastdagi tugmani bosib inson ekanligingizni tasdiqlang.",
        "test_manage_msg": "⚙️ <b>Test boshqaruvi:</b>\n{title}\n\nNima amal bajaramiz?",
        "test_options_msg": "🧩 <b>Test: {title}</b>\n⏱ <b>Vaqt:</b> {time}\n📊 <b>Baholash:</b> {scoring}\n🔄 <b>Urinishlar limiti:</b> {limit}\n\nTestni qayerda ishlashni xohlaysiz?",
        "prem_method_msg": "💎 <b>Premium xarid qilish usulini tanlang:</b>\n\n💳 <b>Karta orqali</b> to'lasangiz, chekni adminga yuborasiz va tasdiqlashini kutasiz.\n⭐️ <b>Telegram Yulduzlari</b> orqali to'lasangiz, Premium <b>avtomatik va shu zahoti</b> beriladi!",
        "prem_card_msg": "💳 <b>Karta orqali to'lov ta'riflari:</b>\n<i>Karta orqali to'lasangiz to'lov chekini yuborishingiz kerak bo'ladi.</i>",
        "prem_stars_msg": "⭐️ <b>Telegram Stars ta'riflari:</b>\n<i>Premium avtomatik tarzda shu zahoti beriladi!</i>",
        "prem_card_info": "💳 <b>To'lov uchun ma'lumotlar:</b>\n\n💰 Summa: <b>{price} so'm</b>\n💳 Karta: <code>{card}</code>\n\n👇 To'lovni amalga oshirgach, <b>CHEK (skrinshot) rasmini shu yerga yuboring!</b>",
        "lock_system": "🔐 <b>Tizim qulflangan!</b>\n\nDavom etish uchun PIN kodingizni kiriting:",
        "unlock_success": "🔓 <b>Qulf ochildi!</b>",
        "wrong_pin": "❌ <b>Noto'g'ri PIN-kod! Qaytadan urinib ko'ring.</b>",
        "no_pin_setup": "⚠️ Sizda hali PIN kod o'rnatilmagan! Qulflash uchun avval <b>Shaxsiy kabinet</b> (sayt) orqali PIN kod o'rnating.",
        "ai_welcome": "🤖 <b>Geo AI</b>ga xush kelibsiz! Savolingizni yozing yoki rasm yuboring.\n\n<i>(Asosiy menyuga qaytish uchun /start ni bosing yoki [Menyuga qaytish] deb yozing)</i>",
        "ai_thinking": "<i>AI javob yozmoqda... ⏳</i>",
        "ai_pic_thinking": "<i>AI rasmni tahlil qilmoqda... ⏳</i>",
        "ai_limit": "❌ Bugungi AI xabarlar limitingiz (10 ta) tugadi.\nAsosiy menyuga qaytarildingiz.",
        "create_where": "📝 Qaysi kanal/guruhga test yuborasiz?",
        "convert_where": "📝 <b>Word dan test yaratish</b>\n\nQaysi joyga yuborilsin? Avval shuni tanlang, keyin sizdan Word fayli so'raladi.",
        "search_prompt": "🔍 @testname orqali qidiruv. @testname kiriting:",
        "search_results": "🔍 Qidiruv natijalari:",
        "search_not_found": "🤷‍♂️ Hech narsa topilmadi. Nomi to'g'riligini tekshiring.",
        "results_empty": "📭 Hali test yaratmagansiz.",
        "results_choose": "📊 <b>Boshqarish uchun testingizni tanlang:</b>",
        "sub_req_msg": "🚫 <b>Majburiy obuna!</b>\n\nBotdan foydalanish uchun <b>{channel}</b> kanaliga obuna bo'lishingiz majburiy.\n\nKanalga qo'shilgach, «Tasdiqlash» tugmasini bosing.",
        "sub_btn_join": "📢 Kanalga qo'shilish",
        "sub_btn_check": "✅ Tasdiqlash",
        "test_not_found": "❌ Test topilmadi yoki o'chirilgan.",
        "test_running_err": "⏳ Sizda bu test bo'yicha sessiya davom etyapti.",
        "test_limit_err": "❌ Siz bu testni ishlash limitingizni tugatgansiz!",
        "place_1": "🥇 👑 1-O'RIN",
        "place_2": "🥈 🌟 2-O'RIN",
        "place_3": "🥉 ✨ 3-O'RIN",
        "place_n": "🎗 {i}-o'rin",
        "correct_ans": "ta to'g'ri",
        "time_min": "daqiqa",
        "time_sec": "soniya",
        "leaderboard_title": "🏆 <b>TEST YAKUNLANDI!</b> 🏆\n━━━━━━━━━━━━━━━━━━━━━━\n📚 <b>Mavzu:</b> {title}\n📝 <b>Savollar soni:</b> {qcount} ta\n👥 <b>Qatnashchilar:</b> {participants} ta\n━━━━━━━━━━━━━━━━━━━━━━\n\n📊 <b>TOP NATIJALAR:</b>\n\n{lines}\n\n━━━━━━━━━━━━━━━━━━━━━━\n🤖 <i>Geo Ustoz - Bilimingizni sinang!</i>",
        "nobody_played": "🤷‍♂️ <i>Hali hech kim qatnashmadi.</i>",
        "main_menu_loaded": "⚙️ Asosiy menyu yuklandi:"
    },
    "uz_cyrl": {
        "welcome": "👋 <b>Салом, {name}!</b>\n\nGeo Ustoz ботига хуш келибсиз.\n\nПастдаги меню орқали керакли бўлимни танланг 👇",
        "account": "👤 <b>Сизнинг ҳисобингиз:</b>\n\n🆔 ID: <code>{user_id}</code>\n💼 Мақом: <b>{status}</b>{premium_text}\n🌐 Тил: 🇺🇿 Ўзбек (Кирилл)\n\n<i>Premium мақоми орқали барча чекловлардан халос бўласиз!</i>",
        "change_lang_btn": "🌐 Тилни ўзгартириш",
        "choose_lang": "Илтимос, ўзингизга қулай тилни танланг:",
        "lang_saved": "✅ Тил муваффақиятли ўзгартирилди!",
        "btn_cabinet": "🖥 Шахсий кабинет",
        "btn_account": "👤 Ҳисобим",
        "btn_wallet": "👛 Менинг Ҳамёним",
        "btn_referral": "🔗 Дўстларни таклиф қилиш",
        "btn_top": "🏆 Рейтинг ва Ўрним",
        "wallet_info": "👛 <b>GEO WALLET (GWT)</b>\n\n🔑 Ҳамён манзили:\n<code>{address}</code>\n\n💰 Баланс: <b>{balance} GWT</b>\n\n<i>Қуйидаги тугмалар орқали токен ўтказишингиз ёки қабул қилишингиз мумкин.</i>",
        "btn_transfer": "💸 Токен ўтказиш",
        "btn_receive": "📥 Қабул қилиш",
        "btn_ai": "🤖 AI билан гаплашиш",
        "btn_search": "🔍 Қидирув",
        "btn_create_manual": "🧩 қўлда тест яратиш",
        "btn_create_word": "📝 Wordдан тест яратиш",
        "btn_results": "📊 Натижалар / Excel",
        "btn_check_chats": "✅ Чатларни текшириш",
        "btn_premium": "💎 Premium сотиб олиш",
        "btn_add_bot": "➕ Ботни гуруҳ/каналга қўшиш",
        "btn_lock": "🔐 Қулфлаш",
        "btn_back": "⬅️ Орқага",
        "btn_back_main": "⬅️ Асосий меню",
        "btn_cancel": "❌ Бекор қилиш",
        "btn_yes": "✅ Ҳа",
        "btn_no": "❌ Йўқ",
        "btn_solve_site": "🌐 Сайтда ишлаш (Қулай)",
        "btn_solve_bot": "💬 Ботда ишлаш",
        "btn_finish_test_session": "🛑 Тестни тугатиш",
        "btn_export_excel": "📤 Excel юклаб олиш",
        "btn_close_test": "⛔ Тестни ёпиш",
        "btn_publish_test": "🌐 Public қилиш (@ном)",
        "btn_delete_test": "🗑 Тестни ўчириш",
        "btn_pay_card": "💳 Карта (Чек юбориш)",
        "btn_pay_stars": "⭐️ Telegram Stars (Автоматик)",
        "btn_pay_token": "🪙 GWT Токен (Блокчейн)",
        "btn_create_private": "🔒 Private тест (фақат ўзим учун)",
        "month_1": "1 Ой", "month_3": "3 Ой", "month_6": "6 Ой", "month_12": "1 Йил",
        "captcha_msg": "🔒 <b>Хавфсизлик текшируви!</b>\n\nGeo Ustoz ботидан фойдаланиш учун, илтимос, пастдаги тугмани босиб инсон эканлигингизни тасдиқланг.",
        "test_manage_msg": "⚙️ <b>Тест бошқаруви:</b>\n{title}\n\nНима амал бажарамиз?",
        "test_options_msg": "🧩 <b>Тест: {title}</b>\n⏱ <b>Вақт:</b> {time}\n📊 <b>Баҳолаш:</b> {scoring}\n🔄 <b>Уринишлар лимити:</b> {limit}\n\nТестни қаерда ишлашни хоҳлайсиз?",
        "prem_method_msg": "💎 <b>Premium харид қилиш усулини танланг:</b>\n\n💳 <b>Карта орқали</b> тўласангиз, чекни админга юборасиз ва тасдиқлашини кутасиз.\n⭐️ <b>Telegram Юлдузлари</b> орқали тўласангиз, Premium <b>автоматик ва шу заҳоти</b> берилади!",
        "prem_card_msg": "💳 <b>Карта орқали тўлов таърифлари:</b>\n<i>Карта орқали тўласангиз тўлов чекини юборишингиз керак бўлади.</i>",
        "prem_stars_msg": "⭐️ <b>Telegram Stars таърифлари:</b>\n<i>Premium автоматик тарзда шу заҳоти берилади!</i>",
        "prem_card_info": "💳 <b>Тўлов учун маълумотлар:</b>\n\n💰 Сумма: <b>{price} сўм</b>\n💳 Карта: <code>{card}</code>\n\n👇 Тўловни амалга оширгач, <b>ЧЕК (скриншот) расмини шу ерга юборинг!</b>",
        "lock_system": "🔐 <b>Тизим қулфланган!</b>\n\nДавом этиш учун PIN кодингизни киритинг:",
        "unlock_success": "🔓 <b>Қулф очилди!</b>",
        "wrong_pin": "❌ <b>Нотўғри PIN-код! Қайтадан уриниб кўринг.</b>",
        "no_pin_setup": "⚠️ Сизда ҳали PIN код ўрнатилмаган! Қулфлаш учун аввал <b>Шахсий кабинет</b> (сайт) орқали PIN код ўрнатинг.",
        "ai_welcome": "🤖 <b>Geo AI</b>га хуш келибсиз! Саволингизни ёзинг ёки расм юборинг.\n\n<i>(Асосий менюга қайтиш учун /start ни босинг ёки [Менюга қайтиш] деб ёзинг)</i>",
        "ai_thinking": "<i>AI жавоб ёзмоқда... ⏳</i>",
        "ai_pic_thinking": "<i>AI расмни таҳлил қилмоқда... ⏳</i>",
        "ai_limit": "❌ Бугунги AI хабарлар лимитингиз (10 та) тугади.\nАсосий менюга қайтарилдингиз.",
        "create_where": "📝 Қайси канал/гуруҳга тест юборасиз?",
        "convert_where": "📝 <b>Word дан тест яратиш</b>\n\nҚайси жойга юборилсин? Аввал шуни танланг, кейин сиздан Word файли сўралади.",
        "search_prompt": "🔍 @testname орқали қидирув. @testname киритинг:",
        "search_results": "🔍 Қидирув натижалари:",
        "search_not_found": "🤷‍♂️ Ҳеч нарса топилмади. Номи тўғрилигини текширинг.",
        "results_empty": "📭 Ҳали тест яратмагансиз.",
        "results_choose": "📊 <b>Бошқариш учун тестингизни танланг:</b>",
        "sub_req_msg": "🚫 <b>Мажбурий обуна!</b>\n\nБотдан фойдаланиш учун <b>{channel}</b> каналига обуна бўлишингиз мажбурий.\n\nКаналга қўшилгач, «Тасдиқлаш» тугмасини босинг.",
        "sub_btn_join": "📢 Каналга қўшилиш",
        "sub_btn_check": "✅ Тасдиқлаш",
        "test_not_found": "❌ Тест топилмади ёки ўчирилган.",
        "test_running_err": "⏳ Сизда бу тест бўйича сессия давом этяпти.",
        "test_limit_err": "❌ Сиз бу тестни ишлаш лимитингизни тугатгансиз!",
        "place_1": "🥇 👑 1-ЎРИН",
        "place_2": "🥈 🌟 2-ЎРИН",
        "place_3": "🥉 ✨ 3-ЎРИН",
        "place_n": "🎗 {i}-ўрин",
        "correct_ans": "та тўғри",
        "time_min": "дақиқа",
        "time_sec": "сония",
        "leaderboard_title": "🏆 <b>ТЕСТ ЯКУНЛАНДИ!</b> 🏆\n━━━━━━━━━━━━━━━━━━━━━━\n📚 <b>Мавзу:</b> {title}\n📝 <b>Саволлар сони:</b> {qcount} та\n👥 <b>Қатнашчилар:</b> {participants} та\n━━━━━━━━━━━━━━━━━━━━━━\n\n📊 <b>ТОП НАТИЖАЛАР:</b>\n\n{lines}\n\n━━━━━━━━━━━━━━━━━━━━━━\n🤖 <i>Geo Ustoz - Билимингизни синанг!</i>",
        "nobody_played": "🤷‍♂️ <i>Ҳали ҳеч ким қатнашмади.</i>",
        "main_menu_loaded": "⚙️ Асосий меню юкланди:"
    },
    "ru": {
        "welcome": "👋 <b>Привет, {name}!</b>\n\nДобро пожаловать в бот Geo Ustoz.\n\nВыберите нужный раздел в меню ниже 👇",
        "account": "👤 <b>Ваш профиль:</b>\n\n🆔 ID: <code>{user_id}</code>\n💼 Статус: <b>{status}</b>{premium_text}\n🌐 Язык: 🇷🇺 Русский\n\n<i>Статус Premium снимает все ограничения!</i>",
        "change_lang_btn": "🌐 Изменить язык",
        "choose_lang": "Пожалуйста, выберите удобный язык:",
        "lang_saved": "✅ Язык успешно изменен!",
        "btn_cabinet": "🖥 Личный кабинет",
        "btn_account": "👤 Мой профиль",
        "btn_wallet": "👛 Мой Кошелек",
        "btn_referral": "🔗 Пригласить друзей",
        "btn_top": "🏆 Рейтинг и Мое место",
        "wallet_info": "👛 <b>GEO WALLET (GWT)</b>\n\n🔑 Адрес кошелька:\n<code>{address}</code>\n\n💰 Баланс: <b>{balance} GWT</b>\n\n<i>Ниже вы можете отправить или получить токены.</i>",
        "btn_transfer": "💸 Перевести токены",
        "btn_receive": "📥 Получить",
        "btn_ai": "🤖 Общение с ИИ",
        "btn_search": "🔍 Поиск",
        "btn_create_manual": "🧩 Создать тест вручную",
        "btn_create_word": "📝 Тест из Word",
        "btn_results": "📊 Результаты / Excel",
        "btn_check_chats": "✅ Проверить чаты",
        "btn_premium": "💎 Купить Premium",
        "btn_add_bot": "➕ Добавить бота в группу/канал",
        "btn_lock": "🔐 Заблокировать",
        "btn_back": "⬅️ Назад",
        "btn_back_main": "⬅️ Главное меню",
        "btn_cancel": "❌ Отмена",
        "btn_yes": "✅ Да",
        "btn_no": "❌ Нет",
        "btn_solve_site": "🌐 Решать на сайте (Удобно)",
        "btn_solve_bot": "💬 Решать в боте",
        "btn_finish_test_session": "🛑 Завершить тест",
        "btn_export_excel": "📤 Скачать Excel",
        "btn_close_test": "⛔ Закрыть тест",
        "btn_publish_test": "🌐 Сделать публичным (@имя)",
        "btn_delete_test": "🗑 Удалить тест",
        "btn_pay_card": "💳 Картой (Отправить чек)",
        "btn_pay_stars": "⭐️ Telegram Stars (Автоматически)",
        "btn_pay_token": "🪙 GWT Токен (Блокчейн)",
        "btn_create_private": "🔒 Private тест (только для себя)",
        "month_1": "1 Месяц", "month_3": "3 Месяца", "month_6": "6 Месяцев", "month_12": "1 Год",
        "captcha_msg": "🔒 <b>Проверка безопасности!</b>\n\nДля использования бота Geo Ustoz, пожалуйста, нажмите кнопку ниже, чтобы подтвердить, что вы человек.",
        "test_manage_msg": "⚙️ <b>Управление тестом:</b>\n{title}\n\nЧто будем делать?",
        "test_options_msg": "🧩 <b>Тест: {title}</b>\n⏱ <b>Время:</b> {time}\n📊 <b>Оценивание:</b> {scoring}\n🔄 <b>Лимит попыток:</b> {limit}\n\nГде вы хотите решить тест?",
        "prem_method_msg": "💎 <b>Выберите способ покупки Premium:</b>\n\n💳 <b>Картой</b>: отправьте чек админу и ждите подтверждения.\n⭐️ <b>Telegram Stars</b>: Premium выдается <b>автоматически и мгновенно</b>!",
        "prem_card_msg": "💳 <b>Тарифы при оплате картой:</b>\n<i>При оплате картой вам нужно будет отправить чек.</i>",
        "prem_stars_msg": "⭐️ <b>Тарифы Telegram Stars:</b>\n<i>Premium выдается автоматически мгновенно!</i>",
        "prem_card_info": "💳 <b>Данные для оплаты:</b>\n\n💰 Сумма: <b>{price} сум</b>\n💳 Карта: <code>{card}</code>\n\n👇 После оплаты <b>отправьте фото ЧЕКА сюда!</b>",
        "lock_system": "🔐 <b>Система заблокирована!</b>\n\nВведите ваш PIN-код для продолжения:",
        "unlock_success": "🔓 <b>Разблокировано!</b>",
        "wrong_pin": "❌ <b>Неверный PIN-код! Попробуйте еще раз.</b>",
        "no_pin_setup": "⚠️ У вас еще не установлен PIN-код! Сначала установите его через <b>Личный кабинет</b> (сайт).",
        "ai_welcome": "🤖 Добро пожаловать в <b>Geo AI</b>! Напишите вопрос или отправьте фото.\n\n<i>(Для возврата в меню нажмите /start или напишите [В главное меню])</i>",
        "ai_thinking": "<i>ИИ пишет ответ... ⏳</i>",
        "ai_pic_thinking": "<i>ИИ анализирует фото... ⏳</i>",
        "ai_limit": "❌ Ваш сегодняшний лимит сообщений ИИ (10 шт) исчерпан.\nВы возвращены в главное меню.",
        "create_where": "📝 В какой канал/группу отправить тест?",
        "convert_where": "📝 <b>Создание теста из Word</b>\n\nКуда отправить? Сначала выберите это, затем вас попросят отправить файл.",
        "search_prompt": "🔍 Поиск по @testname. Введите @testname:",
        "search_results": "🔍 Результаты поиска:",
        "search_not_found": "🤷‍♂️ Ничего не найдено. Проверьте правильность имени.",
        "results_empty": "📭 Вы еще не создали тесты.",
        "results_choose": "📊 <b>Выберите тест для управления:</b>",
        "sub_req_msg": "🚫 <b>Обязательная подписка!</b>\n\nДля использования бота необходимо подписаться на канал <b>{channel}</b>.\n\nПосле подписки нажмите кнопку «Подтвердить».",
        "sub_btn_join": "📢 Подписаться на канал",
        "sub_btn_check": "✅ Подтвердить",
        "test_not_found": "❌ Тест не найден или удален.",
        "test_running_err": "⏳ Sizda bu test bo'yicha sessiya davom etyapti. Avval testni tugating.",
        "test_limit_err": "❌ Вы исчерпали свой лимит попыток для этого теста!",
        "place_1": "🥇 👑 1-Е МЕСТО",
        "place_2": "🥈 🌟 2-Е МЕСТО",
        "place_3": "🥉 ✨ 3-Е МЕСТО",
        "place_n": "🎗 {i}-е место",
        "correct_ans": "верно",
        "time_min": "минут",
        "time_sec": "секунд",
        "leaderboard_title": "🏆 <b>ТЕСТ ЗАВЕРШЕН!</b> 🏆\n━━━━━━━━━━━━━━━━━━━━━━\n📚 <b>Тема:</b> {title}\n📝 <b>Количество вопросов:</b> {qcount} шт\n👥 <b>Участников:</b> {participants} чел\n━━━━━━━━━━━━━━━━━━━━━━\n\n📊 <b>ТОП РЕЗУЛЬТАТЫ:</b>\n\n{lines}\n\n━━━━━━━━━━━━━━━━━━━━━━\n🤖 <i>Geo Ustoz - Проверьте свои знания!</i>",
        "nobody_played": "🤷‍♂️ <i>Пока никто не участвовал.</i>",
        "main_menu_loaded": "⚙️ Главное меню загружено:"
    }
}

def get_bot_text(key, lang="uz", **kwargs):
    text = BOT_LANGS.get(lang, BOT_LANGS["uz"]).get(key, BOT_LANGS["uz"].get(key, key))
    if kwargs:
        return text.format(**kwargs)
    return text

def get_user_lang(user_id: int) -> str:
    try:
        with db._conn() as c:
            row = c.execute("SELECT lang FROM users WHERE user_id=%s", (user_id,)).fetchone()
            if row and "lang" in row and row["lang"]:
                return row["lang"]
    except Exception:
        pass
    return "uz"

def get_all_localized_buttons(key: str) -> list:
    return [BOT_LANGS[l].get(key, "") for l in BOT_LANGS]

# ==========================================
# 3. MATNLAR VA YORDAMCHI FUNKSIYALAR
# ==========================================

def format_display_score(score_val, scoring_type, lang="uz"):
    """ Natijani baholash tizimiga moslashtirib chiroyli formatlaydi """
    score_val = float(score_val or 0)
    if scoring_type == "percentage":
        return f"{score_val:g} %"
    elif scoring_type in ["minus", "custom"]:
        b = "ball" if lang != "ru" else "балл"
        return f"{score_val:g} {b}"
    else:
        t = get_bot_text('correct_ans', lang)
        return f"{int(score_val)} {t}"


class texts:
    @staticmethod
    def medal_line_html(i: int, name: str, score_str: str, dur: str, lang: str = "uz") -> str:
        if i == 1:
            medal = get_bot_text('place_1', lang)
        elif i == 2:
            medal = get_bot_text('place_2', lang)
        elif i == 3:
            medal = get_bot_text('place_3', lang)
        else:
            medal = get_bot_text('place_n', lang, i=i)

        return f"{medal} ➔ 👤 <b>{html.escape(name)}</b> (🎯 {score_str} | ⏱ {dur})"

    @staticmethod
    def leaderboard_template_html(title: str, qcount: int, per_q_sec: int, participants: int, lines: list, lang: str = "uz") -> str:
        lines_str = "\n\n".join(lines) if lines else get_bot_text('nobody_played', lang)
        return get_bot_text('leaderboard_title', lang, title=html.escape(title), qcount=qcount, participants=participants, lines=lines_str)


def now_ts() -> int:
    return int(datetime.now(tz=timezone.utc).timestamp())

def ts_to_local(ts: int) -> str:
    return datetime.fromtimestamp(ts, tz=timezone.utc).astimezone(TZ).strftime("%d/%m/%Y %H:%M")

def parse_deadline(s: str) -> int:
    dt = datetime.strptime(s.strip(), "%d/%m/%Y %H:%M")
    dt = dt.replace(tzinfo=TZ)
    return int(dt.astimezone(timezone.utc).timestamp())

def is_admin_member(m) -> bool:
    return m.status in (ChatMemberStatus.ADMINISTRATOR, ChatMemberStatus.OWNER)

def fmt_duration(sec: int, lang: str = "uz") -> str:
    m = sec // 60
    s = sec % 60
    min_txt = get_bot_text('time_min', lang)
    sec_txt = get_bot_text('time_sec', lang)
    if m <= 0: return f"{s} {sec_txt}"
    return f"{m} {min_txt} {s} {sec_txt}"

def h(s: object) -> str:
    return html.escape(str(s) if s is not None else "")

def format_user_display(username: Optional[str], first_name: Optional[str], last_name: Optional[str], user_id: int) -> str:
    uname = (username or "").strip().lstrip("@")
    fn = (first_name or "").strip()
    ln = (last_name or "").strip()
    full = (f"{fn} {ln}").strip()
    if uname and full: return f"@{uname} ({full})"
    if uname: return f"@{uname}"
    if full: return full
    return f"User{user_id}"

async def upsert_user_from_update(update: Update):
    u = update.effective_user
    if not u: return
    try:
        db.upsert_user(
            user_id=u.id,
            username=u.username or "",
            first_name=u.first_name or "",
            last_name=u.last_name or "",
            now_ts=now_ts()
        )
        db.get_or_create_user_api_key(u.id)
    except Exception as e:
        logging.error(f"Foydalanuvchini bazaga qo'shishda xato: {e}")

async def send_test_options(update: Update, context: ContextTypes.DEFAULT_TYPE, test_id: str):
    user_id = update.effective_user.id
    chat_id = update.effective_chat.id
    lang = get_user_lang(user_id)

    test = db.get_test(test_id)
    if not test:
        await context.bot.send_message(chat_id=chat_id, text=get_bot_text('test_not_found', lang))
        return

    # ----------------- PULLIK TEST TEKSHIRUVI -----------------
    has_price = float(dict(test).get("price_gwt") or 0) > 0 or int(dict(test).get("price_stars") or 0) > 0
    is_owner = int(dict(test).get("owner_user_id", 0)) == user_id
    is_admin = user_id in SUPERADMINS

    if has_price and not is_owner and not is_admin:
        with db._conn() as c:
            purchased = c.execute("SELECT 1 FROM purchased_tests WHERE user_id=%s AND test_id=%s", (user_id, test_id)).fetchone()

        if not purchased:
            token = db.get_or_create_user_api_key(user_id)
            webapp_url = f"{WEB_BASE_URL.rstrip('/')}/solve/{test_id}?token={token}"
            kb = InlineKeyboardMarkup([[InlineKeyboardButton("🛒 Sotib olish (Saytda)", web_app=WebAppInfo(url=webapp_url))]])
            await context.bot.send_message(chat_id=chat_id, text="⛔ <b>Bu pullik test!</b>\n\nUni ishlash uchun avval sotib olishingiz kerak:", reply_markup=kb, parse_mode=ParseMode.HTML)
            return
    # -----------------------------------------------------------

    with db._conn() as c:
        is_running = c.execute("SELECT 1 FROM sessions WHERE user_id=%s AND test_id=%s AND state='running' LIMIT 1", (user_id, test_id)).fetchone()

    if is_running:
        kb_finish = InlineKeyboardMarkup([[InlineKeyboardButton("🛑 Sessiyani tugatish", callback_data=f"force_finish_{test_id}")]])
        await context.bot.send_message(chat_id=chat_id, text=get_bot_text('test_running_err', lang), reply_markup=kb_finish)
        return

    # Yordamchi matnlarni tayyorlash (Pro ma'lumotlari)
    attempts_limit = test.get("attempts_limit")
    if attempts_limit is None: attempts_limit = 1

    if lang == 'ru':
        limit_txt = "Безлимитно" if attempts_limit == 0 else f"{attempts_limit} раз(а)"
    elif lang == 'uz_cyrl':
        limit_txt = "Чексиз" if attempts_limit == 0 else f"{attempts_limit} марта"
    else:
        limit_txt = "Cheksiz" if attempts_limit == 0 else f"{attempts_limit} marta"

    time_limit = int(test.get("time_limit") or 0)
    if time_limit > 0:
        time_txt = f"{time_limit} {get_bot_text('time_min', lang)}"
    else:
        if lang == 'ru': time_txt = "Безлимитно"
        elif lang == 'uz_cyrl': time_txt = "Чексиз"
        else: time_txt = "Cheksiz"

    scoring = test.get("scoring_type", "standard")
    if scoring == "percentage":
        scor_txt = "Foizli (100%)" if lang != 'ru' else "В процентах (100%)"
    elif scoring == "minus":
        scor_txt = "Minus ballik" if lang != 'ru' else "Штрафные баллы"
    elif scoring == "custom":
        scor_txt = "Maxsus ballar" if lang != 'ru' else "Специальные баллы"
    else:
        scor_txt = "Standard" if lang != 'ru' else "Стандарт"

    with db._conn() as c:
        finished_count_row = c.execute("SELECT COUNT(*) as cnt FROM sessions WHERE test_id=%s AND user_id=%s AND state='finished'", (test_id, user_id)).fetchone()
        finished_count = finished_count_row["cnt"] if finished_count_row else 0

    if attempts_limit > 0 and finished_count >= attempts_limit:
        await context.bot.send_message(chat_id=chat_id, text=get_bot_text('test_limit_err', lang))
        return

    webapp_url = f"{WEB_BASE_URL.rstrip('/')}/app-solve/{test_id}"

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton(get_bot_text('btn_solve_site', lang), web_app=WebAppInfo(url=webapp_url), style="success")],
        [InlineKeyboardButton(get_bot_text('btn_solve_bot', lang), callback_data=f"bot_solve:{test_id}", style="primary")]
    ])

    text_msg = get_bot_text('test_options_msg', lang, title=h(test.get('title', '')), limit=limit_txt, time=time_txt, scoring=scor_txt)

    await context.bot.send_message(
        chat_id=chat_id,
        text=text_msg,
        reply_markup=kb,
        parse_mode=ParseMode.HTML
    )

# ==========================================
# 4. AI VA INTERNET FUNKSIYALARI
# ==========================================
def search_internet(query: str) -> str:
    try:
        results = DDGS().text(query, max_results=3)
        res_list = []
        for r in results:
            res_list.append(r)
            if len(res_list) >= 3:
                break

        if not res_list: return ""

        context = "INTERNETDAGI ENG SO'NGGI MA'LUMOTLAR:\n"
        for i, res in enumerate(res_list):
            context += f"{i+1}. {res.get('title', '')}: {res.get('body', '')}\n"
        return context
    except Exception as e:
        logging.error(f"Internet qidiruv xatosi: {e}")
        return ""

async def solve_test_with_ai(questions_text: str) -> str:
    if not GROQ_API_KEY: return ""
    sys_prompt = "Siz test yechuvchi AIsiz. Quyidagi savollarning to'g'ri javoblarini toping va faqat qat'iy formatda qaytaring: 1A 2B 3C ... Ortig'ini yozmang."
    def fetch():
        headers = {"Authorization": f"Bearer {GROQ_API_KEY}", "Content-Type": "application/json"}
        data = {
            "model": GROQ_MODEL,
            "messages": [
                {"role": "system", "content": sys_prompt},
                {"role": "user", "content": questions_text[:15000]}
            ],
            "temperature": 0.1
        }
        res = requests.post(GROQ_URL, headers=headers, json=data, timeout=60)
        res.raise_for_status()
        return res.json()["choices"][0]["message"]["content"]

    try:
        return await asyncio.to_thread(fetch)
    except Exception as e:
        logging.error(f"AI Solving Error: {e}")
        return ""

async def get_theme_from_ai(questions_text: str) -> str:
    if not GROQ_API_KEY: return "Umumiy Test"
    sys_prompt = "Siz test mavzusi o'ylab topuvchisiz. Berilgan test savollariga qarab, eng mos, qisqa va lo'nda mavzu (sarlavha) yozib bering. Faqatgina mavzu nomini yozing, boshqa ortiqcha gap yozmang."

    def fetch():
        headers = {"Authorization": f"Bearer {GROQ_API_KEY}", "Content-Type": "application/json"}
        data = {
            "model": GROQ_MODEL,
            "messages": [
                {"role": "system", "content": sys_prompt},
                {"role": "user", "content": questions_text[:5000]}
            ],
            "temperature": 0.3
        }
        res = requests.post(GROQ_URL, headers=headers, json=data, timeout=30)
        res.raise_for_status()
        return res.json()["choices"][0]["message"]["content"].strip()

    try:
        theme = await asyncio.to_thread(fetch)
        theme = theme.replace('"', '').replace("'", "").replace("*", "")
        return theme if theme else "Umumiy Test"
    except Exception as e:
        logging.error(f"AI Theme Error: {e}")
        return "Umumiy Test"

async def announce_to_channel_with_ai(context: ContextTypes.DEFAULT_TYPE, theme: str, qcount: int, public_name: str):
    if not GROQ_API_KEY:
        return

    sys_prompt = "Siz ta'lim va testlar kanalida professional SMM mutaxassisisiz. Yangi test yaratildi. Odamlarni bu testni ishlashga qiziqtiruvchi qisqa, emojilarga boy, energiya bilan to'la post yozing. Matnda faqat oddiy Telegram HTML (<b>qalin</b>, <i>yotiq</i>) ishlating."
    user_prompt = (
        f"Test mavzusi: {theme}\n"
        f"Savollar soni: {qcount} ta\n"
        f"Test kodi: {public_name}\n\n"
        f"Foydalanuvchilarga ushbu testni ishlash uchun botga kirib aynan {public_name} kodini yozishlari "
        f"yoki ustiga bosishlari kerakligini chiroyli tushuntiring."
    )

    def fetch():
        headers = {"Authorization": f"Bearer {GROQ_API_KEY}", "Content-Type": "application/json"}
        data = {
            "model": GROQ_MODEL,
            "messages": [
                {"role": "system", "content": sys_prompt},
                {"role": "user", "content": user_prompt}
            ],
            "temperature": 0.7
        }
        res = requests.post(GROQ_URL, headers=headers, json=data, timeout=30)
        res.raise_for_status()
        return res.json()["choices"][0]["message"]["content"].strip()

    try:
        post_text = await asyncio.to_thread(fetch)
        bot_username = context.bot.username
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("🚀 Botga o'tish va Testni ishlash", url=f"https://t.me/{bot_username}")]
        ])

        await context.bot.send_message(
            chat_id=PUBLIC_TEST_CHANNEL,
            text=post_text,
            reply_markup=kb,
            parse_mode=ParseMode.HTML
        )
    except Exception as e:
        logging.error(f"Kanalga avtomatik e'lon yuborishda xato: {e}")

async def get_ai_reply(text: str, user_info_str: str, user_id: int, history: list, base64_image: str = None) -> str:
    if not GROQ_API_KEY:
        return "❌ Kechirasiz, AI tizimi hozircha sozlanmagan. Iltimos adminga murojaat qiling."

    sys_prompt = f"""Siz 'Geo Ustoz' Telegram botining aqlli AI yordamchisisiz. Ismingiz: Geo AI. Seni Ulug'bek G'ulomov yasagan.
Asosiy qoida: Javoblaringizni doim faqat ODDIY MATN yoki eng oddiy telegram HTML formatida bering (qalin uchun <b>...</b>, yotiq uchun <i>...</i>). Markdown yulduzchalar va HTML ro'yxatlar ishlatmang! Ro'yxat kerak bo'lsa faqat oddiy nuqta (•) ishlating.
Senda JONLI INTERNETGA ULANISH imkoniyati bor. Agar tizim internet natijalarini bersa, foydalan.
Adminning telegramdagi usernamesi @python_dasturchi_bola buni so'ralsagina ayt.

FOYDALANUVCHINING MA'LUMOTLARI VA GURUHLARI:
{user_info_str}
(Foydalanuvchining o'zining shaxsiy ID raqami: {user_id})

🔥 MAXSUS BUYRUQLAR (Sening pulting):
Agar foydalanuvchi quyidagi amallarni so'rasa, gap orasida yoki oxirida faqat shu buyruqlarni yozing:
- Qidiruv qilish (masalan, kimyo testini top): [CMD:SEARCH:@kimyo]
- Foydalanuvchi o'z natijalarini ko'rmoqchi bo'lsa: [CMD:RESULTS]
- Premium sotib olmoqchi bo'lsa: [CMD:PREMIUM]
- Botni qulflamoqchi bo'lsa: [CMD:LOCK]
- Menyuga qaytmoqchi bo'lsa: [CMD:MAIN_MENU]
- Reklama tarqatish (faqat adminlar uchun): [CMD:BROADCAST]
- Worddan test yaratish (eski usul): [CMD:WORD_MODE:chat_id_raqami]

🛠 TEST YARATISH (YANGI JSON TIZIM):
Agar foydalanuvchi test yaratmoqchi bo'lsa, u bilan suhbatlashib ma'lumotlarni yig'ib ol. Agar foydalanuvchi "O'zing tuz", "Hammasini o'zing kirit" yoki shunga o'xshash mustaqil ishlashni talab qilsa, undan hech narsa so'ramasdan darhol o'zing 5 ta (yoki so'ralgan miqdorda) savolli test o'ylab top.
QAT'IY QOIDA: Testni yaratgach, xabarning eng oxirida ALBATTA quyidagi JSON formatni to'liq va xatosiz yoz. JSON kodni doim "===JSON_START===" va "===JSON_END===" belgilari orasiga yozish SHART!
DIQQAT: JSON atrofida ``` (backtick) belgilarini UMUMAN ishlatma, faqat toza va to'liq JSON yoz:
===JSON_START===
{{"title": "Mavzu", "questions": [{{"q": "Savol 1?", "opts": ["A", "B", "C", "D"], "ans": 0}}]}}
===JSON_END===
"""

    if not base64_image and text:
        internet_context = await asyncio.to_thread(search_internet, text)
        if internet_context:
            sys_prompt += f"\n\n{internet_context}\nYuqoridagi internet ma'lumotlaridan foydalanib javob ber!"

    def fetch_groq():
        headers = {
            "Authorization": f"Bearer {GROQ_API_KEY}",
            "Content-Type": "application/json"
        }

        messages = [{"role": "system", "content": sys_prompt}]
        for msg in history[:-1]:
            messages.append(msg)

        last_msg = history[-1]

        if base64_image:
            messages.append({
                "role": "user",
                "content": [
                    {"type": "text", "text": last_msg["content"]},
                    {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{base64_image}"}}
                ]
            })
            current_model = VISION_MODEL
        else:
            messages.append(last_msg)
            current_model = GROQ_MODEL

        data = {
            "model": current_model,
            "messages": messages,
            "temperature": 0.7
        }

        try:
            response = requests.post(GROQ_URL, headers=headers, json=data, timeout=30)
            response.raise_for_status()
            return response.json()["choices"][0]["message"]["content"]
        except requests.exceptions.RequestException as req_err:
            err_text = req_err.response.text if req_err.response else str(req_err)
            logging.error(f"Groq API XATOSI: {err_text}")
            return f"⚠️ Serverda xatolik yuz berdi. Iltimos biroz kuting."

    try:
        reply = await asyncio.to_thread(fetch_groq)
        return reply
    except Exception as e:
        logging.error(f"AI Xatosi: {e}")
        return "⚠️ Tizimda xatolik yuz berdi."

async def process_ai_message(update: Update, context: ContextTypes.DEFAULT_TYPE, user_id: int, text: str, wait_msg, base64_image: str = None):
    user_row = dict(db.get_user(user_id) or {})
    is_premium = user_row.get("status") == "premium"
    is_admin = user_id in SUPERADMINS or user_id in LOWER_ADMINS
    lang = get_user_lang(user_id)

    chats = db.chats_for_user(user_id)
    eligible_chats = [c for c in chats if c.get("bot_is_admin") == 1]
    user_chats_info = "\n".join([f"- Nomi: {c.get('title')}, ID: {c.get('chat_id')}" for c in eligible_chats]) if eligible_chats else "Guruhlar yo'q."

    user_info_str = f"""
- Premium maqomi: {"Bor (Cheklovsiz)" if is_premium else "Yo'q (Limitli)"}
- Adminlik huquqi: {"Bor" if is_admin else "Yo'q"}
- Guruhlari:\n{user_chats_info}
"""

    history = context.user_data.get("ai_history", [])
    history.append({"role": "user", "content": text})

    if len(history) > 20:
        history = history[-20:]

    reply = await get_ai_reply(text, user_info_str, user_id, history, base64_image)

    json_match = re.search(r'===JSON_START===\s*(.*?)\s*===JSON_END===', reply, re.DOTALL)
    if json_match:
        test_data_str = json_match.group(1).strip()
        reply = reply.replace(json_match.group(0), "").strip()

        test_data_str = re.sub(r'^```json\s*', '', test_data_str, flags=re.IGNORECASE)
        test_data_str = re.sub(r'^```\s*', '', test_data_str)
        test_data_str = re.sub(r'\s*```$', '', test_data_str)
        test_data_str = test_data_str.strip()

        try:
            test_data = json.loads(test_data_str)
            test_id = uuid.uuid4().hex[:10]
            db.create_test(
                test_id=test_id,
                owner_user_id=user_id,
                chat_id=user_id,
                title=test_data.get("title", "AI Test"),
                per_question_sec=60,
                created_at=now_ts(),
                scoring_type="standard",
                time_limit=0,
                is_randomized=0
            )

            for i, q in enumerate(test_data.get("questions", [])):
                db.add_question(
                    test_id=test_id,
                    q_index=i,
                    question=q.get("q", "Savol?"),
                    options_list=q.get("opts", ["A", "B"]),
                    correct_index=q.get("ans", 0),
                    photo_id=None,
                    score=1.0
                )
            reply += f"\n\n✅ AI testingizni muvaffaqiyatli yaratdi! Boshlash uchun bosing: /start test_{test_id}"

        except json.JSONDecodeError as e:
            logging.error(f"AI JSON Parse xatosi: {e}\nKelgan JSON: {test_data_str}")
            reply += "\n\n❌ AI test savollarini shakllantirishda kichik xatoga yo'l qo'ydi. Unga 'Kodni tekshirib qaytadan JSON yubor' deb yozing."
        except Exception as e:
            logging.error(f"AI DB yozish xatosi: {e}")
            reply += "\n\n❌ Testni bazaga saqlashda xatolik yuz berdi."

    cmd_matched = None
    cmd_arg = None
    cmd_pattern = r"\[CMD:([A-Z_]+)(?::([^\]]+))?\]"
    match = re.search(cmd_pattern, reply)
    if match:
        cmd_matched = match.group(1)
        cmd_arg = match.group(2)
        reply = re.sub(cmd_pattern, "", reply).strip()

    reply = re.sub(r'<think>.*?</think>', '', reply, flags=re.DOTALL).strip()
    reply = reply.replace("<think>", "").replace("</think>", "")

    history.append({"role": "assistant", "content": reply})
    context.user_data["ai_history"] = history

    reply_formatted = re.sub(r'\*\*(.*?)\*\*', r'<b>\1</b>', reply)
    reply_formatted = re.sub(r'(?<!\*)\*(?!\*)(.*?)(?<!\*)\*(?!\*)', r'<i>\1</i>', reply_formatted)

    if reply_formatted:
        try:
            await wait_msg.edit_text(reply_formatted, parse_mode=ParseMode.HTML)
        except Exception as e:
            logging.error(f"Telegram HTML xatosi: {e}. Xabar oddiy matn sifatida yuborilmoqda.")
            await wait_msg.edit_text(reply, parse_mode=None)
    else:
        await wait_msg.delete()

    if cmd_matched:
        if cmd_matched == "MAIN_MENU":
            context.user_data.pop(K["mode"], None)
            msg, kbm = await build_main_menu(user_id, context.bot.username, lang)
            await update.effective_chat.send_message(msg, reply_markup=kbm, parse_mode=ParseMode.HTML)

        elif cmd_matched == "SEARCH":
            search_query = cmd_arg.strip().lower() if cmd_arg else ""
            if not search_query.startswith("@"): search_query = "@" + search_query
            results = db.search_public_tests(search_query)
            if results:
                kb = []
                for t in results[:10]:
                    t_title = str(dict(t).get('title', 'Nomsiz Test'))
                    t_name = str(dict(t).get('public_name', ''))
                    t_id = str(dict(t).get('test_id', ''))
                    kb.append([InlineKeyboardButton(f"🧩 {t_title[:30]} ({t_name})", callback_data=f"start_public_test:{t_id}")])
                await update.effective_chat.send_message(get_bot_text('search_results', lang), reply_markup=InlineKeyboardMarkup(kb))
            else:
                await update.effective_chat.send_message(get_bot_text('search_not_found', lang))

        elif cmd_matched == "RESULTS":
            context.user_data.pop(K["mode"], None)
            await results_menu(update, context)

        elif cmd_matched == "PREMIUM":
            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton(get_bot_text('btn_pay_card', lang), callback_data="pay_method_card", style="primary")],
                [InlineKeyboardButton(get_bot_text('btn_pay_stars', lang), callback_data="pay_method_stars", style="success")]
            ])
            await update.effective_chat.send_message(get_bot_text('prem_method_msg', lang), reply_markup=kb, parse_mode=ParseMode.HTML)

        elif cmd_matched == "LOCK":
            context.user_data['is_locked'] = True
            text_msg, kb = get_pin_keyboard("", lang)
            await update.effective_chat.send_message(text_msg, reply_markup=kb, parse_mode=ParseMode.HTML)

        elif cmd_matched == "BROADCAST" and (user_id in SUPERADMINS or user_id in LOWER_ADMINS):
            context.user_data[K["mode"]] = "broadcast"
            context.user_data[K["broadcast"]] = {"stage": "wait_content"}
            kb = InlineKeyboardMarkup([[InlineKeyboardButton(get_bot_text('btn_cancel', lang), callback_data="cancel_action", style="danger")]])
            await update.effective_chat.send_message("📣 Tarqatish uchun Xabar, Rasm, Video yoki Ovozli xabarni yuboring:", reply_markup=kb)

        elif cmd_matched == "WORD_MODE":
            chat_id = int(cmd_arg) if (cmd_arg and cmd_arg.isdigit()) else user_id
            context.user_data[K["mode"]] = "convert"
            context.user_data[K["convert"]] = {"chat_id": chat_id}
            await update.effective_chat.send_message("📝 Test uchun Word faylini yuboring.")

# ==========================================
# 5. PIN KOD VA QULF MIDDLEWARE
# ==========================================
def get_pin_keyboard(current_input: str, lang: str = "uz"):
    display = " ".join(["🟢" if i < len(current_input) else "⚪️" for i in range(4)])
    text = f"{get_bot_text('lock_system', lang)}\n\n{display}"

    kb = [
        [InlineKeyboardButton("1️⃣", callback_data="pin_num_1"), InlineKeyboardButton("2️⃣", callback_data="pin_num_2"), InlineKeyboardButton("3️⃣", callback_data="pin_num_3")],
        [InlineKeyboardButton("4️⃣", callback_data="pin_num_4"), InlineKeyboardButton("5️⃣", callback_data="pin_num_5"), InlineKeyboardButton("6️⃣", callback_data="pin_num_6")],
        [InlineKeyboardButton("7️⃣", callback_data="pin_num_7"), InlineKeyboardButton("8️⃣", callback_data="pin_num_8"), InlineKeyboardButton("9️⃣", callback_data="pin_num_9")],
        [InlineKeyboardButton("🧹", callback_data="pin_clear"), InlineKeyboardButton("0️⃣", callback_data="pin_num_0"), InlineKeyboardButton("⌫", callback_data="pin_del")],
        [InlineKeyboardButton("❓ PIN kodni unutdim", callback_data="forgot_pin")]
    ]
    return text, InlineKeyboardMarkup(kb)

async def handle_pin_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()

    user_id = update.effective_user.id
    lang = get_user_lang(user_id)
    data = q.data
    current = context.user_data.get('current_pin_input', '')

    if data == "forgot_pin":
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("📧 Pochtaga kod yuborish", callback_data="reset_pin_email")],
            [InlineKeyboardButton("🔑 Maxfiy so'z orqali", callback_data="reset_pin_secret")],
            [InlineKeyboardButton("❌ Bekor qilish", callback_data="cancel_reset_pin")]
        ])
        await q.message.edit_text(
            "<b>PIN kodni tiklash usulini tanlang:</b>\n\n"
            "Iltimos, akkauntingizga ulangan pochta yoki maxfiy so'z orqali tasdiqlang.",
            reply_markup=kb,
            parse_mode=ParseMode.HTML
        )
        return

    if data == "reset_pin_email":
        with db._conn() as c:
            row = c.execute("SELECT email FROM users WHERE user_id=%s", (user_id,)).fetchone()

        email = row.get("email") if row else None

        if not email:
            await q.answer("Sizda email o'rnatilmagan! Adminga murojaat qiling.", show_alert=True)
            return

        reset_code = str(random.randint(1000, 9999))
        context.user_data['reset_pin_code'] = reset_code
        context.user_data['reset_pin_mode'] = 'email'

        await q.message.edit_text("⏳ Pochtaga kod yuborilmoqda, kuting...")
        success = send_reset_code_email(email, reset_code)

        if success:
            kb = InlineKeyboardMarkup([[InlineKeyboardButton("❌ Bekor qilish", callback_data="cancel_reset_pin")]])
            await q.message.edit_text(f"✉️ <b>{email}</b> pochtangizga 4 xonali tiklash kodi yuborildi!\n\nKodni shu yerga yozib yuboring:", parse_mode=ParseMode.HTML, reply_markup=kb)
        else:
            await q.message.edit_text("❌ Kod yuborishda xatolik yuz berdi. Pochtani tekshiring.")
        return

    if data == "reset_pin_secret":
        context.user_data['reset_pin_mode'] = 'secret'
        kb = InlineKeyboardMarkup([[InlineKeyboardButton("❌ Bekor qilish", callback_data="cancel_reset_pin")]])
        await q.message.edit_text(
            "🔑 <b>Maxfiy so'zingizni kiriting:</b>\n\n(Masalan: onamni ismi yoki itimni ismi)",
            parse_mode=ParseMode.HTML,
            reply_markup=kb
        )
        return

    if data == "cancel_reset_pin":
        context.user_data.pop('reset_pin_code', None)
        context.user_data.pop('reset_pin_mode', None)
        text, kb = get_pin_keyboard("", lang)
        await q.message.edit_text(text, reply_markup=kb, parse_mode=ParseMode.HTML)
        return

    if data == "pin_clear":
        current = ""
    elif data == "pin_del":
        current = current[:-1]
    elif data.startswith("pin_num_"):
        num = data.split("_")[2]
        if len(current) < 4:
            current += num

    context.user_data['current_pin_input'] = current

    if len(current) == 4:
        with db._conn() as c:
            row = c.execute("SELECT pin_code FROM users WHERE user_id=%s", (user_id,)).fetchone()

        if row and row.get("pin_code"):
            hashed = hashlib.sha256(current.encode()).hexdigest()
            if hashed == row["pin_code"]:
                context.user_data['is_locked'] = False
                context.user_data['current_pin_input'] = ""
                await q.message.delete()

                await context.bot.send_message(chat_id=user_id, text=get_bot_text('unlock_success', lang), parse_mode=ParseMode.HTML)

                pending_payload = context.user_data.pop('pending_payload', None)

                if pending_payload and pending_payload.startswith("test_"):
                    test_id = pending_payload.replace("test_", "", 1)
                    await send_test_options(update, context, test_id)
                elif pending_payload and pending_payload.startswith("export_"):
                    test_id = pending_payload.replace("export_", "", 1)
                    await export_excel(update, context, test_id)
                else:
                    msg, kb = await build_main_menu(user_id, context.bot.username, lang)
                    await context.bot.send_message(chat_id=user_id, text=msg, reply_markup=kb, parse_mode=ParseMode.HTML)
                return
            else:
                context.user_data['current_pin_input'] = ""
                text, kb = get_pin_keyboard("", lang)
                text = get_bot_text('wrong_pin', lang) + "\n\n" + text
                await q.message.edit_text(text, reply_markup=kb, parse_mode=ParseMode.HTML)
                return

    text, kb = get_pin_keyboard(current, lang)
    try:
        await q.message.edit_text(text, reply_markup=kb, parse_mode=ParseMode.HTML)
    except Exception:
        pass

async def check_verified_middleware(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not update.effective_user or update.effective_chat.type != "private":
        return

    user_id = update.effective_user.id

    if context.user_data.get('pending_premium_months') or context.user_data.get('pending_gwt_amount'):
        return

    lang = get_user_lang(user_id)
    user_row = dict(db.get_user(user_id) or {})
    is_verified = user_row.get("is_verified", 1)

    if is_verified == 1:
        return

    now = time.time()
    last_warn = context.user_data.get('last_captcha_warn', 0)

    if now - last_warn > 5:
        context.user_data['last_captcha_warn'] = now
        captcha_url = f"{WEB_BASE_URL.rstrip('/')}/captcha"
        markup = InlineKeyboardMarkup([[InlineKeyboardButton("🤖 Men robot emasman", web_app=WebAppInfo(url=captcha_url))]])

        try:
            del_msg = await update.effective_chat.send_message("🚫", reply_markup=ReplyKeyboardRemove())
            await del_msg.delete()
        except:
            pass

        try:
            if update.callback_query:
                await update.callback_query.answer("Tasdiqlashdan o'ting!", show_alert=True)

            await update.effective_chat.send_message(
                get_bot_text('captcha_msg', lang),
                reply_markup=markup,
                parse_mode=ParseMode.HTML
            )
        except Exception as e:
            logging.error(f"Kapcha xabarini yuborishda xato: {e}")
    else:
        if update.callback_query:
            try:
                await update.callback_query.answer()
            except:
                pass

    raise ApplicationHandlerStop

async def check_lock_middleware(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not update.effective_user:
        return

    if update.effective_chat and update.effective_chat.type != "private":
        return

    if update.callback_query and re.match(r"^(pin_|forgot_pin|cancel_reset_pin)", update.callback_query.data):
        raise ApplicationHandlerStop

    user_id = update.effective_user.id
    lang = get_user_lang(user_id)

    if not context.user_data.get('is_locked', False):
        return

    if context.user_data.get('reset_pin_mode'):
        return

    with db._conn() as c:
        row = c.execute("SELECT pin_code FROM users WHERE user_id=%s", (user_id,)).fetchone()

    if not row or not row.get("pin_code"):
        context.user_data['is_locked'] = False
        return

    if update.message and update.message.text and update.message.text.startswith('/start '):
        payload = update.message.text.split(' ', 1)[1]
        context.user_data['pending_payload'] = payload

    context.user_data['current_pin_input'] = ""
    text, kb = get_pin_keyboard("", lang)

    if update.message:
        await update.message.reply_text(text, reply_markup=kb, parse_mode=ParseMode.HTML)
    elif update.callback_query:
        await update.callback_query.answer("Tizim qulflangan!", show_alert=True)
        await update.callback_query.message.reply_text(text, reply_markup=kb, parse_mode=ParseMode.HTML)

    raise ApplicationHandlerStop

# ==========================================
# SPAM TEKSHIRUVI MIDDLEWARE
# ==========================================
async def spam_check_middleware(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not update.effective_user or not update.effective_chat:
        return
    if update.effective_chat.type != "private":
        return

    user_id = update.effective_user.id
    lang = get_user_lang(user_id)
    now = time.time()

    spam_data = context.user_data.get('spam_track', {
        'start_time': now,
        'count': 0,
        'warnings': 0
    })

    if now - spam_data['start_time'] > 60:
        spam_data['start_time'] = now
        spam_data['count'] = 0

    spam_data['count'] += 1

    if spam_data['count'] > 100:
        spam_data['warnings'] += 1
        spam_data['count'] = 0
        spam_data['start_time'] = now

        warnings = spam_data['warnings']

        if warnings == 1:
            await context.bot.send_message(
                chat_id=user_id,
                text="⚠️ <b>1-ogohlantirish:</b> Siz juda tez harakat qilyapsiz! Iltimos, daqiqasiga 100 ta so'rov limitidan oshmang.",
                parse_mode=ParseMode.HTML
            )
            raise ApplicationHandlerStop

        elif warnings >= 2:
            try:
                with db._conn() as c:
                    c.execute("UPDATE users SET is_verified = 0 WHERE user_id = %s", (user_id,))
            except Exception as e:
                logging.error(f"Anti-spam DB update xatosi: {e}")

            spam_data['warnings'] = 0

            captcha_url = f"{WEB_BASE_URL.rstrip('/')}/captcha"
            markup = InlineKeyboardMarkup([[InlineKeyboardButton("🤖 Men robot emasman", web_app=WebAppInfo(url=captcha_url))]])

            await context.bot.send_message(
                chat_id=user_id,
                text=get_bot_text('captcha_msg', lang),
                reply_markup=markup,
                parse_mode=ParseMode.HTML
            )
            raise ApplicationHandlerStop

    context.user_data['spam_track'] = spam_data

# ==========================================
# 🛑 MAJBURIY OBUNA (FORCE SUB) MIDDLEWARE
# ==========================================
async def check_subscription_middleware(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not update.effective_user or not update.effective_chat:
        return
    if update.effective_chat.type != "private":
        return

    if update.callback_query and update.callback_query.data == "check_sub":
        return

    user_id = update.effective_user.id
    lang = get_user_lang(user_id)

    if user_id in SUPERADMINS:
        return

    last_check = context.user_data.get('last_sub_check_time', 0)
    is_subbed = context.user_data.get('is_subscribed', False)

    now = time.time()

    if now - last_check > 300 or not is_subbed:
        try:
            member = await context.bot.get_chat_member(chat_id=REQUIRED_CHANNEL, user_id=user_id)
            if member.status in [ChatMemberStatus.LEFT, ChatMemberStatus.BANNED]:
                is_subbed = False
            else:
                is_subbed = True

            context.user_data['last_sub_check_time'] = now
            context.user_data['is_subscribed'] = is_subbed
        except Exception as e:
            logging.error(f"Kanal obunasini tekshirishda xato: {e}")
            is_subbed = True
            context.user_data['is_subscribed'] = True

    if not is_subbed:
        last_prompt = context.user_data.get('last_sub_prompt_time', 0)

        if now - last_prompt > 5:
            context.user_data['last_sub_prompt_time'] = now

            channel_link = f"https://t.me/{REQUIRED_CHANNEL.replace('@', '')}"
            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton(get_bot_text('sub_btn_join', lang), url=channel_link)],
                [InlineKeyboardButton(get_bot_text('sub_btn_check', lang), callback_data="check_sub")]
            ])
            text = get_bot_text('sub_req_msg', lang, channel=REQUIRED_CHANNEL)

            try:
                del_msg = await update.effective_chat.send_message("⏳", reply_markup=ReplyKeyboardRemove())
                await del_msg.delete()
            except:
                pass

            try:
                if update.callback_query:
                    await update.callback_query.message.reply_text(text, reply_markup=kb, parse_mode=ParseMode.HTML)
                elif update.message:
                    await update.message.reply_text(text, reply_markup=kb, parse_mode=ParseMode.HTML)
            except Exception as e:
                logging.error(f"Obuna xabarini yuborishda xato: {e}")

        raise ApplicationHandlerStop

# ==========================================
# TELEGRAM YULDUZLAR (STARS) HANDLERLARI
# ==========================================
async def precheckout_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.pre_checkout_query
    if query.invoice_payload.startswith("premium_") or query.invoice_payload.startswith("buygwt_"):
        await query.answer(ok=True)
    else:
        await query.answer(ok=False, error_message="Noma'lum to'lov.")

async def successful_payment_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await upsert_user_from_update(update)
    payload = update.message.successful_payment.invoice_payload
    user_id = update.effective_user.id

    if payload.startswith("premium_"):
        months = int(payload.split("_")[1])
        db.add_premium_months(user_id, months)

        try:
            admin_id = MAIN_ADMIN_ID if MAIN_ADMIN_ID else (list(SUPERADMINS)[0] if SUPERADMINS else None)
            if admin_id:
                await context.bot.send_message(
                    chat_id=admin_id,
                    text=f"⭐️ <b>YANGI YULDUZLI TO'LOV (AVTOMATIK TASDIQLANDI)!</b>\n👤 Foydalanuvchi: {update.effective_user.first_name} (ID: <code>{user_id}</code>)\n📦 Ta'rif: <b>{months} oylik Premium</b>",
                    parse_mode=ParseMode.HTML
                )
        except Exception:
            pass

        await update.message.reply_text(
            f"🎉 <b>Tabriklaymiz!</b> To'lov muvaffaqiyatli amalga oshirildi.\n\nSizga avtomatik ravishda <b>{months} oylik PREMIUM</b> taqdim etildi! 💎",
            parse_mode=ParseMode.HTML
        )

    elif payload.startswith("buygwt_"):
        amount = float(payload.split("_")[1])
        db.system_sell_token(user_id, amount, method="STARS")
        await update.message.reply_text(
            f"🎉 <b>Tabriklaymiz!</b> To'lov muvaffaqiyatli amalga oshirildi.\n\nSizning hamyoningizga avtomatik ravishda <b>{amount} GWT</b> tashlab berildi! 💰",
            parse_mode=ParseMode.HTML
        )


# ==========================================
# 6. ASOSIY MENYU VA CHAT HANDLERLARI
# ==========================================
async def build_main_menu(user_id: int, bot_username: str, lang: str = "uz") -> tuple[str, ReplyKeyboardMarkup]:
    is_superadmin = user_id in SUPERADMINS
    is_lower_admin = user_id in LOWER_ADMINS

    kb = [
        [KeyboardButton(text=get_bot_text('btn_cabinet', lang)), KeyboardButton(text=get_bot_text('btn_account', lang))],
        [KeyboardButton(text=get_bot_text('btn_wallet', lang))],
        [KeyboardButton(text=get_bot_text('btn_ai', lang)), KeyboardButton(text=get_bot_text('btn_search', lang))],
        [KeyboardButton(text=get_bot_text('btn_create_manual', lang)), KeyboardButton(text=get_bot_text('btn_create_word', lang))],
        [KeyboardButton(text=get_bot_text('btn_results', lang)), KeyboardButton(text=get_bot_text('btn_check_chats', lang))],
        [KeyboardButton(text=get_bot_text('btn_premium', lang)), KeyboardButton(text=get_bot_text('btn_top', lang))],
        [KeyboardButton(text=get_bot_text('btn_add_bot', lang)), KeyboardButton(text=get_bot_text('btn_referral', lang))],
        [KeyboardButton(text=get_bot_text('btn_lock', lang)), KeyboardButton(text="👨‍💻 Adminga murojaat")]
    ]

    if is_superadmin or is_lower_admin:
        kb.append([KeyboardButton(text="📣 Reklama Tarqatish"), KeyboardButton(text="📊 Mening reklamalarim")])
    if is_superadmin:
        kb.append([KeyboardButton(text="📈 Statistika"), KeyboardButton(text="📋 Kanallar ro'yxati")])

    start_text = get_bot_text('welcome', lang, name="")
    return start_text, ReplyKeyboardMarkup(kb, resize_keyboard=True, is_persistent=True)

async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_chat.type != "private":
        try:
            await update.effective_chat.send_message("✅ Bot faol!")
        except Exception:
            pass
        return

    user = update.effective_user
    user_id = user.id
    lang = get_user_lang(user_id)
    user_row = db.get_user(user_id)

    payload = context.args[0].strip() if context.args else None
    referrer_id = None

    if payload and payload.startswith("ref_"):
        try:
            referrer_id = int(payload.split("_")[1])
        except ValueError:
            pass

    # Yangi foydalanuvchi va Referal tekshiruvi (Bonuslar beriladi)
    if not user_row:
        try:
            # Diqqat: Bu funksiyani db.py da yaratgan bo'lishingiz kerak
            is_new, ref_count = db.register_new_user_with_bonus(
                user_id=user.id,
                username=user.username or "",
                first_name=user.first_name or "",
                last_name=user.last_name or "",
                now_ts=now_ts(),
                referrer_id=referrer_id
            )
            db.get_or_create_user_api_key(user.id)

            if is_new:
                # 1 GWT Bonus xabari
                try:
                    await context.bot.send_message(chat_id=user_id, text="🎉 <b>Xush kelibsiz!</b>\n\nSizga ro'yxatdan o'tganingiz uchun <b>1 GWT</b> miqdorida kripto bonus taqdim etildi! 💰\n\nBu mablag'ni hamyoningizda ko'rishingiz mumkin.", parse_mode=ParseMode.HTML)
                except Exception: pass

                # Referal egasiga bonus xabari (agar u 10 ta yig'gan bo'lsa)
                if referrer_id and ref_count > 0 and ref_count % 10 == 0:
                    db.add_premium_months(referrer_id, 1)
                    try:
                        await context.bot.send_message(chat_id=referrer_id, text=f"🎉 <b>Tabriklaymiz!</b>\n\nSiz 10 ta do'stingizni taklif qildingiz va <b>1 Oylik Premium</b> yutib oldingiz! 💎", parse_mode=ParseMode.HTML)
                    except Exception: pass
        except Exception as e:
            logging.error(f"Foydalanuvchini ro'yxatdan o'tkazishda xato: {e}")
            await upsert_user_from_update(update)
    else:
        await upsert_user_from_update(update)

    context.user_data.pop(K["mode"], None)
    context.user_data.pop("pending_premium_months", None)
    context.user_data.pop('pending_gwt_amount', None)

    # Kapcha tekshiruvi (is_verified)
    user_row = dict(db.get_user(user_id) or {})
    is_verified = user_row.get("is_verified", 0)

    if payload == 'guide':
        if not is_verified:
            db.set_pending_payload(user_id, "guide")
        else:
            token = db.get_or_create_user_api_key(user_id)
            guide_url = f"{WEB_BASE_URL.rstrip('/')}/guide?token={token}"
            kb = InlineKeyboardMarkup([[InlineKeyboardButton("📖 Yo'riqnomani ochish", web_app=WebAppInfo(url=guide_url))]])
            await update.message.reply_text(
                "📚 <b>Geo Ustoz - Yo'riqnoma va Yordam</b>\n\n"
                "Platformadan foydalanish bo'yicha to'liq qo'llanmani o'qish va bevosita adminga xabar yozish uchun pastdagi tugmani bosing:",
                reply_markup=kb,
                parse_mode=ParseMode.HTML
            )
            return

    if not is_verified:
        if payload and payload != 'guide':
            db.set_pending_payload(user_id, payload)

        try:
            del_msg = await update.effective_chat.send_message("🔄", reply_markup=ReplyKeyboardRemove())
            await del_msg.delete()
        except:
            pass

        captcha_url = f"{WEB_BASE_URL.rstrip('/')}/captcha"
        markup = InlineKeyboardMarkup([[InlineKeyboardButton("🤖 Men robot emasman", web_app=WebAppInfo(url=captcha_url))]])
        await update.effective_chat.send_message(
            get_bot_text('captcha_msg', lang),
            reply_markup=markup,
            parse_mode=ParseMode.HTML
        )
        return

    if payload and payload.startswith("export_"):
        test_id = payload.replace("export_", "", 1)
        await export_excel(update, context, test_id)
        return

    if payload and payload.startswith("test_"):
        test_id = payload.replace("test_", "", 1)
        await send_test_options(update, context, test_id)
        return

    msg, kb_main = await build_main_menu(user_id, context.bot.username, lang)

    webapp_url = f"{WEB_BASE_URL.rstrip('/')}/telegram-login"
    markup = InlineKeyboardMarkup([
        [InlineKeyboardButton(text=get_bot_text('btn_cabinet', lang), web_app=WebAppInfo(url=webapp_url))],
        [InlineKeyboardButton(text=get_bot_text('change_lang_btn', lang), callback_data="change_lang")]
    ])

    await update.effective_chat.send_message(get_bot_text('main_menu_loaded', lang), reply_markup=kb_main)

    await update.effective_chat.send_message(
        get_bot_text('welcome', lang, name=user.first_name),
        reply_markup=markup,
        parse_mode="HTML"
    )

async def cmd_cabinet(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_chat.type != "private":
        return
    await upsert_user_from_update(update)
    user_id = update.effective_user.id
    lang = get_user_lang(user_id)
    webapp_url = f"{WEB_BASE_URL.rstrip('/')}/telegram-login"
    kb = InlineKeyboardMarkup([[InlineKeyboardButton(get_bot_text('btn_cabinet', lang), web_app=WebAppInfo(url=webapp_url))]])
    await update.message.reply_text("Kabinetingizga kirish uchun quyidagi tugmani bosing:", reply_markup=kb)

# ADMIN KOMANDASI: Oylik reyting g'oliblariga GWT tarqatish
async def cmd_reward_top(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if user_id not in SUPERADMINS:
        return

    await update.message.reply_text("⏳ Oylik reyting hisoblanmoqda va mukofotlar tarqatilmoqda. Iltimos kuting...")
    try:
        # db.py da yozilgan funksiya chaqiriladi
        winners = db.distribute_monthly_rewards()

        if not winners:
            await update.message.reply_text("Hech kim topilmadi yoki bu oyda natijalar yo'q.")
            return

        text = "🏆 <b>OYLIK TOP-100 YAKUNLANDI!</b>\n\nEng ko'p va sifatli test yechgan bilimdonlar taqdirlandi:\n\n"
        for w in winners[:15]:
            text += f"<b>{w['rank']}-o'rin:</b> {html.escape(w['name'])} - <b>{w['reward']} GWT</b> 💰\n"
        if len(winners) > 15:
            text += f"\n... va yana {len(winners) - 15} ta foydalanuvchiga 1 GWT dan berildi!"

        text += "\n\n🎉 Barcha g'oliblarga mukofotlar ularning <b>Geo Wallet</b> hamyoniga o'tkazildi!"

        await update.message.reply_text(text, parse_mode=ParseMode.HTML)

        # Kanallarga tarqatish
        try:
            await context.bot.send_message(chat_id=PUBLIC_TEST_CHANNEL, text=text, parse_mode=ParseMode.HTML)
        except Exception: pass

        # Har bir g'olibga lichkasiga yozish
        for w in winners:
            try:
                await context.bot.send_message(
                    chat_id=w['user_id'],
                    text=f"🎉 <b>Tabriklaymiz!</b>\n\nSiz oylik reytingda <b>{w['rank']}-o'rinni</b> egalladingiz va <b>{w['reward']} GWT</b> yutib oldingiz!\n\nTangalar hamyoningizga o'tkazildi. Bundanda yuqori natijalarga erishishda davom eting! 🚀",
                    parse_mode=ParseMode.HTML
                )
            except Exception: pass

    except Exception as e:
        await update.message.reply_text(f"❌ Xatolik yuz berdi: {e}")

async def on_my_chat_member(update: Update, context: ContextTypes.DEFAULT_TYPE):
    cmu = update.my_chat_member
    if not cmu: return
    chat = cmu.chat
    from_user = cmu.from_user
    new_status = cmu.new_chat_member.status
    ctype = chat.type
    title = chat.title or (chat.username or "")

    bot_is_admin = 1 if new_status in (ChatMemberStatus.ADMINISTRATOR, ChatMemberStatus.OWNER) else 0
    added_by = from_user.id if from_user else None

    db.upsert_chat(chat.id, title, ctype, added_by, bot_is_admin, now_ts())

async def on_chat_member(update: Update, context: ContextTypes.DEFAULT_TYPE):
    cmu = update.chat_member
    if not cmu: return

    user_id = cmu.new_chat_member.user.id
    new_status = cmu.new_chat_member.status

    if new_status in [ChatMemberStatus.MEMBER, ChatMemberStatus.RESTRICTED]:
        if user_id in SUPERADMINS:
            try:
                await context.bot.promote_chat_member(
                    chat_id=cmu.chat.id, user_id=user_id, can_manage_chat=True,
                    can_change_info=True, can_post_messages=True, can_edit_messages=True,
                    can_delete_messages=True, can_invite_users=True, can_restrict_members=True,
                    can_pin_messages=True, can_manage_video_chats=True
                )
            except Exception as e:
                logging.error(f"SuperAdminni admin qilishda xatolik: {e}")

# ==========================================
# 7. BROADCAST (REKLAMA) FUNKSIYALARI
# ==========================================
async def show_channel_selection(chat_id_to_send: int, context: ContextTypes.DEFAULT_TYPE, msg_id_to_edit: int = None):
    b = context.user_data.get(K["broadcast"])
    if not b: return

    targets = b.get("targets", {})
    selected = b.get("selected", {})

    kb = []
    if targets:
        all_selected = all(selected.values())
        toggle_all_btn = "❌ Barchasini o'chirish" if all_selected else "✅ Barchasini tanlash"
        kb.append([InlineKeyboardButton(toggle_all_btn, callback_data="bc_toggle_all")])

    for cid, title in targets.items():
        is_sel = selected.get(cid, False)
        prefix = "✅" if is_sel else "❌"
        short_title = title[:30] + "..." if len(title) > 30 else title
        kb.append([InlineKeyboardButton(f"{prefix} {short_title}", callback_data=f"bc_toggle:{cid}")])

    sel_count = sum(1 for v in selected.values() if v)
    kb.append([InlineKeyboardButton(f"🚀 Yuborish ({sel_count} ta chat)", callback_data="bc_send")])
    kb.append([InlineKeyboardButton("❌ Bekor qilish", callback_data="cancel_action")])

    text = f"📢 <b>Reklama qaysi kanallarga yuborilsin?</b>\nKeraksizlarini ustiga bosib o'chirib qo'yishingiz mumkin.\n\nTanlangan kanallar soni: {sel_count} ta"

    if msg_id_to_edit:
        try:
            await context.bot.edit_message_text(text, chat_id=chat_id_to_send, message_id=msg_id_to_edit, reply_markup=InlineKeyboardMarkup(kb), parse_mode=ParseMode.HTML)
        except Exception:
            pass
    else:
        await context.bot.send_message(chat_id=chat_id_to_send, text=text, reply_markup=InlineKeyboardMarkup(kb), parse_mode=ParseMode.HTML)

# ==========================================
# 8. CALLBACK QUERY HANDLER (TUGMALAR)
# ==========================================
async def on_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await upsert_user_from_update(update)
    q = update.callback_query
    if not q: return

    data = q.data or ""
    user_id = update.effective_user.id
    lang = get_user_lang(user_id)

    if data == "check_sub":
        try:
            member = await context.bot.get_chat_member(chat_id=REQUIRED_CHANNEL, user_id=user_id)
            if member.status not in [ChatMemberStatus.LEFT, ChatMemberStatus.BANNED]:
                context.user_data['is_subscribed'] = True
                context.user_data['last_sub_check_time'] = time.time()
                try:
                    await q.message.delete()
                except:
                    pass
                await q.answer("✅ Obuna tasdiqlandi! Botdan foydalanishingiz mumkin.", show_alert=True)

                msg, kb_main = await build_main_menu(user_id, context.bot.username, lang)
                await update.effective_chat.send_message("✅ Obuna muvaffaqiyatli tasdiqlandi!", reply_markup=kb_main)
            else:
                await q.answer("❌ Hali kanalga qo'shilmadingiz! Iltimos, kanalga qo'shilib tasdiqlang.", show_alert=True)
        except Exception as e:
            logging.error(f"Kanalni tekshirishda xatolik: {e}")
            await q.answer("⚠️ XATOLIK: Botingiz kanalingizga ADMIN qilinmagan! Iltimos botni avval kanalga admin qiling.", show_alert=True)
        return

    is_superadmin = user_id in SUPERADMINS
    is_lower_admin = user_id in LOWER_ADMINS

    user_row = dict(db.get_user(user_id) or {})
    is_verified = user_row.get("is_verified", 1)

    if not is_verified and data != "change_lang" and not data.startswith("setlang_"):
        try:
            await q.answer("Iltimos, avval botning shaxsiy sahifasida kapchadan o'ting!", show_alert=True)
        except:
            pass
        return

    if data == "change_lang":
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("🇺🇿 O'zbek (Lotin)", callback_data="setlang_uz")],
            [InlineKeyboardButton("🇺🇿 Ўзбек (Кирилл)", callback_data="setlang_uz_cyrl")],
            [InlineKeyboardButton("🇷🇺 Русский", callback_data="setlang_ru")]
        ])
        await q.message.edit_text(get_bot_text('choose_lang', lang), reply_markup=kb)
        return

    if data.startswith("setlang_"):
        new_lang = data.split("_", 1)[1]
        try:
            with db._conn() as c:
                c.execute("UPDATE users SET lang=%s WHERE user_id=%s", (new_lang, user_id))
        except Exception:
            pass

        await q.answer(get_bot_text('lang_saved', new_lang), show_alert=False)

        user_row = dict(db.get_user(user_id) or {})
        status = user_row.get("status", "free")
        status_text = "💎 PREMIUM" if status == "premium" else ("Oddiy" if new_lang == 'uz' else ("Оддий" if new_lang == 'uz_cyrl' else "Обычный"))
        premium_text = f"\n⏳ Premium: <b>{ts_to_local(user_row.get('premium_expire_at'))}</b>" if status == "premium" and user_row.get("premium_expire_at") else ""

        text_acc = get_bot_text('account', new_lang, user_id=user_id, name=update.effective_user.first_name, status=status_text, premium_text=premium_text)

        markup = InlineKeyboardMarkup([
            [InlineKeyboardButton(text=get_bot_text('btn_cabinet', new_lang), web_app=WebAppInfo(url=f"{WEB_BASE_URL}/telegram-login"))],
            [InlineKeyboardButton(text=get_bot_text('change_lang_btn', new_lang), callback_data="change_lang")]
        ])

        await q.message.edit_text(text_acc, reply_markup=markup, parse_mode=ParseMode.HTML)

        msg_text, kb_reply = await build_main_menu(user_id, context.bot.username, new_lang)
        await context.bot.send_message(chat_id=user_id, text=get_bot_text('main_menu_loaded', new_lang), reply_markup=kb_reply)
        return

    if data == "mass_verify_1":
        if user_id not in SUPERADMINS: return
        try:
            with db._conn() as c:
                c.execute("UPDATE users SET is_verified=1")
            await q.message.edit_text("✅ <b>Barcha foydalanuvchilar tasdiqdan ozod qilindi!</b>\n\nEndi ulardan kapcha so'ralmaydi.", parse_mode=ParseMode.HTML)
        except Exception as e:
            await q.message.edit_text(f"❌ Xatolik: {e}")
        return

    if data == "withdraw_gwt":
        balance = db.get_token_balance(user_id)
        if balance <= 0:
            await q.answer("❌ Balansingizda pul yechish uchun GWT yo'q!", show_alert=True)
            return

        context.user_data[K["mode"]] = "withdraw_typing_card"
        kb = InlineKeyboardMarkup([[InlineKeyboardButton(get_bot_text('btn_cancel', lang), callback_data="wallet_refresh")]])
        await q.message.edit_text(
            f"💸 <b>GWT ni naqd pulga almashtirish</b>\n\n"
            f"Sizning balansingiz: <b>{balance} GWT</b>\n"
            f"Kurs: 1 GWT = 120,000 so'm\n\n"
            f"💳 Iltimos, pul o'tkazib berishimiz uchun <b>Karta raqamingizni</b> yuboring (16 ta raqam):",
            reply_markup=kb, parse_mode=ParseMode.HTML
        )
        return

    if data == "mass_verify_0":
        if user_id not in SUPERADMINS: return
        try:
            with db._conn() as c:
                c.execute("UPDATE users SET is_verified=0")
            await q.message.edit_text("🤖 <b>Barcha foydalanuvchilar bot deb belgilandi!</b>\n\nEndi hammadan botni ishlatishdan oldin kapcha so'raladi.", parse_mode=ParseMode.HTML)
        except Exception as e:
            await q.message.edit_text(f"❌ Xatolik: {e}")
        return

    if data.startswith("ad:"):
        ad_id = data.split(":")[1]
        try:
            ad = db.get_ad(ad_id)
        except Exception:
            ad = None

        if ad:
            try:
                db.register_ad_click(ad_id, user_id)
            except Exception:
                pass

            rt = str(ad["reply_text"])
            if len(rt) > 190: rt = rt[:190] + "..."
            try:
                await q.answer(rt, show_alert=True)
            except Exception:
                pass
        else:
            try:
                await q.answer("⚠️ Bu reklama eskirgan.", show_alert=True)
            except Exception:
                pass
        return
    # --- ADMIN PUL YECHISHNI TASDIQLASHI ---
    if data.startswith("approve_withdraw_"):
        parts = data.split("_")
        target_user = int(parts[2])
        amount = float(parts[3])

        await q.message.edit_text(q.message.text + "\n\n✅ <b>HOLAT: TO'LAB BERILDI VA TASDIQLANDI!</b>", parse_mode=ParseMode.HTML)
        try:
            await context.bot.send_message(
                chat_id=target_user,
                text=f"🎉 <b>Pul yechish so'rovingiz tasdiqlandi!</b>\n\nSiz yechib olgan <b>{amount} GWT</b> puli kartangizga tushirib berildi. Moliya bo'limini tekshiring!",
                parse_mode=ParseMode.HTML
            )
        except: pass
        return

    if data.startswith("reject_withdraw_"):
        parts = data.split("_")
        target_user = int(parts[2])
        amount = float(parts[3])

        # Rad etilgani uchun GWT ni qaytarib beramiz
        db.system_sell_token(target_user, amount, method="REFUND")

        await q.message.edit_text(q.message.text + "\n\n❌ <b>HOLAT: RAD ETILDI (GWT QAYTARILDI)!</b>", parse_mode=ParseMode.HTML)
        try:
            await context.bot.send_message(
                chat_id=target_user,
                text=f"❌ <b>Pul yechish so'rovingiz rad etildi!</b>\n\nKarta raqamida xatolik bo'lishi mumkin. Siz yechmoqchi bo'lgan <b>{amount} GWT</b> balansingizga qaytarildi.",
                parse_mode=ParseMode.HTML
            )
        except: pass
        return

    # ==========================================
    # 👛 HAMYON VA 💎 PREMIUM XARID QILISH
    # ==========================================

    if data == "btn_premium":
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton(get_bot_text('btn_pay_card', lang), callback_data="pay_method_card")],
            [InlineKeyboardButton("💰 GWT Token (Avtomatik)", callback_data="pay_method_gwt")],
            [InlineKeyboardButton(get_bot_text('btn_pay_stars', lang), callback_data="pay_method_stars")],
            [InlineKeyboardButton(get_bot_text('btn_back_main', lang), callback_data="back_to_main")]
        ])

        text = (
            "💎 <b>Premium xarid qilish usulini tanlang:</b>\n\n"
            "💳 <b>Karta orqali:</b> To'lov qilgach chekni adminga yuborasiz va tasdiqlashni kutasiz.\n"
            "💰 <b>GWT orqali:</b> Balansingizdan yechiladi va Premium <b>avtomatik va shu zahoti</b> beriladi!\n"
            "⭐ <b>Telegram Stars:</b> To'lov qilsangiz, Premium <b>avtomatik va shu zahoti</b> beriladi!"
        )
        await q.message.edit_text(text, reply_markup=kb, parse_mode=ParseMode.HTML)
        return

    if data == "pay_method_gwt":
        balance = db.get_token_balance(user_id)
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("🥉 1 Oylik — 0.16 GWT", callback_data="buy_prem_gwt:1")],
            [InlineKeyboardButton("🥈 3 Oylik — 0.44 GWT", callback_data="buy_prem_gwt:3")],
            [InlineKeyboardButton("🥇 6 Oylik — 0.87 GWT", callback_data="buy_prem_gwt:6")],
            [InlineKeyboardButton("💎 12 Oylik — 1.59 GWT", callback_data="buy_prem_gwt:12")],
            [InlineKeyboardButton(get_bot_text('btn_back', lang), callback_data="btn_premium")]
        ])
        await q.message.edit_text(
            f"💰 <b>GWT Token orqali Premium:</b>\n\n"
            f"Sizning balansingiz: <b>{balance} GWT</b>\n"
            f"Kurs: 1 GWT ≈ 126,000 so'm\n\n"
            "Kerakli muddatni tanlang:",
            reply_markup=kb, parse_mode=ParseMode.HTML
        )
        return

    if data.startswith("buy_prem_gwt:"):
        months = int(data.split(":")[1])
        gwt_prices = {1: 0.16, 3: 0.44, 6: 0.87, 12: 1.59}
        price = gwt_prices.get(months, 0.16)

        try:
            wallet = db.get_wallet(user_id)
        except AttributeError:
            wallet = None

        if not wallet:
            await q.answer("❌ Sizda hamyon yo'q! Avval hamyon yarating.", show_alert=True)
            return

        if db.get_token_balance(user_id) < price:
            await q.answer(f"❌ Balansda yetarli GWT yo'q! (Kerak: {price} GWT)", show_alert=True)
            return

        transaction_data = f"{wallet['public_key']}->GENESIS:{price}".encode('utf-8')
        signature = crypto_mgr.sign_transaction(wallet['encrypted_private_key'], transaction_data)

        success, err_msg = db.transfer_token_by_address_or_id(
            sender_id=user_id,
            target="GENESIS",
            amount=price,
            signature=signature
        )

        if success:
            db.add_premium_months(user_id, months)
            await q.message.edit_text(
                f"🎉 <b>Tabriklaymiz!</b>\n\n"
                f"Hisobingizdan <b>{price} GWT</b> muvaffaqiyatli yechildi.\n"
                f"Sizga <b>{months} oylik</b> Premium maqomi berildi! ✅",
                parse_mode=ParseMode.HTML
            )
            await q.answer("To'lov muvaffaqiyatli!", show_alert=False)
        else:
            await q.answer(f"❌ Xatolik: {err_msg}", show_alert=True)
        return

    if data == "buy_gwt_menu":
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("💳 Karta orqali", callback_data="buy_gwt_card_prompt"),
             InlineKeyboardButton("⭐️ Stars orqali", callback_data="buy_gwt_stars_prompt")],
            [InlineKeyboardButton(get_bot_text('btn_back', lang), callback_data="wallet_refresh")]
        ])
        await q.message.edit_text("➕ <b>GWT Sotib olish:</b>\n\n💎 1 GWT ≈ 126,000 so'm\n\nTo'lov usulini tanlang:", reply_markup=kb, parse_mode=ParseMode.HTML)
        return

    if data == "buy_gwt_card_prompt":
        context.user_data[K["mode"]] = "buy_gwt_typing_card"
        kb = InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Bekor qilish", callback_data="buy_gwt_menu")]])
        await q.message.edit_text("💳 <b>Karta orqali xarid:</b>\n\nQancha GWT sotib olmoqchisiz? Miqdorni yozing (masalan, <code>0.5</code>):", reply_markup=kb, parse_mode=ParseMode.HTML)
        return

    if data == "buy_gwt_stars_prompt":
        context.user_data[K["mode"]] = "buy_gwt_typing_stars"
        kb = InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Bekor qilish", callback_data="buy_gwt_menu")]])
        await q.message.edit_text("⭐️ <b>Telegram Stars orqali xarid:</b>\n\nQancha GWT sotib olmoqchisiz? Miqdorni yozing:", reply_markup=kb, parse_mode=ParseMode.HTML)
        return

    if data == "wallet_refresh":
        wallet = db.get_wallet(user_id)
        balance = db.get_token_balance(user_id)
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("➕ GWT Sotib olish", callback_data="buy_gwt_menu")],
            [InlineKeyboardButton("💳 Pul yechish (Naqd qilish)", callback_data="withdraw_gwt")], # <-- YANGI TUGMA
            [InlineKeyboardButton(get_bot_text('btn_transfer', lang), callback_data="wallet_transfer"),
             InlineKeyboardButton(get_bot_text('btn_receive', lang), callback_data="wallet_receive")],
            [InlineKeyboardButton(get_bot_text('btn_back_main', lang), callback_data="back_to_main")]
        ])
        await q.message.edit_text(get_bot_text('wallet_info', lang, address=wallet['public_key'], balance=balance), reply_markup=kb, parse_mode=ParseMode.HTML)
        return

    if data == "wallet_transfer":
        context.user_data[K["mode"]] = "wallet_transfer_amount"
        await q.message.edit_text("💸 O'tkazmoqchi bo'lgan token (GWT) miqdorini kiriting:\n\n<i>Masalan: 1.5 yoki 0.2</i>", parse_mode=ParseMode.HTML)
        return

    if data == "wallet_receive":
        try:
            wallet = db.get_wallet(user_id)
            if wallet:
                await q.message.edit_text(f"📥 <b>Sizning qabul qilish manzilingiz (Public Key):</b>\n\n<code>{wallet['public_key']}</code>\n\nShu manzilni to'lovchiga yuboring.", parse_mode=ParseMode.HTML)
        except Exception:
            pass
        return

    if data == "pdf_ai_theme":
        pending = context.user_data.get(K["convert"], {}).get("pending_test")
        if not pending:
            await q.answer("Xatolik: Test ma'lumotlari topilmadi.", show_alert=True)
            return

        await q.message.edit_text("⏳ AI mavzu o'ylamoqda...")

        q_text = "\n".join([qu["question"] for qu in pending["questions"][:5]])
        ai_theme = await get_theme_from_ai(q_text)
        pending["theme"] = ai_theme

        if pending.get("missing_keys"):
            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton("🤖 AI o'zi yechib bersin", callback_data="pdf_ai_solve")],
                [InlineKeyboardButton("✍️ Kalitlarni o'zim kiritaman", callback_data="pdf_manual_keys")]
            ])
            await q.message.edit_text(
                f"✅ AI mavzu tanladi: <b>{ai_theme}</b>\n\n⚠️ Lekin kalitlar topilmadi. Nima qilamiz?",
                reply_markup=kb, parse_mode=ParseMode.HTML
            )
        else:
            await q.message.edit_text(f"✅ AI mavzu tanladi: <b>{ai_theme}</b>\nTest yaratilmoqda...", parse_mode=ParseMode.HTML)
            await build_and_send_test(update, context, ai_theme, pending["questions"], pending["target_chat"])
        return

    # --- ADMIN TASDIQLASHI ---
    if data.startswith("appr_wd_"):
        p = data.split("_")
        t_user, amt = int(p[2]), float(p[3])
        await q.message.edit_text(q.message.text + "\n\n✅ <b>TASDIQLANDI: PUL TO'LANDI</b>", parse_mode=ParseMode.HTML)
        try: await context.bot.send_message(chat_id=t_user, text=f"🎉 <b>Xushxabar!</b>\n\nSiz yechgan <b>{amt} GWT</b> uchun pul kartangizga o'tkazib berildi!", parse_mode=ParseMode.HTML)
        except: pass
        return

    if data.startswith("rejt_wd_"):
        p = data.split("_")
        t_user, amt = int(p[2]), float(p[3])
        db.system_sell_token(t_user, amt, method="WITHDRAW_REFUND") # GWTni qaytarish
        await q.message.edit_text(q.message.text + "\n\n❌ <b>RAD ETILDI: GWT QAYTARILDI</b>", parse_mode=ParseMode.HTML)
        try: await context.bot.send_message(chat_id=t_user, text=f"⚠️ <b>Sizning so'rovingiz rad etildi.</b>\n\nKarta raqami xato bo'lishi mumkin. <b>{amt} GWT</b> hisobingizga qaytarildi.", parse_mode=ParseMode.HTML)
        except: pass
        return

    if data == "pdf_manual_theme":
        context.user_data[K["mode"]] = "wait_pdf_theme"
        await q.message.edit_text("✍️ Iltimos, test mavzusini yozib yuboring:")
        return

    if data == "pdf_ai_solve":
        pending = context.user_data.get(K["convert"], {}).get("pending_test")
        if not pending:
            await q.answer("Test ma'lumotlari topilmadi.", show_alert=True)
            return

        await q.message.edit_text("⏳ AI testni yechmoqda. Iltimos kuting...")

        q_text = ""
        for i, qu in enumerate(pending["questions"]):
            q_text += f"{i+1}. {qu['question']}\n"
            for j, opt in enumerate(qu['options']):
                q_text += f"{chr(97+j)}) {opt}\n"

        ai_keys_text = await solve_test_with_ai(q_text)
        keys = re.findall(r"(\d+)[ \-\.\)]*([a-eA-E])", ai_keys_text)

        if not keys:
            await q.message.edit_text("❌ AI kalitlarni topa olmadi. Iltimos o'zingiz qo'lda kiriting.")
            context.user_data[K["mode"]] = "wait_pdf_keys"
            return

        for k in keys:
            q_idx = int(k[0]) - 1
            ans_letter = k[1].lower()
            if 0 <= q_idx < len(pending["questions"]):
                pending["questions"][q_idx]["correct_index"] = ord(ans_letter) - ord('a')

        missing = sum(1 for qu in pending["questions"] if qu.get("correct_index") is None or qu.get("correct_index") < 0)
        if missing > 0:
            await q.message.edit_text(f"⚠️ AI {missing} ta savolga kalit topa olmadi. Iltimos, o'zingiz to'liq yuboring:\nFormat: 1A 2B 3C...")
            context.user_data[K["mode"]] = "wait_pdf_keys"
            return

        await q.message.edit_text("✅ AI barcha kalitlarni topdi!")
        await build_and_send_test(update, context, pending["theme"], pending["questions"], pending["target_chat"])
        return

    if data == "pdf_manual_keys":
        context.user_data[K["mode"]] = "wait_pdf_keys"
        await q.message.edit_text("✍️ Iltimos, to'g'ri javoblarni quyidagi formatda yuboring:\n\n1A 2B 3C 4D...\n(yoki 1a 2b 3c...)")
        return

    try:
        await q.answer()
    except Exception:
        pass

    if data.startswith("bot_solve:"):
        test_id = data.split(":")[1]

        test = db.get_test(test_id)
        if not test:
            await q.message.edit_text("❌ Test topilmadi.")
            return

        has_price = float(dict(test).get("price_gwt") or 0) > 0 or int(dict(test).get("price_stars") or 0) > 0
        is_owner = int(dict(test).get("owner_user_id", 0)) == user_id
        is_admin = user_id in SUPERADMINS

        if has_price and not is_owner and not is_admin:
            with db._conn() as c:
                purchased = c.execute("SELECT 1 FROM purchased_tests WHERE user_id=%s AND test_id=%s", (user_id, test_id)).fetchone()
            if not purchased:
                await q.answer("⛔ Bu pullik test! Avval sayt orqali sotib oling.", show_alert=True)
                return

        await start_test_private(update, context, test_id)
        return

    if data.startswith("force_finish_"):
        test_id = data.split("_", 2)[2]
        with db._conn() as c:
            running = c.execute("SELECT session_id FROM sessions WHERE user_id=%s AND test_id=%s AND state='running' LIMIT 1", (user_id, test_id)).fetchone()
        if running:
            await finish_test_session(context, running['session_id'])
            try:
                await q.message.delete()
            except: pass
        else:
            await q.answer("Sessiya topilmadi yoki allaqachon tugatilgan.", show_alert=True)
        return

    if data == "back_to_main":
        context.user_data.pop(K["mode"], None)
        context.user_data.pop("pending_premium_months", None)
        context.user_data.pop("pending_gwt_amount", None)
        msg, kb = await build_main_menu(user_id, context.bot.username, lang)
        try:
            await q.message.delete()
        except Exception:
            pass
        await context.bot.send_message(chat_id=user_id, text=get_bot_text('main_menu_loaded', lang), parse_mode=ParseMode.HTML, reply_markup=kb)
        return

    if data == "pay_method_card":
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton(f"🥉 {get_bot_text('month_1', lang)} - 20,000 UZS", callback_data="buy_premium_1")],
            [InlineKeyboardButton(f"🥈 {get_bot_text('month_3', lang)} - 55,000 UZS", callback_data="buy_premium_3")],
            [InlineKeyboardButton(f"🥇 {get_bot_text('month_6', lang)} - 110,000 UZS", callback_data="buy_premium_6")],
            [InlineKeyboardButton(f"💎 {get_bot_text('month_12', lang)} - 200,000 UZS", callback_data="buy_premium_12")],
            [InlineKeyboardButton(get_bot_text('btn_back', lang), callback_data="btn_premium")]
        ])
        await q.message.edit_text(get_bot_text('prem_card_msg', lang), reply_markup=kb, parse_mode=ParseMode.HTML)
        return

    if data == "pay_method_stars":
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton(f"🥉 {get_bot_text('month_1', lang)} - 80 ⭐️", callback_data="buy_stars_1")],
            [InlineKeyboardButton(f"🥈 {get_bot_text('month_3', lang)} - 220 ⭐️", callback_data="buy_stars_3")],
            [InlineKeyboardButton(f"🥇 {get_bot_text('month_6', lang)} - 440 ⭐️", callback_data="buy_stars_6")],
            [InlineKeyboardButton(f"💎 {get_bot_text('month_12', lang)} - 800 ⭐️", callback_data="buy_stars_12")],
            [InlineKeyboardButton(get_bot_text('btn_back', lang), callback_data="btn_premium")]
        ])
        await q.message.edit_text(get_bot_text('prem_stars_msg', lang), reply_markup=kb, parse_mode=ParseMode.HTML)
        return

    if data.startswith("buy_premium_"):
        months = int(data.split("_")[2])
        prices_dict = {1: 20000, 3: 55000, 6: 110000, 12: 200000}
        price = f"{prices_dict.get(months, 20000):,}"

        context.user_data['pending_premium_months'] = months

        text = get_bot_text('prem_card_info', lang, price=price, card=ADMIN_CARD)
        kb = InlineKeyboardMarkup([[InlineKeyboardButton(get_bot_text('btn_cancel', lang), callback_data="cancel_action")]])
        await q.message.edit_text(text, parse_mode=ParseMode.HTML, reply_markup=kb)
        return

    if data.startswith("buy_stars_"):
        months = int(data.split("_")[2])
        prices_stars = {1: 80, 3: 220, 6: 440, 12: 800}
        stars_amount = prices_stars.get(months, 80)

        title = f"Premium maqom - {months} oy"
        description = f"Geo Ustoz botida {months} oylik Premium maqomini xarid qilish."
        payload = f"premium_{months}"
        currency = "XTR"
        prices = [LabeledPrice(title, stars_amount)]

        try:
            await context.bot.send_invoice(
                chat_id=user_id, title=title, description=description, payload=payload,
                provider_token="", currency=currency, prices=prices
            )
        except Exception as e:
            logging.error(f"Telegram Stars invoice xatosi: {e}")
            await q.message.reply_text("❌ Telegram Yulduzlari orqali to'lov yuborishda xatolik yuz berdi.")

        return

    # ==========================================
    # 👑 TASDIQLASH VA RAD ETISH TUGMALARI
    # ==========================================
    if data.startswith("approve_req_"):
        parts = data.split("_")
        req_id = int(parts[2])
        req_type = parts[3]
        target_user = int(parts[4])
        val = float(parts[5])

        with db._conn() as c:
            req = c.execute("SELECT * FROM premium_requests WHERE id=%s", (req_id,)).fetchone()

        if not req or req.get('status') != 'pending':
            await q.answer("Bu so'rov allaqachon ko'rib chiqilgan!", show_alert=True)
            await q.message.edit_reply_markup(reply_markup=None)
            return

        with db._conn() as c:
            c.execute("UPDATE premium_requests SET status='approved' WHERE id=%s", (req_id,))

        if req_type == "gwt":
            db.system_sell_token(target_user, val, method="CARD")
            try:
                await context.bot.send_message(chat_id=target_user, text=f"🎉 <b>Tabriklaymiz!</b> To'lovingiz tasdiqlandi.\n\nSizning hamyoningizga <b>{val} GWT</b> tashlab berildi! 💰", parse_mode="HTML")
            except: pass
        elif req_type == "prem":
            db.add_premium_months(target_user, int(val))
            try:
                await context.bot.send_message(chat_id=target_user, text=f"🎉 <b>Tabriklaymiz!</b> To'lovingiz admin tomonidan tasdiqlandi.\n\nSizga <b>{int(val)} oylik PREMIUM</b> maqomi berildi! 💎", parse_mode="HTML")
            except: pass

        # Update all admin messages
        status_text = "✅ <b>TASDIQLANGAN VA BERILDI</b>"
        admin_msg_ids = req.get("admin_msg_ids", "")
        new_caption = (q.message.caption or "") + f"\n\n📝 Holat: {status_text}\n👨‍💻 Admin: {update.effective_user.first_name}"

        if admin_msg_ids:
            for pair in admin_msg_ids.split(","):
                if ":" in pair:
                    c_id, m_id = pair.split(":")
                    try:
                        await context.bot.edit_message_caption(chat_id=c_id, message_id=m_id, caption=new_caption, reply_markup=None, parse_mode=ParseMode.HTML)
                    except Exception:
                        pass
        else:
            try:
                await q.edit_message_caption(caption=new_caption, reply_markup=None, parse_mode=ParseMode.HTML)
            except:
                pass
        await q.answer("Muvaffaqiyatli tasdiqlandi!")
        return

    if data.startswith("reject_req_"):
        parts = data.split("_")
        req_id = int(parts[2])
        target_user = int(parts[3])

        with db._conn() as c:
            req = c.execute("SELECT * FROM premium_requests WHERE id=%s", (req_id,)).fetchone()

        if not req or req.get('status') != 'pending':
            await q.answer("Bu so'rov allaqachon ko'rib chiqilgan!", show_alert=True)
            await q.message.edit_reply_markup(reply_markup=None)
            return

        with db._conn() as c:
            c.execute("UPDATE premium_requests SET status='rejected' WHERE id=%s", (req_id,))

        try:
            await context.bot.send_message(chat_id=target_user, text="❌ <b>To'lovingiz tasdiqlanmadi!</b>\n\nKiritilgan chek xato yoki to'lov o'tmagan. Iltimos adminga murojaat qiling.", parse_mode="HTML")
        except: pass

        status_text = "❌ <b>RAD ETILDI</b>"
        admin_msg_ids = req.get("admin_msg_ids", "")
        new_caption = (q.message.caption or "") + f"\n\n📝 Holat: {status_text}\n👨‍💻 Admin: {update.effective_user.first_name}"

        if admin_msg_ids:
            for pair in admin_msg_ids.split(","):
                if ":" in pair:
                    c_id, m_id = pair.split(":")
                    try:
                        await context.bot.edit_message_caption(chat_id=c_id, message_id=m_id, caption=new_caption, reply_markup=None, parse_mode=ParseMode.HTML)
                    except Exception:
                        pass
        else:
            try:
                await q.edit_message_caption(caption=new_caption, reply_markup=None, parse_mode=ParseMode.HTML)
            except:
                pass
        await q.answer("Rad etildi!")
        return

    with db._conn() as c:
        running = c.execute(
            "SELECT session_id, test_id FROM sessions WHERE user_id=%s AND state='running' ORDER BY started_at DESC LIMIT 1",
            (user_id,)
        ).fetchone()

    if running:
        if data.startswith("start_public_test:") or data.startswith("start_private_test:"):
            try:
                await update.effective_chat.send_message(get_bot_text('test_running_err', lang))
            except Exception:
                pass
            return

    if data.startswith("ans:"):
        parts = data.split(":")
        if len(parts) != 4: return
        session_id = parts[1]
        q_index = int(parts[2])
        opt_index = int(parts[3])
        await handle_inline_answer(update, context, session_id, q_index, opt_index)
        return

    if data == "bc_toggle_all":
        b = context.user_data.get(K["broadcast"])
        if not b or b.get("stage") != "select_channels": return
        all_selected = all(b["selected"].values())
        new_val = not all_selected
        for cid in b["selected"]:
            b["selected"][cid] = new_val
        await show_channel_selection(update.effective_chat.id, context, q.message.message_id)
        return

    if data.startswith("bc_toggle:"):
        b = context.user_data.get(K["broadcast"])
        if not b or b.get("stage") != "select_channels": return
        cid = int(data.split(":")[1])
        if cid in b["selected"]:
            b["selected"][cid] = not b["selected"][cid]
            await show_channel_selection(update.effective_chat.id, context, q.message.message_id)
        return

    if data == "bc_send":
        b = context.user_data.get(K["broadcast"])
        if not b or b.get("stage") != "select_channels": return

        selected_cids = [cid for cid, is_sel in b["selected"].items() if is_sel]
        if not selected_cids:
            try:
                await q.answer("Kamida 1 ta kanalni tanlang!", show_alert=True)
            except: pass
            return

        await q.message.edit_text(f"📣 Yuborish boshlandi. Chatlar: {len(selected_cids)} ta...\n⏳ Iltimos kuting...")

        reply_text = b["button_text"]
        msg_id = b["msg_id"]
        src_chat_id = update.effective_chat.id

        ad_id = uuid.uuid4().hex[:8]
        db.create_ad(ad_id, user_id, reply_text)

        ok, fail, reach = 0, 0, 0
        err_reasons = []

        if reply_text and (reply_text.startswith("http://") or reply_text.startswith("https://")):
            markup = InlineKeyboardMarkup([[InlineKeyboardButton("🔗 Batafsil", url=reply_text)]])
        elif reply_text:
            markup = InlineKeyboardMarkup([[InlineKeyboardButton("🔗 Batafsil", callback_data=f"ad:{ad_id}")]])
        else:
            markup = None

        for chat_id in selected_cids:
            try:
                await context.bot.copy_message(chat_id=chat_id, from_chat_id=src_chat_id, message_id=msg_id, reply_markup=markup)
                ok += 1
                try:
                    count = await context.bot.get_chat_member_count(chat_id)
                    reach += count
                except Exception:
                    pass
            except Exception as e:
                logging.error(f"Broadcast xatosi ({chat_id}): {e}")
                fail += 1
                err_msg = str(e)
                if "forbidden" in err_msg.lower() or "not found" in err_msg.lower() or "kicked" in err_msg.lower():
                    db.set_bot_admin(chat_id, 0, now_ts())

                if err_msg not in err_reasons:
                    err_reasons.append(err_msg)
            await asyncio.sleep(0.05)

        context.user_data.pop(K["broadcast"], None)
        context.user_data.pop(K["mode"], None)

        err_text = ""
        if err_reasons:
            err_text = f"\n\n⚠️ <b>Xato sababi:</b> {html.escape(err_reasons[0])}"
            if len(err_reasons) > 1: err_text += " (va boshqa xatolar)"
            err_text += "\n<i>(Tizim xatoli kanallarni avtomatik tozalamoqda.)</i>"

        kb = InlineKeyboardMarkup([[InlineKeyboardButton(get_bot_text('btn_back_main', lang), callback_data="back_to_main")]])
        await q.message.edit_text(
            f"✅ Tarqatish tugadi.\n\n✅ Muvaffaqiyatli: {ok} ta\n❌ Xatolik: {fail} ta\n👁 Umumiy qamrov: ~{reach} kishi{err_text}",
            parse_mode=ParseMode.HTML, reply_markup=kb
        )
        return

    create = context.user_data.get(K["create"])

    if data == "add_question" and create:
        await prompt_question(update, context)
        return

    if data == "finish_questions" and create:
        await ask_deadline(update, context)
        return

    if data.startswith("set_correct:") and create:
        idx = int(data.split(":", 1)[1])
        cur = create.get("current")
        if not cur or not cur.get("options"):
            await q.message.edit_text("Variantlar yo‘q.")
            return

        cur["correct_index"] = idx
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("➕ Yana savol", callback_data="add_question")],
            [InlineKeyboardButton("✅ Yakunlash", callback_data="finish_questions")],
        ])
        await q.message.edit_text(f"✅ To‘g‘ri javob: {idx+1}-variant tanlandi.\n\nTanlang:", reply_markup=kb)
        return

    if data == "deadline_yes" and create:
        create["stage"] = "ask_deadline_value"
        await q.message.edit_text("⏰ Deadline vaqtini yuboring:\nMasalan: 03/03/2026 22:00")
        return

    if data == "deadline_no" and create:
        create["deadline_ts"] = None
        await ask_private_button(update, context)
        return

    if data == "confirm_publish" and create:
        await publish_to_chat(update, context)
        return

    if data in ("cancel_action", "cancel_premium"):
        context.user_data.pop("delete_test", None)
        context.user_data.pop("publish", None)
        context.user_data.pop(K["create"], None)
        context.user_data.pop(K["convert"], None)
        context.user_data.pop(K["mode"], None)
        context.user_data.pop("pending_premium_months", None)
        context.user_data.pop("pending_gwt_amount", None)
        await q.message.delete()
        return

    if data.startswith("manage_test:"):
        test_id = data.split(":")[1]
        test = db.get_test(test_id)
        if not test or (int(test["owner_user_id"]) != user_id and not is_superadmin):
            await q.message.edit_text("❌ Test topilmadi yoki ruxsat yo'q.")
            return

        title = test["title"]
        text = get_bot_text('test_manage_msg', lang, title=html.escape(title))

        kb = [
            [InlineKeyboardButton(get_bot_text('btn_export_excel', lang), callback_data=f"export_test:{test_id}")],
            [InlineKeyboardButton(get_bot_text('btn_close_test', lang), callback_data=f"close_test:{test_id}")],
        ]
        if not test["public_name"]:
            kb.append([InlineKeyboardButton(get_bot_text('btn_publish_test', lang), callback_data=f"publish_test:{test_id}")])

        kb.append([InlineKeyboardButton(get_bot_text('btn_delete_test', lang), callback_data=f"delete_test:{test_id}")])
        kb.append([InlineKeyboardButton(get_bot_text('btn_back', lang), callback_data="results_menu")])

        await q.message.edit_text(text, parse_mode=ParseMode.HTML, reply_markup=InlineKeyboardMarkup(kb))
        return

    if data.startswith("publish_test:"):
        test_id = data.split(":", 1)[1]
        test = db.get_test(test_id)
        if not test or (int(test.get("owner_user_id", 0)) != user_id and not is_superadmin):
            await q.message.edit_text("❌ Test topilmadi yoki ruxsat yo'q.")
            return

        if dict(test).get("public_name"):
            await q.message.edit_text("⚠️ Bu test allaqachon public qilingan.")
            return

        context.user_data[K["mode"]] = "publish"
        context.user_data["publish"] = {"test_id": test_id, "stage": "ask_public_name"}
        kb = InlineKeyboardMarkup([[InlineKeyboardButton(get_bot_text('btn_cancel', lang), callback_data=f"manage_test:{test_id}")]])
        await q.message.edit_text("🌐 Testni public qilish.\n\nTest uchun maxsus nom kiriting (masalan: @kimyo_2026):", reply_markup=kb)
        return

    if data == "publish_no_password":
        pub = context.user_data.get("publish")
        if not pub: return
        test_id = pub["test_id"]
        public_name = pub["public_name"]

        try:
            success = db.set_public_link(test_id, public_name, None)
        except Exception as e:
            logging.error(f"DB xatosi: {e}")
            success = False

        kb = InlineKeyboardMarkup([[InlineKeyboardButton(get_bot_text('btn_back', lang), callback_data=f"manage_test:{test_id}")]])

        if success:
            await q.message.edit_text(f"✅ Test public qilindi: {public_name}", reply_markup=kb)
            context.user_data.pop("publish", None)
            context.user_data.pop(K["mode"], None)

            try:
                test_data = dict(db.get_test(test_id) or {})
                qcount, _ = db.stats(test_id)
                asyncio.create_task(
                    announce_to_channel_with_ai(context, test_data.get("title", "Umumiy test"), qcount, public_name)
                )
            except Exception as e:
                logging.error(f"AI e'lon xatosi: {e}")

        else:
            pub["stage"] = "ask_public_name"
            await q.message.edit_text(f"❌ Bu nom band: {public_name}\n\nIltimos, boshqa nom yuboring (masalan: @yangi_test):", reply_markup=kb)
        return

    if data == "publish_set_password":
        pub = context.user_data.get("publish")
        if not pub: return
        pub["stage"] = "ask_password_value"
        kb = InlineKeyboardMarkup([[InlineKeyboardButton(get_bot_text('btn_back', lang), callback_data=f"manage_test:{pub['test_id']}")]])
        await q.message.edit_text("🔒 Parolni kiriting:", reply_markup=kb)
        return

    if data.startswith("delete_test:"):
        test_id = data.split(":", 1)[1]
        test = db.get_test(test_id)

        if not test or (int(test.get("owner_user_id", 0)) != user_id and not is_superadmin):
            await q.message.edit_text("❌ Siz bu testni o'chirish huquqiga ega emassiz.")
            return

        if dict(test).get("manage_password") and not is_superadmin:
            context.user_data[K["mode"]] = "delete_test_password"
            context.user_data["delete_test"] = {"test_id": test_id}
            kb = InlineKeyboardMarkup([[InlineKeyboardButton(get_bot_text('btn_cancel', lang), callback_data=f"manage_test:{test_id}")]])
            await q.message.edit_text(f"🗑 <b>{h(dict(test).get('title'))}</b> testini o'chirish uchun BOSHQARUV parolini kiriting:", reply_markup=kb, parse_mode=ParseMode.HTML)
        else:
            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton(get_bot_text('btn_yes', lang), callback_data=f"confirm_delete_test:{test_id}")],
                [InlineKeyboardButton(get_bot_text('btn_cancel', lang), callback_data=f"manage_test:{test_id}")],
            ])
            msg_prefix = "👑 ADMIN: " if is_superadmin and int(test.get("owner_user_id", 0)) != user_id else ""
            await q.message.edit_text(f"{msg_prefix}🗑 Rostdan ham <b>{h(dict(test).get('title'))}</b> testini va uning barcha natijalarini butunlay o'chirmoqchimisiz?", reply_markup=kb, parse_mode=ParseMode.HTML)
        return

    if data.startswith("confirm_delete_test:"):
        test_id = data.split(":", 1)[1]
        test = db.get_test(test_id)
        if not test or (int(test.get("owner_user_id", 0)) != user_id and not is_superadmin):
            await q.message.edit_text("❌ Ruxsat yo'q yoki test topilmadi.")
            return
        success = db.delete_test(test_id, user_id)
        kb = InlineKeyboardMarkup([[InlineKeyboardButton(get_bot_text('btn_back_main', lang), callback_data="back_to_main")]])
        if success:
            await q.message.edit_text("✅ Test va uning natijalari butunlay o'chirildi.", reply_markup=kb)
        else:
            await q.message.edit_text("❌ O'chirishda xatolik yuz berdi.", reply_markup=kb)
        return

    if data.startswith("start_public_test:"):
        test_id = data.split(":", 1)[1]
        test = db.get_test(test_id)

        if not test or not dict(test).get("public_name"):
            await q.message.edit_text("❌ Test topilmadi yoki yopilgan.")
            return

        # --- PULLIK TEST TEKSHIRUVI ---
        has_price = float(dict(test).get("price_gwt") or 0) > 0 or int(dict(test).get("price_stars") or 0) > 0
        is_owner = int(dict(test).get("owner_user_id", 0)) == user_id
        is_admin = user_id in SUPERADMINS

        if has_price and not is_owner and not is_admin:
            with db._conn() as c:
                purchased = c.execute("SELECT 1 FROM purchased_tests WHERE user_id=%s AND test_id=%s", (user_id, test_id)).fetchone()

            if not purchased:
                token = db.get_or_create_user_api_key(user_id)
                webapp_url = f"{WEB_BASE_URL.rstrip('/')}/solve/{test_id}?token={token}"
                kb = InlineKeyboardMarkup([[InlineKeyboardButton("🛒 Sotib olish va ishlash", web_app=WebAppInfo(url=webapp_url))]])
                await q.message.edit_text("⛔ <b>Bu pullik test!</b>\n\nUni ishlash uchun quyidagi tugma orqali bozorga kiring va xarid qiling:", reply_markup=kb, parse_mode=ParseMode.HTML)
                return
        # ------------------------------

        if dict(test).get("password") and not is_superadmin:
            context.user_data[K["mode"]] = "enter_password"
            context.user_data["enter_password"] = {
                "test_id": test_id,
                "msg_id": q.message.message_id
            }
            await q.message.edit_text("🔒 Bu test parolli. Iltimos, parolni kiriting:")
            return

        try:
            await q.message.delete()
        except:
            pass

        await start_public_test(update, context, test_id)
        return

    if data.startswith("export_test:"):
        test_id = data.split(":", 1)[1]
        await export_excel(update, context, test_id)
        return

    if data.startswith("close_test:"):
        test_id = data.split(":", 1)[1]
        test = db.get_test(test_id)
        if not test:
            await q.message.edit_text("Test topilmadi.")
            return
        if int(test["owner_user_id"]) != user_id and not is_superadmin:
            await q.message.edit_text("❌ Siz bu testni yopishga ruxsatga ega emassiz.")
            return
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("📢 Ha, e’lon qilinsin", callback_data=f"close_announce_yes:{test_id}")],
            [InlineKeyboardButton("🔇 Yo‘q, faqat yopish", callback_data=f"close_announce_no:{test_id}")],
            [InlineKeyboardButton(get_bot_text('btn_back', lang), callback_data=f"manage_test:{test_id}")]
        ])
        await q.message.edit_text("⛔ Testni yopmoqchimisiz?\nNatijani chatga e’lon qilinsinmi?", reply_markup=kb)
        return

    if data.startswith("close_announce_yes:"):
        test_id = data.split(":", 1)[1]
        await finalize_test(update, context, test_id, manual=True, announce=True)
        kb = InlineKeyboardMarkup([[InlineKeyboardButton(get_bot_text('btn_back', lang), callback_data=f"manage_test:{test_id}")]])
        await q.message.edit_text("✅ Test yopildi va natija chatga joylandi.", reply_markup=kb)
        return

    if data.startswith("close_announce_no:"):
        test_id = data.split(":", 1)[1]
        await finalize_test(update, context, test_id, manual=True, announce=False)
        kb = InlineKeyboardMarkup([[InlineKeyboardButton(get_bot_text('btn_back', lang), callback_data=f"manage_test:{test_id}")]])
        await q.message.edit_text("✅ Test yopildi (natija e’lon qilinmadi).", reply_markup=kb)
        return

    if data.startswith("create_chat:") or data == "create_private":
        chat_id = int(data.split(":", 1)[1]) if ":" in data else None
        context.user_data[K["mode"]] = "creating"
        context.user_data[K["create"]] = {
            "stage": "ask_title",
            "chat_id": chat_id,
            "title": None,
            "per_q_sec": 60,
            "questions": [],
            "current": None,
            "test_id": None,
            "deadline_ts": None
        }
        await q.message.edit_text("📝 Test nomini yuboring:")
        return

    if data.startswith("convert_chat:") or data == "convert_private":
        chat_id = int(data.split(":", 1)[1]) if ":" in data else None
        context.user_data[K["mode"]] = "convert"
        context.user_data[K["convert"]] = {"chat_id": chat_id}
        await q.message.edit_text(
            get_bot_text('convert_where', lang),
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("✅ Tanlandi, davom etish", callback_data="none")]
            ])
        )

        await update.effective_chat.send_message(
            "✅ Manzil tanlandi!\n\nEndi <b>Word (.docx)</b> faylini yuboring.\n\n"
            "Fayl ichidagi format xuddi shunday bo'lishi kerak:\n\n"
            "<code>theme:Mavzu nomi\n\n"
            "1. Savol matni\n"
            "[shu joyda ixtiyoriy rasm bo'lishi mumkin]\n"
            "a) Variant 1\n"
            "b) Variant 2\n"
            "true:a\n\n"
            "2. Rasm qatnashmagan oddiy savol\n"
            "a) O'zbekiston\n"
            "b) Tojikiston\n"
            "true:a</code>",
            parse_mode=ParseMode.HTML
        )
        return

async def on_callback_postprocess(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    if not q: return
    data = q.data or ""
    create = context.user_data.get(K["create"])
    if not create: return

    if data.startswith("set_correct:"):
        cur = create.get("current")
        if cur and cur.get("question") and cur.get("options") and cur.get("correct_index") is not None:
            create["questions"].append(cur)
            create["current"] = None
            create["stage"] = "idle_questions"

# ==========================================
# 9. MEDIA HANDLER (Rasm, Video, To'lov cheki)
# ==========================================
async def on_media(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_chat.type != "private":
        return

    await upsert_user_from_update(update)
    mode = context.user_data.get(K["mode"])
    user_id = update.effective_user.id
    is_superadmin = user_id in SUPERADMINS
    is_lower_admin = user_id in LOWER_ADMINS
    lang = get_user_lang(user_id)

    user_row = dict(db.get_user(user_id) or {})
    is_verified = user_row.get("is_verified", 1)

    if not is_verified:
        try:
            del_msg = await update.effective_chat.send_message("🔄", reply_markup=ReplyKeyboardRemove())
            await del_msg.delete()
        except:
            pass

        captcha_url = f"{WEB_BASE_URL.rstrip('/')}/captcha"
        markup = InlineKeyboardMarkup([[InlineKeyboardButton("🤖 Men robot emasman", web_app=WebAppInfo(url=captcha_url))]])
        await update.effective_chat.send_message(
            get_bot_text('captcha_msg', lang),
            reply_markup=markup,
            parse_mode=ParseMode.HTML
        )
        return

    if mode == "creating":
        create = context.user_data.get(K["create"])
        if create and create.get("stage") == "ask_question" and update.message.photo:
            caption = update.message.caption
            if not caption:
                await update.effective_chat.send_message("❌ Iltimos, rasm yuborganda uning tagiga savol matnini ham (caption qilib) yozib yuboring.")
                return

            create["current"]["question"] = caption
            create["current"]["photo_id"] = update.message.photo[-1].file_id
            create["stage"] = "ask_option"
            await update.effective_chat.send_message("✅ Rasm va savol qabul qilindi!\n\nEndi variantlarni alohida yuboring. Tugatganda `done` deb yozing.", parse_mode=ParseMode.HTML)
            return

    if mode == "ai_chat" and update.message.photo:
        user_row = db.get_user(user_id)
        if not (user_row and dict(user_row).get("status") == "premium") and not is_superadmin and not is_lower_admin:
            today_str = datetime.now(tz=TZ).strftime("%Y-%m-%d")
            if db.get_ai_usage(user_id, today_str) >= 10:
                context.user_data.pop(K["mode"], None)
                msg, kb = await build_main_menu(user_id, context.bot.username, lang)
                await update.effective_chat.send_message(get_bot_text('ai_limit', lang), reply_markup=kb)
                return
            db.increment_ai_usage(user_id, today_str)

        wait_msg = await update.effective_chat.send_message(get_bot_text('ai_pic_thinking', lang), parse_mode=ParseMode.HTML)

        try:
            photo_file = await update.message.photo[-1].get_file()
            photo_bytes = await photo_file.download_as_bytearray()
            base64_image = base64.b64encode(photo_bytes).decode('utf-8')
        except Exception as e:
            logging.error(f"Rasm yuklashda xato: {e}")
            await wait_msg.edit_text("❌ Rasmni o'qishda xatolik yuz berdi. Boshqa rasm yuborib ko'ring.")
            return

        text = update.message.caption or "Bu rasmda nima tasvirlanganini batafsil tushuntirib ber."
        await process_ai_message(update, context, user_id, text, wait_msg, base64_image)
        return

    if mode == "broadcast":
        if not is_superadmin and not is_lower_admin: return
        b = context.user_data.get(K["broadcast"]) or {}
        if b.get("stage") == "wait_content":
            b["msg_id"] = update.message.message_id
            b["stage"] = "wait_button_text"
            await update.effective_chat.send_message(
                "Juda soz! 🎯\n\nEndi tugma bosilganda nima chiqishini yuboring:\n\n"
                "1️⃣ <b>Matn (Tugma):</b> Ixtiyoriy matn yozsangiz, ekranda sakrab (Pop-up) chiqadi va kliklar hisoblanadi.\n"
                "2️⃣ <b>Link:</b> <code>https://</code> bilan boshlanadigan link yuborsangiz o'sha saytni ochadi.\n\n"
                "<i>Agar tugma umuman kerak bo'lmasa, shunchaki 'yoq' deb yozing.</i>",
                parse_mode=ParseMode.HTML
            )
        return

    # ==========================================
    # TO'LOV CHEKI (PREMIUM VA GWT UCHUN) YUBORISH
    # ==========================================
    pending_months = context.user_data.get('pending_premium_months')
    if pending_months and update.message.photo:
        photo_id = update.message.photo[-1].file_id
        user = update.effective_user

        photo_id_str = f"BOT_{pending_months}_{photo_id}"
        req_id = 0
        try:
            with db._conn() as c:
                c.execute("""
                    INSERT INTO premium_requests(user_id, photo_id, status, created_at)
                    VALUES(%s, %s, 'pending', %s)
                """, (user.id, photo_id_str, now_ts()))
                req_id = c.lastrowid
        except Exception as db_err:
            logging.error(f"Bazaga yozishda xato: {db_err}")

        caption = (
            f"🆕 <b>YANGI TO'LOV CHEKI!</b>\n\n"
            f"👤 Kimdan: {user.first_name} (@{user.username if user.username else 'yoq'})\n"
            f"🆔 ID: <code>{user.id}</code>\n"
            f"📦 Ta'rif: <b>{pending_months} oylik Premium</b>"
        )

        kb = InlineKeyboardMarkup([
            [
                InlineKeyboardButton("✅ Tasdiqlash", callback_data=f"approve_req_{req_id}_prem_{user.id}_{pending_months}"),
                InlineKeyboardButton("❌ Rad etish", callback_data=f"reject_req_{req_id}_{user.id}")
            ]
        ])

        admin_msgs = []
        for adm in SUPERADMINS:
            try:
                sent_msg = await context.bot.send_photo(chat_id=adm, photo=photo_id, caption=caption, reply_markup=kb, parse_mode=ParseMode.HTML)
                admin_msgs.append(f"{adm}:{sent_msg.message_id}")
            except Exception as e:
                logging.error(f"Adminga chek yuborishda xato: {e}")

        if admin_msgs:
            admin_msg_ids_str = ",".join(admin_msgs)
            with db._conn() as c:
                c.execute("UPDATE premium_requests SET admin_msg_ids=%s WHERE id=%s", (admin_msg_ids_str, req_id))
            await update.effective_chat.send_message("✅ Chekingiz adminga yuborildi. Iltimos, tasdiqlanishini kuting (Odatda 5-10 daqiqa).")
        else:
            await update.effective_chat.send_message("❌ Tizimda admin topilmadi.")

        context.user_data.pop('pending_premium_months', None)
        return

    pending_gwt = context.user_data.get('pending_gwt_amount')
    if pending_gwt and update.message.photo:
        photo_id = update.message.photo[-1].file_id
        user = update.effective_user

        photo_id_str = f"GWT_{pending_gwt}_{photo_id}"
        req_id = 0
        try:
            with db._conn() as c:
                c.execute("""
                    INSERT INTO premium_requests(user_id, photo_id, status, created_at)
                    VALUES(%s, %s, 'pending', %s)
                """, (user.id, photo_id_str, now_ts()))
                req_id = c.lastrowid
        except Exception as db_err:
            logging.error(f"GWT Bazaga yozishda xato: {db_err}")

        caption = f"🆕 <b>YANGI GWT TO'LOV CHEKI!</b>\n\n👤 Kimdan: {user.first_name} (@{user.username})\n🆔 ID: <code>{user.id}</code>\n📦 Miqdor: <b>{pending_gwt} GWT</b>"

        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("✅ Tasdiqlash", callback_data=f"approve_req_{req_id}_gwt_{user.id}_{pending_gwt}"),
             InlineKeyboardButton("❌ Rad etish", callback_data=f"reject_req_{req_id}_{user.id}")]
        ])

        admin_msgs = []
        for adm in SUPERADMINS:
            try:
                sent_msg = await context.bot.send_photo(chat_id=adm, photo=photo_id, caption=caption, reply_markup=kb, parse_mode=ParseMode.HTML)
                admin_msgs.append(f"{adm}:{sent_msg.message_id}")
            except Exception as e:
                logging.error(f"Adminga GWT chekini yuborishda xato: {e}")

        if admin_msgs:
            admin_msg_ids_str = ",".join(admin_msgs)
            with db._conn() as c:
                c.execute("UPDATE premium_requests SET admin_msg_ids=%s WHERE id=%s", (admin_msg_ids_str, req_id))
            await update.effective_chat.send_message("✅ Chekingiz adminga yuborildi. Tasdiqlanishini kuting (Odatda 5-10 daqiqa).")
        else:
            await update.effective_chat.send_message("❌ Tizimda admin topilmadi.")

        context.user_data.pop('pending_gwt_amount', None)
        return

# ==========================================
# 10. ASOSIY TEXT HANDLER
# ==========================================
async def on_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_chat.type != "private":
        return

    await upsert_user_from_update(update)
    user_id = update.effective_user.id
    text = (update.message.text or "").strip()
    lang = get_user_lang(user_id)

    # --- Yangi qo'shilgan: PIN kodni elektron pochta yoki maxfiy so'z orqali tiklash jarayoni ---
    if context.user_data.get('reset_pin_mode'):
        mode = context.user_data.get('reset_pin_mode')
        is_valid = False

        # 1. Email orqali kelgan kodni tekshirish
        if mode == 'email':
            expected_code = context.user_data.get('reset_pin_code')
            if text == expected_code:
                is_valid = True
            else:
                kb = InlineKeyboardMarkup([[InlineKeyboardButton("❌ Bekor qilish", callback_data="cancel_reset_pin")]])
                await update.effective_chat.send_message("❌ Noto'g'ri kod! Pochtani qayta tekshiring yoki bekor qiling.", reply_markup=kb)
                return

        # 2. Maxfiy so'zni tekshirish
        elif mode == 'secret':
            with db._conn() as c:
                row = c.execute("SELECT secret_word FROM users WHERE user_id=%s", (user_id,)).fetchone()
            secret_word = row.get("secret_word") if row else None

            if secret_word and text.lower() == secret_word.lower():
                is_valid = True
            else:
                kb = InlineKeyboardMarkup([[InlineKeyboardButton("❌ Bekor qilish", callback_data="cancel_reset_pin")]])
                await update.effective_chat.send_message("❌ Maxfiy so'z noto'g'ri! Qaytadan urinib ko'ring yoki bekor qiling.", reply_markup=kb)
                return

        # Agar parollar to'g'ri bo'lsa (Email yoki Maxfiy so'z orqali)
        if is_valid:
            with db._conn() as c:
                c.execute("UPDATE users SET pin_code=NULL WHERE user_id=%s", (user_id,))
                c.execute("COMMIT")

            context.user_data['is_locked'] = False
            context.user_data.pop('reset_pin_mode', None)
            context.user_data.pop('reset_pin_code', None)
            context.user_data.pop('current_pin_input', None)

            await update.effective_chat.send_message(
                "✅ <b>Qoyil! PIN kodingiz bekor qilindi va tizim ochildi.</b>\nKabinetdan yangi PIN o'rnatishingiz mumkin.",
                parse_mode=ParseMode.HTML
            )

            msg_text, kb_reply = await build_main_menu(user_id, context.bot.username, lang)
            await context.bot.send_message(chat_id=user_id, text=msg_text, reply_markup=kb_reply, parse_mode=ParseMode.HTML)
        return
    # ------------------------------------------------------------------------
    # ------------------------------------------------------------------------

    user_row = dict(db.get_user(user_id) or {})

    create = context.user_data.get(K["create"])
    mode = context.user_data.get(K["mode"])

    pending = context.application.bot_data.get(RUNTIME["pending_ready"], {})
    if user_id in pending:
        if text.lower() in ("ha", "xa", "haa", "yes", "tayyor", "ha.", "да"):
            test_id = pending.pop(user_id)
            await start_test_private(update, context, test_id)
        else:
            await update.effective_chat.send_message("Boshlash uchun <b>ha</b> deb yozing.", parse_mode=ParseMode.HTML)
        return

    # --- TESTNI TUGATISH KNOPKASINI TEKSHIRISH ---
    if text in get_all_localized_buttons("btn_finish_test_session"):
        with db._conn() as c:
            running = c.execute("SELECT session_id, test_id FROM sessions WHERE user_id=%s AND state='running' ORDER BY started_at DESC LIMIT 1", (user_id,)).fetchone()

        if running:
            await finish_test_session(context, running['session_id'])
        else:
            msg, kb = await build_main_menu(user_id, context.bot.username, lang)
            await update.effective_chat.send_message(get_bot_text('main_menu_loaded', lang), reply_markup=kb)
        return

    all_btns = []
    for k in ["btn_cabinet", "btn_account", "btn_wallet", "btn_ai", "btn_search", "btn_premium", "btn_create_manual", "btn_create_word", "btn_results", "btn_check_chats", "btn_add_bot", "btn_lock", "btn_finish_test_session", "btn_referral", "btn_top"]:
        all_btns.extend(get_all_localized_buttons(k))

    admin_btns = ["📣 Reklama Tarqatish", "📊 Mening reklamalarim", "📈 Statistika", "📋 Kanallar ro'yxati"]
    all_btns.extend(admin_btns)
    all_btns.extend(["👨‍💻 Adminga murojaat", "🚪 Chatdan chiqish"])

    if text in all_btns:
        context.user_data.pop(K["mode"], None)
        context.user_data.pop("pending_premium_months", None)
        context.user_data.pop("pending_gwt_amount", None)
        context.user_data.pop(K["create"], None)
        context.user_data.pop(K["convert"], None)
        context.user_data.pop(K["broadcast"], None)

    if text in get_all_localized_buttons("btn_cabinet"):
        webapp_url = f"{WEB_BASE_URL.rstrip('/')}/telegram-login"
        kb = InlineKeyboardMarkup([[InlineKeyboardButton("🌐 Tizimga kirish", web_app=WebAppInfo(url=webapp_url))]])
        await update.effective_chat.send_message("🖥 <b>Shaxsiy kabinet</b>\n\nPastdagi tugmani bosib platformaga kiring:", reply_markup=kb, parse_mode=ParseMode.HTML)
        return

    elif text in get_all_localized_buttons("btn_wallet"):
        try:
            wallet = db.get_wallet(user_id)
        except AttributeError:
            await update.effective_chat.send_message("❌ Tizim bazasida hamyon funksiyalari o'rnatilmagan (get_wallet topilmadi).")
            return

        if not wallet:
            wait_msg = await update.effective_chat.send_message("⏳ Hamyon yaratilmoqda. Blokcheyn kalitlar generatsiya qilinmoqda...")
            pub_hex, enc_priv = crypto_mgr.generate_wallet()
            db.create_wallet(user_id, pub_hex, enc_priv)
            wallet = {"public_key": pub_hex}
            await wait_msg.delete()

        balance = db.get_token_balance(user_id)

        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("➕ GWT Sotib olish", callback_data="buy_gwt_menu")],
            [InlineKeyboardButton("💳 Pul yechish (Naqd qilish)", callback_data="withdraw_gwt")], # <-- YANGI TUGMA
            [InlineKeyboardButton(get_bot_text('btn_transfer', lang), callback_data="wallet_transfer"), InlineKeyboardButton(get_bot_text('btn_receive', lang), callback_data="wallet_receive")]
        ])

        msg = get_bot_text('wallet_info', lang, address=wallet['public_key'], balance=balance)
        await update.effective_chat.send_message(msg, reply_markup=kb, parse_mode=ParseMode.HTML)
        return

    elif text in get_all_localized_buttons("btn_add_bot"):
        bot_username = context.bot.username
        group_url = f"https://t.me/{bot_username}?startgroup=true&admin=change_info+post_messages+edit_messages+delete_messages+invite_users+restrict_members+pin_messages+manage_video_chats"
        channel_url = f"https://t.me/{bot_username}?startchannel=true&admin=change_info+post_messages+edit_messages+delete_messages+invite_users+restrict_members+pin_messages+manage_video_chats"

        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("👥 Guruhga qo'shish", url=group_url)],
            [InlineKeyboardButton("📢 Kanalga qo'shish", url=channel_url)]
        ])
        await update.effective_chat.send_message(
            "🤖 <b>Botni test o'tkaziladigan joyga qo'shish:</b>\n\nPastdagi tugmalardan keraklisini tanlang. Bot testlarni yuborishi va boshqarishi uchun unga <b>ADMIN</b> huquqlarini berishni unutmang!",
            reply_markup=kb, parse_mode=ParseMode.HTML
        )
        return

    # --- GWT PUL YECHISH JARAYONI ---
    if mode == "withdraw_typing_card":
        card_number = text.replace(" ", "")
        if not card_number.isdigit() or len(card_number) != 16:
            await update.effective_chat.send_message("❌ Karta raqami xato! Iltimos, 16 xonali raqamni to'g'ri kiriting:")
            return

        context.user_data['withdraw_card'] = card_number
        context.user_data[K["mode"]] = "withdraw_typing_amount"

        balance = db.get_token_balance(user_id)
        await update.effective_chat.send_message(
            f"✅ Karta raqami qabul qilindi: <code>{card_number}</code>\n\n"
            f"Sizda <b>{balance} GWT</b> bor. Qancha miqdorni yechmoqchisiz?\n"
            f"(Masalan: 0.5 yoki 1)",
            parse_mode=ParseMode.HTML
        )
        return

    if mode == "withdraw_typing_amount":
        try:
            amount = float(text)
            if amount <= 0: raise ValueError
        except ValueError:
            await update.effective_chat.send_message("❌ Noto'g'ri miqdor! Faqat raqam kiriting (masalan: 1.2):")
            return

        balance = db.get_token_balance(user_id)
        if balance < amount:
            await update.effective_chat.send_message(f"❌ Balansingizda yetarli GWT yo'q! (Mavjud: {balance} GWT)")
            return

        card_num = context.user_data.get('withdraw_card')
        context.user_data.pop(K["mode"], None)
        context.user_data.pop('withdraw_card', None)

        # 1. Foydalanuvchi hamyonidan GWTni tizimga (GENESIS) o'tkazamiz
        wallet = db.get_wallet(user_id)
        tx_data = f"{wallet['public_key']}->GENESIS:{amount}".encode('utf-8')
        signature = crypto_mgr.sign_transaction(wallet['encrypted_private_key'], tx_data)

        success, msg_err = db.transfer_token_by_address_or_id(user_id, "GENESIS", amount, signature)

        if success:
            summa_uzs = int(amount * 120000)
            # 2. ADMINLARGA XABAR YUBORISH
            admin_msg = (
                f"💸 <b>YANGI PUL YECHISH SO'ROVI!</b>\n\n"
                f"👤 Foydalanuvchi: {update.effective_user.first_name}\n"
                f"🆔 ID: <code>{user_id}</code>\n"
                f"💳 Karta: <code>{card_num}</code>\n"
                f"🪙 Miqdor: <b>{amount} GWT</b>\n"
                f"💵 To'lanishi kerak: <b>{summa_uzs:,} so'm</b>\n\n"
                f"<i>Pulni o'tkazib bo'lgach, pastdagi tasdiqlash tugmasini bosing.</i>"
            )

            kb_admin = InlineKeyboardMarkup([
                [InlineKeyboardButton("✅ To'landim (Tasdiqlash)", callback_data=f"appr_wd_{user_id}_{amount}")],
                [InlineKeyboardButton("❌ Rad etish (Qaytarish)", callback_data=f"rejt_wd_{user_id}_{amount}")]
            ])

            for adm in SUPERADMINS:
                try: await context.bot.send_message(chat_id=adm, text=admin_msg, reply_markup=kb_admin, parse_mode=ParseMode.HTML)
                except: pass

            await update.effective_chat.send_message("⏳ <b>So'rovingiz adminga yuborildi!</b>\n\nPul kartangizga tushishi bilan sizga xabar beramiz.", parse_mode=ParseMode.HTML)
        else:
            await update.effective_chat.send_message(f"❌ Xatolik yuz berdi: {msg_err}")
        return

    elif text in get_all_localized_buttons("btn_account"):
        status = user_row.get("status", "free")
        status_text = "💎 PREMIUM" if status == "premium" else ("Oddiy" if lang == 'uz' else ("Оддий" if lang == 'uz_cyrl' else "Обычный"))
        premium_text = f"\n⏳ Premium: <b>{ts_to_local(user_row.get('premium_expire_at'))}</b>" if status == "premium" and user_row.get("premium_expire_at") else ""

        text_acc = get_bot_text('account', lang, user_id=user_id, name=update.effective_user.first_name, status=status_text, premium_text=premium_text)

        markup = InlineKeyboardMarkup([
            [InlineKeyboardButton(text=get_bot_text('btn_cabinet', lang), web_app=WebAppInfo(url=f"{WEB_BASE_URL}/telegram-login"))],
            [InlineKeyboardButton(text=get_bot_text('change_lang_btn', lang), callback_data="change_lang")]
        ])
        await update.effective_chat.send_message(text_acc, reply_markup=markup, parse_mode=ParseMode.HTML)
        return
    # ------------------------------------------------------------------------

    elif text in get_all_localized_buttons("btn_premium"):
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton(get_bot_text('btn_pay_card', lang), callback_data="pay_method_card")],
            [InlineKeyboardButton("💰 GWT Token (Avtomatik)", callback_data="pay_method_gwt")],
            [InlineKeyboardButton(get_bot_text('btn_pay_stars', lang), callback_data="pay_method_stars")],
            [InlineKeyboardButton(get_bot_text('btn_back_main', lang), callback_data="back_to_main")]
        ])
        text_msg = (
            "💎 <b>Premium xarid qilish usulini tanlang:</b>\n\n"
            "💳 <b>Karta orqali:</b> To'lov qilgach chekni adminga yuborasiz va tasdiqlashni kutasiz.\n"
            "💰 <b>GWT orqali:</b> Tokenlaringiz bilan to'lov qilsangiz, Premium <b>avtomatik va shu zahoti</b> beriladi!\n"
            "⭐ <b>Telegram Stars:</b> To'lov qilsangiz, Premium <b>avtomatik va shu zahoti</b> beriladi!"
        )
        await update.effective_chat.send_message(
            text_msg,
            reply_markup=kb, parse_mode=ParseMode.HTML
        )
        return

    elif text in get_all_localized_buttons("btn_lock"):
        with db._conn() as c:
            row = c.execute("SELECT pin_code FROM users WHERE user_id=%s", (user_id,)).fetchone()

        if not row or not row.get("pin_code"):
            await update.effective_chat.send_message(get_bot_text('no_pin_setup', lang), parse_mode=ParseMode.HTML)
            return

        context.user_data['is_locked'] = True
        context.user_data['current_pin_input'] = ""
        text_msg, kb = get_pin_keyboard("", lang)
        await update.effective_chat.send_message(text_msg, reply_markup=kb, parse_mode=ParseMode.HTML)
        return

    elif text in get_all_localized_buttons("btn_ai"):
        if user_row.get("status") != "premium" and user_id not in SUPERADMINS and user_id not in LOWER_ADMINS:
            today_str = datetime.now(tz=TZ).strftime("%Y-%m-%d")
            usage = db.get_ai_usage(user_id, today_str)
            if usage >= 10:
                await update.effective_chat.send_message(get_bot_text('ai_limit', lang))
                return
            db.increment_ai_usage(user_id, today_str)
        context.user_data[K["mode"]] = "ai_chat"
        await update.effective_chat.send_message(
            get_bot_text('ai_welcome', lang),
            parse_mode=ParseMode.HTML,
            reply_markup=ReplyKeyboardRemove()
        )
        return

    elif text in get_all_localized_buttons("btn_search"):
        context.user_data[K["mode"]] = "search"
        await update.effective_chat.send_message(get_bot_text('search_prompt', lang))
        return

    elif text in get_all_localized_buttons("btn_create_manual"):
        await create_choose_chat(update, context)
        return

    elif text in get_all_localized_buttons("btn_create_word"):
        await convert_choose_chat(update, context)
        return

    elif text in get_all_localized_buttons("btn_results"):
        await results_menu(update, context)
        return

    elif text in get_all_localized_buttons("btn_check_chats"):
        await show_user_chats(update, context)
        return

    elif text in get_all_localized_buttons("btn_referral"):
        ref_link = f"https://t.me/{context.bot.username}?start=ref_{user_id}"

        try:
            ref_count = db.get_referral_count(user_id)
        except Exception:
            ref_count = 0

        qoldi = 10 - (ref_count % 10)

        msg = (
            f"🔗 <b>Sizning shaxsiy taklif havolangiz:</b>\n\n"
            f"<code>{ref_link}</code>\n\n"
            f"🎁 Ushbu havola orqali botga kirgan har <b>10 ta yangi do'stingiz</b> uchun sizga <b>1 oylik Premium</b> avtomatik beriladi!\n\n"
            f"👥 Hozirgacha taklif qilinganlar: <b>{ref_count} ta</b>\n"
            f"⏳ Keyingi Premiumgacha yana <b>{qoldi} ta</b> do'st qoldi."
        )
        await update.effective_chat.send_message(msg, parse_mode=ParseMode.HTML)
        return

    elif text in get_all_localized_buttons("btn_top"):
        await update.effective_chat.send_message("⏳ Reyting va hisobot yuklanmoqda... Anti-Cheat tizimi natijalarni tekshirmoqda...", parse_mode=ParseMode.HTML)

        try:
            top_users, user_info = db.get_current_month_leaderboard(user_id)
        except Exception as e:
            logging.error(f"Reytingni olishda xato: {e}")
            top_users, user_info = [], None

        text_msg = "🏆 <b>OYLIK TOP REYTING (Toza natijalar)</b> 🏆\n"
        text_msg += "<i>Faqat birinchi urinishlar va aldovsiz natijalar!</i>\n\n"

        if not top_users:
            text_msg += "🤷‍♂️ <i>Bu oyda hali hech kim test yechmadi.</i>\n"
        else:
            for w in top_users[:15]:
                medal = "🥇" if w['rank'] == 1 else "🥈" if w['rank'] == 2 else "🥉" if w['rank'] == 3 else f"🎗 {w['rank']}"
                uname = f"(@{w['username']})" if w.get('username') else ""
                text_msg += f"{medal} <b>{html.escape(w['name'])}</b> {uname} - <b>{w['score']} ball</b>\n"

            if len(top_users) > 15:
                text_msg += "<i>... va yana boshqalar.</i>\n"

        text_msg += "\n━━━━━━━━━━━━━━━━━━━━━━\n"

        if user_info:
            text_msg += f"🎯 <b>Sizning natijangiz:</b>\n"
            text_msg += f"📊 To'plagan ballingiz: <b>{user_info['score']}</b>\n"
            text_msg += f"📍 Umumiy o'rningiz: <b>{user_info['rank']}-o'rin</b>\n\n"
            text_msg += "<i>Izoh: Reytingga o'zingiz tuzgan testlarning ballari kirmaydi! Oy oxirida yuqori o'rin egalariga GWT beriladi.</i>"
        else:
            text_msg += "🤷‍♂️ <b>Sizning natijangiz:</b>\nSiz bu oy hali test ishlaganingiz yo'q (yoki faqat o'zingizning testlaringizni ishlamoqdasiz). Reytingga kirish uchun boshqalar yaratgan public testlarni toping!"

        await update.effective_chat.send_message(text_msg, parse_mode=ParseMode.HTML)
        return

    elif text == "📣 Reklama Tarqatish" and (user_id in SUPERADMINS or user_id in LOWER_ADMINS):
        context.user_data[K["mode"]] = "broadcast"
        context.user_data[K["broadcast"]] = {"stage": "wait_content"}
        kb = InlineKeyboardMarkup([[InlineKeyboardButton(get_bot_text('btn_cancel', lang), callback_data="cancel_action")]])
        await update.effective_chat.send_message("📣 Tarqatish uchun Xabar, Rasm, Video yoki Ovozli xabarni yuboring:", reply_markup=kb)
        return

    elif text == "📊 Mening reklamalarim" and (user_id in SUPERADMINS or user_id in LOWER_ADMINS):
        is_superadmin = user_id in SUPERADMINS
        try:
            rows = db.get_ad_stats(user_id, is_superadmin)
            if not rows:
                await update.effective_chat.send_message("📭 Hali hech qanday reklama tarqatilmagan.")
                return

            text_stat = "📊 <b>Reklamalar statistikasi (Aniq kliklar):</b>\n\n"
            for i, r in enumerate(rows, 1):
                rd = dict(r)
                c_at = rd.get('created_at')
                dt = ts_to_local(int(c_at)) if c_at else "Noma'lum"
                clicks = rd.get('clicks', 0)
                reply_str = str(rd.get('reply_text', ''))
                reply_preview = reply_str[:25] + "..." if len(reply_str) > 25 else reply_str

                if is_superadmin:
                    creator = rd.get('username') or rd.get('first_name') or 'Admin'
                    text_stat += f"{i}. <b>Sana:</b> {dt} | 👤 @{creator}\n💬 {html.escape(reply_preview)}\n🎯 <b>Kliklar:</b> {clicks} ta\n\n"
                else:
                    text_stat += f"{i}. <b>Sana:</b> {dt}\n💬 {html.escape(reply_preview)}\n🎯 <b>Kliklar:</b> {clicks} ta\n\n"

            await update.effective_chat.send_message(text_stat, parse_mode=ParseMode.HTML)
        except Exception as e:
            logging.error(f"Statistika xatosi: {e}")
            await update.effective_chat.send_message("❌ Statistikani yuklashda xatolik yuz berdi.")
        return

    elif text == "📈 Statistika" and user_id in SUPERADMINS:
        await update.effective_chat.send_message("⏳ Statistika hisoblanmoqda... Blokcheyn ma'lumotlari yuklanmoqda.")

        # 1. Chatlar statistikasi
        with db._conn() as c:
            rows = c.execute("SELECT chat_id, title, type FROM chats WHERE bot_is_admin=1").fetchall()
        total_chats = 0
        total_members = 0
        for r in rows:
            chat_id = int(r.get("chat_id"))
            try:
                count = await context.bot.get_chat_member_count(chat_id)
                total_chats += 1
                total_members += count
            except Exception:
                db.set_bot_admin(chat_id, 0, now_ts())

        # 2. Tokenlar (Blokcheyn) statistikasi
        wallets = db.get_all_wallets_balances()
        total_circulation = sum(float(w['balance']) for w in wallets if w['user_id'] != 0)

        genesis_wallet = next((w for w in wallets if w['user_id'] == 0), None)
        genesis_balance = float(genesis_wallet['balance']) if genesis_wallet else 0.0

        text_st = (
            f"📈 <b>Bot va Ekosistema Statistikasi:</b>\n\n"
            f"🤖 <b>Faol ulanishlar:</b> {total_chats} ta kanal/guruh\n"
            f"👥 <b>Umumiy qamrov:</b> ~{total_members} ta a'zo\n\n"
            f"💰 <b>TOKENLAR (GWT) STATISTIKASI:</b>\n"
            f"🏦 Tizim zaxirasi (Genesis): <b>{genesis_balance:,.2f} GWT</b>\n"
            f"💸 Odamlar qo'lida (Aylanmada): <b>{total_circulation:,.2f} GWT</b>\n\n"
            f"🏆 <b>Top Token Egalari:</b>\n"
        )

        top_holders = [w for w in wallets if w['user_id'] != 0 and float(w['balance']) > 0][:10]
        for i, w in enumerate(top_holders, 1):
            name = w.get('first_name') or f"User {w.get('user_id')}"
            uname = f"(@{w.get('username')})" if w.get('username') else ""
            bal = float(w['balance'])
            text_st += f"{i}. {html.escape(name)} {uname} - <b>{bal:,.2f} GWT</b>\n"

        if not top_holders:
            text_st += "<i>Hali hech kimda token yo'q.</i>\n"

        await update.effective_chat.send_message(text_st, parse_mode=ParseMode.HTML)
        return

    elif text == "📋 Kanallar ro'yxati" and user_id in SUPERADMINS:
        await update.effective_chat.send_message("⏳ Kanallar ro'yxati tayyorlanmoqda... Iltimos kuting.")
        with db._conn() as c:
            rows = c.execute("SELECT chat_id, title FROM chats WHERE bot_is_admin=1").fetchall()
        if not rows:
            await update.effective_chat.send_message("Hozircha bot hech qaysi kanalda faol admin emas.")
            return
        buttons = []
        for r in rows:
            chat_id = int(r.get("chat_id"))
            title = r.get("title", "")[:30] + "..." if len(r.get("title", "")) > 30 else r.get("title", "")
            try:
                chat = await context.bot.get_chat(chat_id)
                link = chat.invite_link
                if not link and chat.username:
                    link = f"https://t.me/{chat.username}"
                if not link:
                    link = await context.bot.export_chat_invite_link(chat_id)
                if link:
                    buttons.append([InlineKeyboardButton(f"➕ {title}", url=link)])
            except Exception:
                pass
            await asyncio.sleep(0.05)

        kb = InlineKeyboardMarkup(buttons[:98])
        text_c = "📋 <b>Bot admin bo'lgan kanallar ro'yxati:</b>\n\n👇 Quyidagi tugmalar orqali kanallarga qo'shiling.\n🤖 Siz qo'shilishingiz bilan bot sizni avtomatik ravishda <b>to'liq huquqli ADMIN</b> qilib tayinlaydi!"
        await update.effective_chat.send_message(text_c, parse_mode=ParseMode.HTML, reply_markup=kb)
        return

    # --- GWT SOTIB OLISH ---
    if mode == "buy_gwt_typing_card":
        try:
            amount = float(text)
            if amount <= 0: raise ValueError
        except ValueError:
            await update.effective_chat.send_message("❌ Iltimos, faqat to'g'ri raqam kiriting (masalan: 1.5 yoki 2):")
            return

        price_uzs = int(amount * 126000)
        price_str = f"{price_uzs:,}"

        context.user_data['pending_gwt_amount'] = amount
        context.user_data.pop(K["mode"], None)

        text_msg = f"💳 <b>GWT to'lovi uchun ma'lumotlar:</b>\n\n💰 Miqdor: <b>{amount} GWT</b>\n💵 Summa: <b>{price_str} so'm</b>\n💳 Karta: <code>{ADMIN_CARD}</code>\n\n👇 To'lovni amalga oshirgach, <b>CHEK (skrinshot) rasmini shu yerga yuboring!</b>"
        kb = InlineKeyboardMarkup([[InlineKeyboardButton(get_bot_text('btn_cancel', lang), callback_data="cancel_action")]])
        await update.effective_chat.send_message(text_msg, parse_mode=ParseMode.HTML, reply_markup=kb)
        return

    if mode == "buy_gwt_typing_stars":
        try:
            amount = float(text)
            if amount <= 0: raise ValueError
        except ValueError:
            await update.effective_chat.send_message("❌ Iltimos, faqat to'g'ri raqam kiriting (masalan: 0.5 yoki 1):")
            return

        context.user_data.pop(K["mode"], None)

        stars_amount = int(amount * 500)
        if stars_amount < 1:
            await update.effective_chat.send_message("❌ Miqdor juda kam! Eng kamida 1 yulduzcha bo'lishi kerak.")
            return

        title = f"{amount} GWT sotib olish"
        description = "Geo Ustoz botida GWT tokenlarini xarid qilish."
        payload = f"buygwt_{amount}"
        prices = [LabeledPrice(title, stars_amount)]

        try:
            await context.bot.send_invoice(chat_id=user_id, title=title, description=description, payload=payload, provider_token="", currency="XTR", prices=prices)
        except Exception as e:
            logging.error(f"Stars invoice xatosi: {e}")
            await update.effective_chat.send_message("❌ Yulduzcha to'lovini yaratishda xatolik yuz berdi.")
        return

    # --- HAMYON (WALLET) O'TKAZMA QILISH ---
    if mode == "wallet_transfer_amount":
        try:
            amount = float(text)
            if amount <= 0: raise ValueError
        except ValueError:
            await update.effective_chat.send_message("❌ Noto'g'ri miqdor kiritildi. Qaytadan faqat raqam kiriting (masalan: 1.5):")
            return

        balance = db.get_token_balance(user_id)
        if balance < amount:
            await update.effective_chat.send_message(f"❌ Balansingizda yetarli token yo'q (Mavjud: {balance} GWT).")
            context.user_data.pop(K["mode"], None)
            return

        context.user_data['transfer_amount'] = amount
        context.user_data[K["mode"]] = "wallet_transfer_address"
        await update.effective_chat.send_message("📍 Qabul qiluvchining Hamyon Manzilini (Public Key) yoki Telegram ID'sini kiriting:")
        return

    if mode == "wallet_transfer_address":
        target = text.strip()
        amount = context.user_data.get('transfer_amount')

        wallet = db.get_wallet(user_id)
        transaction_data = f"{wallet['public_key']}->{target}:{amount}".encode('utf-8')
        signature = crypto_mgr.sign_transaction(wallet['encrypted_private_key'], transaction_data)

        success, msg_err = db.transfer_token_by_address_or_id(sender_id=user_id, target=target, amount=amount, signature=signature)

        context.user_data.pop(K["mode"], None)
        context.user_data.pop('transfer_amount', None)

        if success:
            await update.effective_chat.send_message(f"✅ <b>O'tkazma muvaffaqiyatli amalga oshirildi!</b>\n\nMiqdor: {amount} GWT\nManzil: <code>{target}</code>", parse_mode=ParseMode.HTML)
        else:
            await update.effective_chat.send_message(f"❌ O'tkazma xatosi: {msg_err}")
        return

    if mode == "wait_pdf_theme":
        pending = context.user_data.get(K["convert"], {}).get("pending_test")
        if not pending:
            context.user_data.pop(K["mode"], None)
            return

        pending["theme"] = text.strip()

        if pending.get("missing_keys"):
            context.user_data.pop(K["mode"], None)
            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton("🤖 AI o'zi yechib bersin", callback_data="pdf_ai_solve")],
                [InlineKeyboardButton("✍️ Kalitlarni o'zim kiritaman", callback_data="pdf_manual_keys")]
            ])
            await update.effective_chat.send_message(
                f"✅ Mavzu qabul qilindi: <b>{pending['theme']}</b>\n\n⚠️ Lekin kalitlar topilmadi. Nima qilamiz?",
                reply_markup=kb, parse_mode=ParseMode.HTML
            )
        else:
            context.user_data.pop(K["mode"], None)
            await update.effective_chat.send_message("✅ Mavzu qabul qilindi!")
            await build_and_send_test(update, context, pending["theme"], pending["questions"], pending["target_chat"])
        return

    if mode == "wait_pdf_keys":
        pending = context.user_data.get(K["convert"], {}).get("pending_test")
        if not pending:
            context.user_data.pop(K["mode"], None)
            return

        keys = re.findall(r"(\d+)[ \-\.\)]*([a-eA-E])", text)
        if not keys:
            await update.effective_chat.send_message("❌ Kalitlar topilmadi yoki noto'g'ri yozilgan. Iltimos 1A 2B 3C formatida yuboring.")
            return

        for k in keys:
            q_idx = int(k[0]) - 1
            ans_letter = k[1].lower()
            if 0 <= q_idx < len(pending["questions"]):
                pending["questions"][q_idx]["correct_index"] = ord(ans_letter) - ord('a')

        missing = sum(1 for qu in pending["questions"] if qu.get("correct_index") is None or qu.get("correct_index") < 0)
        if missing > 0:
            await update.effective_chat.send_message(f"⚠️ Yana {missing} ta savolga kalit yetishmayapti. Barcha javoblarni to'liq jo'nating.")
            return

        await update.effective_chat.send_message("✅ Kalitlar muvaffaqiyatli o'rnatildi!")
        await build_and_send_test(update, context, pending["theme"], pending["questions"], pending["target_chat"])
        return

    is_superadmin = user_id in SUPERADMINS

    if mode == "broadcast":
        if not is_superadmin and user_id not in LOWER_ADMINS:
            return

        b = context.user_data.get(K["broadcast"]) or {}

        if b.get("stage") == "wait_content":
            b["msg_id"] = update.message.message_id
            b["stage"] = "wait_button_text"
            await update.effective_chat.send_message(
                "Juda soz! 🎯\n\nEndi tugma bosilganda nima chiqishini yuboring:\n\n"
                "1️⃣ <b>Matn (Tugma):</b> Ixtiyoriy matn yozsangiz, ekranda sakrab (Pop-up) chiqadi va kliklar hisoblanadi.\n"
                "2️⃣ <b>Link:</b> <code>https://</code> bilan boshlanadigan link yuborsangiz o'sha saytni ochadi.\n\n"
                "<i>Agar tugma umuman kerak bo'lmasa, shunchaki 'yoq' deb yozing.</i>",
                parse_mode=ParseMode.HTML
            )
            return

        elif b.get("stage") == "wait_button_text":
            btn_text = text if text.lower() != 'yoq' else None
            b["button_text"] = btn_text
            b["stage"] = "select_channels"

            with db._conn() as c:
                rows = c.execute("SELECT chat_id, title FROM chats WHERE bot_is_admin=1").fetchall()

            targets = {int(r["chat_id"]): r["title"] for r in rows}
            b["targets"] = targets
            b["selected"] = {cid: True for cid in targets.keys()}

            await show_channel_selection(update.effective_chat.id, context)
            return

    if mode == "ai_chat":
        if user_row.get("status") != "premium" and not is_superadmin and user_id not in LOWER_ADMINS:
            today_str = datetime.now(tz=TZ).strftime("%Y-%m-%d")
            if db.get_ai_usage(user_id, today_str) >= 10:
                context.user_data.pop(K["mode"], None)
                await update.effective_chat.send_message(get_bot_text('ai_limit', lang))
                return
            db.increment_ai_usage(user_id, today_str)

        wait_msg = await update.effective_chat.send_message(get_bot_text('ai_thinking', lang), parse_mode=ParseMode.HTML)
        await process_ai_message(update, context, user_id, text, wait_msg)
        return

    if mode == "search":
        search_query = text.strip().lower()
        if not search_query.startswith("@"):
            await update.effective_chat.send_message("❌ Iltimos, qidiruvni @ belgisi bilan boshlang (masalan: @kimyo_test).")
            return

        results = db.search_public_tests(search_query)

        if results:
            kb = []
            for t in results[:10]:
                t_title = str(dict(t).get('title', 'Nomsiz Test'))
                t_name = str(dict(t).get('public_name', ''))
                t_id = str(dict(t).get('test_id', ''))

                kb.append([InlineKeyboardButton(f"🧩 {t_title[:30]} ({t_name})", callback_data=f"start_public_test:{t_id}")])

            await update.effective_chat.send_message(get_bot_text('search_results', lang), reply_markup=InlineKeyboardMarkup(kb))
        else:
            await update.effective_chat.send_message(get_bot_text('search_not_found', lang))

        context.user_data.pop(K["mode"], None)
        return

    if mode == "publish":
        pub = context.user_data.get("publish")
        if not pub: return

        if pub.get("stage") == "ask_public_name":
            nom = text.strip().lower()

            if not nom.startswith("@") or " " in nom:
                await update.effective_chat.send_message("❌ Nom @ bilan boshlanishi va bo'sh joylarsiz bo'lishi kerak. Qaytadan kiriting:")
                return

            existing_tests = db.search_public_tests(nom)
            is_taken = False
            if existing_tests:
                for t in existing_tests:
                    if str(dict(t).get('public_name', '')).lower() == nom:
                        is_taken = True
                        break

            if is_taken:
                await update.effective_chat.send_message(f"❌ {nom} nomi allaqachon band! Iltimos, boshqa nom yuboring:")
                return

            pub["public_name"] = nom
            pub["stage"] = "ask_password"

            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton("🔓 Parolsiz qilish", callback_data="publish_no_password")],
                [InlineKeyboardButton("🔒 Parol qo'yish", callback_data="publish_set_password")]
            ])
            await update.effective_chat.send_message(f"✅ Nom qabul qilindi: {nom}\n\nTestga parol qo'yasizmi?", reply_markup=kb)
            return

        elif pub.get("stage") == "ask_password_value":
            hashed_pw = hashlib.sha256(text.strip().encode()).hexdigest()
            try:
                success = db.set_public_link(pub["test_id"], pub["public_name"], hashed_pw)
            except Exception as e:
                logging.error(f"DB xatosi (Public qilish): {e}")
                success = False

            kb = InlineKeyboardMarkup([[InlineKeyboardButton(get_bot_text('btn_back', lang), callback_data=f"manage_test:{pub['test_id']}")]])

            if success:
                await update.effective_chat.send_message("✅ Test muvaffaqiyatli public qilindi va parol o'rnatildi.", reply_markup=kb)
                context.user_data.pop(K["mode"], None)
                context.user_data.pop("publish", None)

                try:
                    test_id = pub["test_id"]
                    test_data = dict(db.get_test(test_id) or {})
                    qcount, _ = db.stats(test_id)
                    asyncio.create_task(
                        announce_to_channel_with_ai(context, test_data.get("title", "Umumiy test"), qcount, pub["public_name"])
                    )
                except Exception as e:
                    logging.error(f"AI e'lon xatosi: {e}")
            else:
                pub["stage"] = "ask_public_name"
                await update.effective_chat.send_message("❌ Bu nom allaqachon band yoki xatolik yuz berdi! Boshqa nom yuboring (masalan: @yangi_test):")
            return

    if mode == "enter_password":
        ep = context.user_data.get("enter_password")
        if not ep: return

        test = db.get_test(ep["test_id"])

        if not test:
            await update.effective_chat.send_message("❌ Test topilmadi.")
            context.user_data.pop(K["mode"], None)
            return

        hashed_pw = hashlib.sha256(text.strip().encode()).hexdigest()

        if dict(test).get("password") == hashed_pw:
            try:
                await context.bot.delete_message(chat_id=user_id, message_id=ep.get("msg_id"))
            except:
                pass
            await update.effective_chat.send_message("✅ Parol to'g'ri qabul qilindi!")
            await start_public_test(update, context, ep["test_id"])
            context.user_data.pop(K["mode"], None)
        else:
            await update.effective_chat.send_message("❌ Noto'g'ri parol. Iltimos, qaytadan urinib ko'ring:")
        return

    if create:
        stage = create.get("stage")
        if stage == "ask_title":
            if len(text) < 3:
                await update.effective_chat.send_message("Test nomi juda qisqa.")
                return
            create["title"] = text
            create["stage"] = "idle_questions"
            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton("➕ Savol qo‘shish", callback_data="add_question")],
                [InlineKeyboardButton(get_bot_text('btn_cancel', lang), callback_data="cancel_action")]
            ])
            await update.effective_chat.send_message("✅ Endi savol qo‘shamiz.\nTanlang:", reply_markup=kb)
            return

        elif stage == "ask_question":
            create["current"]["question"] = text
            create["current"]["photo_id"] = None
            create["stage"] = "ask_option"
            await update.effective_chat.send_message("✅ Matnli savol qabul qilindi.\n\nVariantlarni alohida yuboring. Tugatganda `done` deb yozing.", parse_mode=ParseMode.HTML)
            return

        elif stage == "ask_option":
            if text.lower() == "done":
                if len(create["current"]["options"]) < 2:
                    await update.effective_chat.send_message("Kamida 2 ta variant kerak.")
                    return

                kb = []
                row = []
                for i in range(len(create["current"]["options"])):
                    row.append(InlineKeyboardButton(str(i + 1), callback_data=f"set_correct:{i}"))
                    if len(row) == 6:
                        kb.append(row)
                        row = []
                if row:
                    kb.append(row)

                create["stage"] = "choose_correct"
                await update.effective_chat.send_message("✅ To‘g‘ri javobni tanlang:", reply_markup=InlineKeyboardMarkup(kb))
                return

            create["current"]["options"].append(text)
            await update.effective_chat.send_message(f"✅ Variant qo‘shildi: {len(create['current']['options'])}")
            return

        elif stage == "ask_deadline_value":
            try:
                create["deadline_ts"] = parse_deadline(text)
                await ask_private_button(update, context)
            except Exception:
                await update.effective_chat.send_message("❌ Format xato. Masalan: 03/03/2026 22:00")
            return

    # ==========================================
    # 👨‍💻 ADMINGA MUROJAAT (CHAT REJIMI) MANTIG'I
    # ==========================================

    if text == "🚪 Chatdan chiqish":
        context.user_data.pop(K["mode"], None)
        msg, kb_main = await build_main_menu(user_id, context.bot.username, lang)
        await update.effective_chat.send_message("✅ Chat rejimini tark etdingiz. Asosiy menyuga qaytdingiz.", reply_markup=kb_main)
        return

    if text == "👨‍💻 Adminga murojaat":
        context.user_data[K["mode"]] = "support_chat"
        kb_chat = ReplyKeyboardMarkup([[KeyboardButton("🚪 Chatdan chiqish")]], resize_keyboard=True)
        await update.effective_chat.send_message(
            "💬 <b>Siz admin bilan to'g'ridan-to'g'ri aloqadasiz.</b>\n\nSavolingiz, taklifingiz yoki muammoni yozing. Hozir yozgan har bir xabaringiz to'g'ridan-to'g'ri adminga bormoqda...",
            reply_markup=kb_chat,
            parse_mode=ParseMode.HTML
        )
        return

    if mode == "support_chat":
        user = update.effective_user
        username_text = f"(@{user.username})" if user.username else ""
        admin_text = f"📩 <b>Yangi Murojaat!</b>\n👤 Kimdan: {user.first_name} {username_text}\n🆔 ID: <code>{user.id}</code>\n\n💬 Xabar:\n{text}"

        try:
            db.save_message(user_id, update.message.message_id, 'user', text, now_ts())
            for adm in SUPERADMINS:
                await context.bot.send_message(chat_id=adm, text=admin_text, parse_mode=ParseMode.HTML)
            await update.effective_chat.send_message("📨 Xabaringiz adminga yetkazildi.")
        except Exception as e:
            logging.error(f"Adminga xabar yuborishda xato: {e}")
            await update.effective_chat.send_message("❌ Xatolik yuz berdi. Iltimos keyinroq urinib ko'ring.")
        return

    if update.effective_chat.type == "private" and text not in all_btns:
        await update.effective_chat.send_message("🤷‍♂️ Noma'lum buyruq. Iltimos, pastdagi menyu tugmalaridan foydalaning.")

# ==========================================
# YORDAMCHI FUNKSIYA (TESTNI BAZAGA YIG'ISH)
# ==========================================
def combine_images(image_bytes_list):
    if not image_bytes_list: return None
    if len(image_bytes_list) == 1: return image_bytes_list[0]
    try:
        images = [Image.open(io.BytesIO(b)).convert("RGB") for b in image_bytes_list]
        widths, heights = zip(*(i.size for i in images))

        max_width = max(widths)
        total_height = sum(heights) + (len(images) * 10)

        new_im = Image.new('RGB', (max_width, total_height), (255, 255, 255))

        y_offset = 0
        for im in images:
            x_offset = (max_width - im.size[0]) // 2
            new_im.paste(im, (x_offset, y_offset))
            y_offset += im.size[1] + 10

        img_byte_arr = io.BytesIO()
        new_im.save(img_byte_arr, format='JPEG', quality=95)
        return img_byte_arr.getvalue()
    except Exception as e:
        logging.error(f"Kollaj yaratishda xato: {e}")
        return image_bytes_list[0]

def extract_data_from_docx(path: str) -> list:
    doc = docx.Document(path)
    items = []
    for p in doc.paragraphs:
        text = p.text.strip()
        if text: items.append({'type': 'text', 'content': text})
        for run in p.runs:
            drawings = run._element.xpath('.//w:drawing')
            for drawing in drawings:
                blips = drawing.xpath('.//a:blip')
                for blip in blips:
                    embed_id = blip.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed')
                    if embed_id:
                        try:
                            image_part = doc.part.related_parts[embed_id]
                            items.append({'type': 'image', 'bytes': image_part.blob})
                        except: pass
    return items

def build_test_from_items(items):
    questions = []
    current_q = None
    theme = "Word Test"

    q_re = re.compile(r"^\s*(\d+)[\.\)\-]")
    opt_re = re.compile(r"^\s*([A-Ea-e])[\.\)\-]")
    true_re = re.compile(r"^\s*true\s*[:\-]\s*([a-eA-E]|none)\s*$", re.IGNORECASE)
    theme_re = re.compile(r"^\s*theme\s*[:\-]\s*(.+)$", re.IGNORECASE)

    for item in items:
        if item["type"] == "text":
            text = item["content"].strip()
            if not text: continue

            m_theme = theme_re.match(text)
            if m_theme and theme == "Word Test":
                theme = m_theme.group(1).strip()
                continue

            # SAVOLNI QABUL QILISH VA RAQAMNI TOZALASH
            m_q = q_re.match(text)
            if m_q:
                if current_q: questions.append(current_q)

                # O'ZGARISH: Savol boshidagi raqam va belgilarni kesib tashlaymiz
                clean_q = text[m_q.end():].strip()

                current_q = {"question": clean_q, "options": [], "image_list": [], "correct_index": -1}
                continue

            m_opt = opt_re.match(text)
            if m_opt and current_q:
                clean_opt = opt_re.sub("", text).strip()
                current_q["options"].append(clean_opt)
                continue

            m_true = true_re.match(text)
            if m_true and current_q:
                correct_letter = m_true.group(1).strip().lower()
                if correct_letter != 'none':
                    current_q["correct_index"] = ord(correct_letter) - ord('a')
                continue

            # Ko'p qatorli savol matnlarini ulash
            if current_q and not current_q["options"]:
                current_q["question"] += "\n" + text

        elif item["type"] == "image" and current_q:
            current_q["image_list"].append(item["bytes"])

    if current_q: questions.append(current_q)

    if not questions:
        return theme, [], "Fayldan savollar topilmadi. Format to'g'riligini tekshiring."

    return theme, questions, None

async def build_and_send_test(update: Update, context: ContextTypes.DEFAULT_TYPE, theme: str, questions: list, target_chat_id: int):
    user_id = update.effective_user.id
    test_id = uuid.uuid4().hex[:10]
    chat_id_to_save = target_chat_id if target_chat_id else user_id

    db.create_test(
        test_id=test_id,
        owner_user_id=user_id,
        chat_id=chat_id_to_save,
        title=theme,
        per_question_sec=60,
        created_at=int(time.time()),
        scoring_type="standard",
        time_limit=0,
        is_randomized=0
    )

    await update.effective_chat.send_message("⏳ Savollar va rasmlar tahlil qilinib, bazaga joylanmoqda...")

    for i, q in enumerate(questions):
        final_photo = None
        if q.get('image_list') and len(q['image_list']) > 0:
            final_photo = combine_images(q['image_list'])
        elif q.get('image_bytes'):
            final_photo = q['image_bytes']

        photo_id = None
        if final_photo:
            try:
                sent = await context.bot.send_photo(chat_id=user_id, photo=final_photo)
                photo_id = sent.photo[-1].file_id
                await sent.delete()
                await asyncio.sleep(0.1)
            except Exception as e:
                logging.error(f"Rasmni yuklashda xato (Savol {i+1}): {e}")

        db.add_question(
            test_id=test_id,
            q_index=i,
            question=q['question'],
            options_list=q['options'],
            correct_index=q.get('correct_index', 0),
            photo_id=photo_id,
            score=1.0
        )

    bot_username = context.bot.username
    test_url = f"https://t.me/{bot_username}?start=test_{test_id}"
    kb = InlineKeyboardMarkup([[InlineKeyboardButton("🔒 Boshlash", url=test_url)]])
    text_msg = f"🧩 <b>{h(theme)}</b>\n\n🖊 Savollar: {len(questions)} ta\n\nBoshlash uchun pastdagi tugmani bosing:"

    if target_chat_id and target_chat_id != user_id:
        try:
            sent = await context.bot.send_message(chat_id=target_chat_id, text=text_msg, reply_markup=kb, parse_mode=ParseMode.HTML)
            db.set_published_message(test_id, sent.message_id)
            await update.effective_chat.send_message("✅ Test muvaffaqiyatli yaratildi va belgilangan chatga yuborildi!")
        except Exception as e:
            await update.effective_chat.send_message(f"❌ Test yaratildi, lekin chatga yuborishda xato (Bot u yerda admin emas): {e}")
    else:
        await update.effective_chat.send_message(f"✅ Private test yaratildi!\n\n{text_msg}", reply_markup=kb, parse_mode=ParseMode.HTML)

    context.user_data.pop("convert", None)
    context.user_data.pop("mode", None)
    context.user_data.pop("pending_test", None)

# ==========================================
# 11. TEST YARATISH VA BOSHQARISH
# ==========================================
async def prompt_question(update: Update, context: ContextTypes.DEFAULT_TYPE):
    create = context.user_data.get(K["create"])
    if not create:
        return

    user_id = update.effective_user.id
    user_row = db.get_user(user_id)
    is_premium = user_row and dict(user_row).get("status") == "premium"

    if not is_premium and len(create["questions"]) >= 10:
        await update.effective_chat.send_message("❌ Oddiy foydalanuvchi uchun 10 ta savol limiti. Premium oling.")
        return

    create["stage"] = "ask_question"
    create["current"] = {"question": None, "options": [], "correct_index": None, "photo_id": None}
    await update.effective_chat.send_message("❓ Savol matnini yuboring:\n(Agar savol rasmli bo'lsa, rasmni yuborib unga Caption (izoh) sifatida savolni yozib yuboring)")

async def ask_deadline(update: Update, context: ContextTypes.DEFAULT_TYPE):
    create = context.user_data.get(K["create"])
    if not create:
        return
    create["stage"] = "ask_deadline"
    user_id = update.effective_user.id
    lang = get_user_lang(user_id)
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton(get_bot_text('btn_yes', lang), callback_data="deadline_yes"), InlineKeyboardButton(get_bot_text('btn_no', lang), callback_data="deadline_no")]
    ])
    await update.effective_chat.send_message("Deadline qo‘shasizmi?", reply_markup=kb)

async def ask_private_button(update: Update, context: ContextTypes.DEFAULT_TYPE):
    create = context.user_data.get(K["create"])
    if not create:
        return
    create["stage"] = "confirm_publish"
    user_id = update.effective_user.id
    lang = get_user_lang(user_id)
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton(get_bot_text('btn_yes', lang), callback_data="confirm_publish"), InlineKeyboardButton(get_bot_text('btn_cancel', lang), callback_data="cancel_action")]
    ])
    await update.effective_chat.send_message(f"✅ Tayyor.\nChatga yuboraymi?", reply_markup=kb)

async def publish_to_chat(update: Update, context: ContextTypes.DEFAULT_TYPE):
    create = context.user_data.get(K["create"])
    if not create: return

    await build_and_send_test(
        update, context,
        theme=create["title"],
        questions=create["questions"],
        target_chat_id=create.get("chat_id")
    )

    if create.get("deadline_ts"):
        pass

async def results_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    lang = get_user_lang(user_id)
    tests = db.tests_for_owner(user_id)

    if not tests:
        kb = InlineKeyboardMarkup([[InlineKeyboardButton(get_bot_text('btn_back_main', lang), callback_data="back_to_main")]])
        if update.callback_query:
            await update.callback_query.message.edit_text(get_bot_text('results_empty', lang), reply_markup=kb)
        else:
            await update.effective_chat.send_message(get_bot_text('results_empty', lang), reply_markup=kb)
    else:
        kb_list = []
        for t in tests[:30]:
            title = (t.get('title') or '')[:35]
            kb_list.append([InlineKeyboardButton(f"📝 {title}", callback_data=f"manage_test:{t.get('test_id')}")])
        kb_list.append([InlineKeyboardButton(get_bot_text('btn_back_main', lang), callback_data="back_to_main")])

        if update.callback_query:
            await update.callback_query.message.edit_text(get_bot_text('results_choose', lang), reply_markup=InlineKeyboardMarkup(kb_list), parse_mode=ParseMode.HTML)
        else:
            await update.effective_chat.send_message(get_bot_text('results_choose', lang), reply_markup=InlineKeyboardMarkup(kb_list), parse_mode=ParseMode.HTML)

async def create_choose_chat(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    lang = get_user_lang(user_id)
    chats = db.chats_for_user(user_id)
    filtered = [c for c in chats if c.get("bot_is_admin") == 1]

    kb = []
    for c in filtered[:30]:
        kb.append([InlineKeyboardButton((c.get("title") or "")[:40], callback_data=f"create_chat:{c.get('chat_id')}")])

    kb.append([InlineKeyboardButton(get_bot_text('btn_create_private', lang), callback_data="create_private")])
    kb.append([InlineKeyboardButton(get_bot_text('btn_cancel', lang), callback_data="cancel_action")])

    await update.effective_chat.send_message(get_bot_text('create_where', lang), reply_markup=InlineKeyboardMarkup(kb))

async def convert_choose_chat(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    lang = get_user_lang(user_id)
    chats = db.chats_for_user(user_id)
    filtered = [c for c in chats if c.get("bot_is_admin") == 1]

    kb = []
    for c in filtered[:30]:
        kb.append([InlineKeyboardButton((c.get("title") or "")[:40], callback_data=f"convert_chat:{c.get('chat_id')}")])

    kb.append([InlineKeyboardButton(get_bot_text('btn_create_private', lang), callback_data="convert_private")])
    kb.append([InlineKeyboardButton(get_bot_text('btn_cancel', lang), callback_data="cancel_action")])

    await update.effective_chat.send_message(get_bot_text('convert_where', lang), reply_markup=InlineKeyboardMarkup(kb), parse_mode=ParseMode.HTML)

async def show_user_chats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chats = db.chats_for_user(update.effective_user.id)
    if not chats:
        await update.effective_chat.send_message("Hali botni hech qaysi kanal/guruhga qo‘shmagansiz.")
        return

    lines = []
    for c in chats:
        status = "✅" if c.get("bot_is_admin") else "⚠️"
        lines.append(f"{status} {h(c.get('title'))}")

    await update.effective_chat.send_message("📌 Siz qo‘shgan chatlar:\n\n" + "\n".join(lines), parse_mode=ParseMode.HTML)

# ==========================================
# 12. SESSIONS (Test ishlash jarayoni)
# ==========================================
async def start_test_private(update: Update, context: ContextTypes.DEFAULT_TYPE, test_id: str):
    user_id = update.effective_user.id
    lang = get_user_lang(user_id)
    test = db.get_test(test_id)
    if not test or test.get("status") == "closed":
        await update.effective_chat.send_message("⛔ Test yopilgan yoki yo'q.")
        return

    with db._conn() as c:
        is_running = c.execute("SELECT 1 FROM sessions WHERE test_id=%s AND user_id=%s AND state='running' LIMIT 1", (test_id, user_id)).fetchone()
        if is_running:
            kb = InlineKeyboardMarkup([[InlineKeyboardButton("🛑 Sessiyani tugatish", callback_data=f"force_finish_{test_id}")]])
            await update.effective_chat.send_message(get_bot_text('test_running_err', lang), reply_markup=kb)
            return

        fc_row = c.execute("SELECT COUNT(*) as cnt FROM sessions WHERE test_id=%s AND user_id=%s AND state='finished'", (test_id, user_id)).fetchone()
        fc = fc_row["cnt"] if fc_row else 0

        limit = test.get("attempts_limit")
        if limit is None:
            limit = 1

        if limit > 0 and fc >= limit:
            await update.effective_chat.send_message(get_bot_text('test_limit_err', lang))
            return

    session_id = uuid.uuid4().hex
    db.create_session(session_id, test_id, user_id, now_ts())

    if RUNTIME["running_session"] not in context.application.bot_data:
        context.application.bot_data[RUNTIME["running_session"]] = {}

    context.application.bot_data[RUNTIME["running_session"]][user_id] = session_id
    context.application.bot_data[f"session:{session_id}:user_chat"] = update.effective_chat.id

    kb_finish = ReplyKeyboardMarkup(
        [[KeyboardButton(get_bot_text('btn_finish_test_session', lang))]],
        resize_keyboard=True
    )

    await update.effective_chat.send_message(f"✅ Boshlandi: <b>{h(test.get('title'))}</b>", parse_mode=ParseMode.HTML, reply_markup=kb_finish)
    await send_inline_question(context, session_id, test_id, 0)

async def start_public_test(update: Update, context: ContextTypes.DEFAULT_TYPE, test_id: str):
    await start_test_private(update, context, test_id)

async def send_inline_question(context: ContextTypes.DEFAULT_TYPE, session_id: str, test_id: str, q_index: int):
    qs = db.get_questions(test_id)
    if q_index >= len(qs):
        await finish_test_session(context, session_id)
        return

    chat_id = context.application.bot_data.get(f"session:{session_id}:user_chat")
    if not chat_id:
        try:
            with db._conn() as c:
                row = c.execute("SELECT user_id FROM sessions WHERE session_id=%s", (session_id,)).fetchone()
                if row:
                    chat_id = row["user_id"]
        except Exception:
            pass

    if not chat_id:
        return

    q = qs[q_index]
    options = json.loads(q["options_json"])
    photo_id = q.get("photo_id")

    kb = []
    for i in range(len(options)):

        letter = chr(65 + i)
        kb.append([InlineKeyboardButton(letter, callback_data=f"ans:{session_id}:{q_index}:{i}")])

    msg = build_question_html(q_index, q["question"], options)
    context.application.bot_data[f"qstart:{session_id}:{q_index}"] = now_ts()

    if photo_id:
        await context.bot.send_photo(chat_id=chat_id, photo=photo_id, caption=msg, parse_mode=ParseMode.HTML, reply_markup=InlineKeyboardMarkup(kb))
    else:
        await context.bot.send_message(chat_id=chat_id, text=msg, parse_mode=ParseMode.HTML, reply_markup=InlineKeyboardMarkup(kb))

    db.set_session_current_q(session_id, q_index)

def build_question_html(q_index: int, qtext: str, options: list[str]) -> str:
    lines = [f"<b>{q_index+1})</b> {h(qtext)}", ""]
    for i, opt in enumerate(options):

        letter = chr(65 + i)
        lines.append(f"<b>{letter})</b> {h(opt)}")
    return "\n".join(lines)

async def handle_inline_answer(update: Update, context: ContextTypes.DEFAULT_TYPE, session_id: str, q_index: int, opt_index: int):
    with db._conn() as c:
        s = c.execute("SELECT * FROM sessions WHERE session_id=%s", (session_id,)).fetchone()

    if not s or s.get("state") != "running":
        return

    qs = db.get_questions(s.get("test_id"))
    if q_index >= len(qs):
        return

    is_correct = 1 if opt_index == int(qs[q_index]["correct_index"]) else 0
    db.upsert_answer(session_id, q_index, opt_index, is_correct, now_ts(), 0)

    try:
        opt = json.loads(qs[q_index]["options_json"])
        base = build_question_html(q_index, qs[q_index]["question"], opt)
        extra = f"\n\n{'✅' if is_correct else '❌'} Tanlangani: <b>{h(opt[opt_index])}</b>"
        photo_id = qs[q_index].get("photo_id")

        if photo_id:
            await update.callback_query.message.edit_caption(caption=base + extra, parse_mode=ParseMode.HTML)
        else:
            await update.callback_query.message.edit_text(text=base + extra, parse_mode=ParseMode.HTML)
    except Exception:
        pass

    await asyncio.sleep(0.4)
    await send_inline_question(context, session_id, s.get("test_id"), q_index + 1)

async def finish_test_session(context: ContextTypes.DEFAULT_TYPE, session_id: str):
    with db._conn() as c:
        s = c.execute("SELECT * FROM sessions WHERE session_id=%s", (session_id,)).fetchone()
        ans = c.execute("SELECT COUNT(*) as total, SUM(is_correct) as correct FROM answers WHERE session_id=%s", (session_id,)).fetchone()

    if not s:
        return

    duration = max(1, now_ts() - int(s.get("started_at", 0)))
    db.finish_session(session_id, now_ts(), int(ans.get("correct") or 0), duration)

    chat_id = context.application.bot_data.get(f"session:{session_id}:user_chat")
    if not chat_id:
        chat_id = s.get("user_id")

    if chat_id:
        user_id = s.get("user_id")
        lang = get_user_lang(user_id)
        msg, kb_main = await build_main_menu(user_id, context.bot.username, lang)

        await context.bot.send_message(
            chat_id=chat_id,
            text=f"✅ Yakunlandi!\nNatija saqlandi!\n\n{msg}",
            reply_markup=kb_main,
            parse_mode=ParseMode.HTML
        )

async def job_finalize_deadline(context: ContextTypes.DEFAULT_TYPE):
    await finalize_test_by_id(context, context.job.data["test_id"], announce=True)

async def finalize_test(update: Update, context: ContextTypes.DEFAULT_TYPE, test_id: str, manual: bool = False, announce: bool = True):
    await finalize_test_by_id(context, test_id, announce)
    if manual:
        await update.effective_chat.send_message("✅ Test yopildi.")

async def finalize_test_by_id(context: ContextTypes.DEFAULT_TYPE, test_id: str, announce: bool = True):
    test = db.get_test(test_id)
    if not test:
        return

    lang = get_user_lang(int(test.get("owner_user_id", 0)))
    scoring_type = test.get("scoring_type", "standard")

    qcount, participants = db.stats(test_id)
    top = db.leaderboard(test_id, limit=20)

    lines = []
    for i, r in enumerate(top, 1):
        score_str = format_display_score(r.get("score"), scoring_type, lang)
        lines.append(texts.medal_line_html(i, format_user_display(r.get("username"), r.get("first_name"), r.get("last_name"), r.get("user_id")), score_str, fmt_duration(r.get("duration_sec") or 0, lang), lang))

    msg = texts.leaderboard_template_html(test.get("title"), qcount, test.get("per_question_sec"), participants, lines, lang)

    if announce:
        if test.get("chat_id"):
            try:
                await context.bot.send_message(chat_id=test.get("chat_id"), text=msg, parse_mode=ParseMode.HTML)
            except Exception:
                pass

        try:
            await context.bot.send_message(chat_id=REQUIRED_CHANNEL, text=msg, parse_mode=ParseMode.HTML)
        except Exception as e:
            logging.error(f"Kanalga natija tashlashda xatolik: {e}")

    db.close_test(test_id)

# ==========================================
# 13. EXCEL VA WORD PARSE
# ==========================================
async def export_excel(update: Update, context: ContextTypes.DEFAULT_TYPE, test_id: str):
    test = db.get_test(test_id)
    if not test:
        await update.effective_chat.send_message("Test topilmadi.")
        return

    user_id = update.effective_user.id
    lang = get_user_lang(user_id)
    scoring_type = test.get("scoring_type", "standard")

    allr = db.all_results(test_id)
    wb = Workbook()
    ws = wb.active

    headers = ["Place", "User", "Score", "Duration"]
    for col, hname in enumerate(headers, 1):
        ws.cell(row=1, column=col).value = hname

    for i, r in enumerate(allr, 1):
        ws.cell(row=i+1, column=1).value = i
        ws.cell(row=i+1, column=2).value = format_user_display(r.get("username"), r.get("first_name"), r.get("last_name"), r.get("user_id"))
        ws.cell(row=i+1, column=3).value = format_display_score(r.get("score"), scoring_type, lang)
        ws.cell(row=i+1, column=4).value = fmt_duration(r.get("duration_sec") or 0, lang)

    path = f"export_{test_id}.xlsx"
    wb.save(path)

    with open(path, "rb") as f:
        await update.effective_chat.send_document(document=f, filename=path)

    os.remove(path)

async def on_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_chat.type != "private":
        return

    await upsert_user_from_update(update)
    doc = update.message.document
    file_name = (doc.file_name or "").lower()

    if not file_name.endswith(".docx"):
        await update.effective_chat.send_message("❌ Iltimos, faqat Word (.docx) formatidagi faylni yuboring.")
        return

    if context.user_data.get(K["mode"]) != "convert":
        context.user_data[K["mode"]] = "convert"
        context.user_data[K["convert"]] = {"chat_id": None}

    target_chat_id = context.user_data[K["convert"]].get("chat_id")

    await update.effective_chat.send_message("⏳ Word fayli qabul qilindi, matn va rasmlar tekshirilmoqda...")

    file = await context.bot.get_file(doc.file_id)
    with tempfile.NamedTemporaryFile(delete=False, suffix=".docx") as tmp:
        await file.download_to_drive(tmp.name)
        try:
            items = extract_data_from_docx(tmp.name)
            theme, questions, error = build_test_from_items(items)
        except Exception as e:
            await update.effective_chat.send_message(f"❌ Faylni o'qishda xato yuz berdi: {e}")
            if os.path.exists(tmp.name): os.remove(tmp.name)
            return

    if os.path.exists(tmp.name): os.remove(tmp.name)

    if error:
        await update.effective_chat.send_message(f"❌ Xato: {error}")
        return

    missing_keys = any(q.get("correct_index") is None or q.get("correct_index") < 0 for q in questions)

    context.user_data[K["convert"]]["pending_test"] = {
        "theme": theme,
        "questions": questions,
        "target_chat": target_chat_id,
        "missing_keys": missing_keys
    }

    if not theme or theme.strip().upper() == "UNKNOWN":
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("🤖 AI mavzu topsin", callback_data="pdf_ai_theme")],
            [InlineKeyboardButton("✍️ O'zim kiritaman", callback_data="pdf_manual_theme")]
        ])
        await update.effective_chat.send_message(
            "⚠️ Test mavzusi aniqlanmadi (yoki fayl nomi mavzuga mos emas).\n\nNima qilamiz?",
            reply_markup=kb, parse_mode=ParseMode.HTML
        )
        return

    if missing_keys:
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("🤖 AI o'zi yechib bersin", callback_data="pdf_ai_solve")],
            [InlineKeyboardButton("✍️ Kalitlarni o'zim kiritaman", callback_data="pdf_manual_keys")]
        ])
        await update.effective_chat.send_message(
            "⚠️ Hujjat muvaffaqiyatli o'qildi, lekin <b>to'g'ri javoblar (kalitlar)</b> topilmadi!\n\n"
            "<i>(Eslatma: Agar testda grafiklar yoki rasmli savollar bo'lsa, AI ularni noto'g'ri yechishi mumkin. Shunday paytda kalitlarni o'zingiz kiritganingiz ma'qul)</i>\n\nNima qilamiz?",
            reply_markup=kb, parse_mode=ParseMode.HTML
        )
        return

    await build_and_send_test(update, context, theme, questions, target_chat_id)

# ==========================================
# 14. TIZIMNI ISHGA TUSHIRISH
# ==========================================
async def verify_all_chats_on_startup(app: Application):
    logging.info("Barcha chatlardagi adminlik huquqlari tekshirilmoqda...")
    with db._conn() as c:
        rows = c.execute("SELECT chat_id FROM chats WHERE bot_is_admin=1").fetchall()

    if not rows:
        logging.info("Tekshirish uchun faol chatlar topilmadi.")
        return

    bot_id = app.bot.id
    removed_count = 0

    for r in rows:
        chat_id = int(r["chat_id"])
        try:
            member = await app.bot.get_chat_member(chat_id, bot_id)
            if member.status not in [ChatMemberStatus.ADMINISTRATOR, ChatMemberStatus.OWNER]:
                db.set_bot_admin(chat_id, 0, now_ts())
                removed_count += 1
        except Exception as e:
            db.set_bot_admin(chat_id, 0, now_ts())
            removed_count += 1
        await asyncio.sleep(0.05)

    logging.info(f"Tekshiruv tugadi! {len(rows)} ta chatdan {removed_count} tasida bot adminlikdan tushgan (yoki haydalgan).")

async def restore_deadlines(app: Application):
    with db._conn() as c:
        rows = c.execute("SELECT test_id, deadline_ts FROM tests WHERE deadline_ts IS NOT NULL AND status != 'closed'").fetchall()
    for r in rows:
        delay = int(r.get("deadline_ts", 0)) - now_ts()
        if delay <= 0:
            pass
        else:
            app.job_queue.run_once(job_finalize_deadline, delay, data={"test_id": r.get("test_id")})

async def cmd_ommaviy_tekshiruv(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    lang = get_user_lang(user_id)

    if user_id not in SUPERADMINS:
        return

    try:
        with db._conn() as c:
            row = c.execute("SELECT COUNT(*) as cnt FROM users").fetchone()
            total_users = row['cnt'] if row else 0
    except Exception as e:
        total_users = 0
        logging.error(f"Foydalanuvchilarni sanashda xato: {e}")

    text = (
        f"📊 <b>Foydalanuvchilarni ommaviy boshqarish!</b>\n\n"
        f"Bazada jami <b>{total_users}</b> ta foydalanuvchi bor.\n\n"
        f"Ularning barchasini bot deb belgilaysizmi (hammadan kapcha so'raladi) "
        f"yoki bot tasdiqlashidan ozod qilasizmi (hech kimdan so'ralmaydi)?"
    )

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ Hammaga ruxsat berish (Ozod qilish)", callback_data="mass_verify_1")],
        [InlineKeyboardButton("🤖 Hammani bot deb belgilash", callback_data="mass_verify_0")],
        [InlineKeyboardButton(get_bot_text('btn_cancel', lang), callback_data="cancel_action")]
    ])

    await update.message.reply_text(text, reply_markup=kb, parse_mode=ParseMode.HTML)

if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s')

    async def post_init(app: Application):
        try:
            with db._conn() as c:
                c.execute("ALTER TABLE users ADD COLUMN pin_code VARCHAR(255) DEFAULT NULL;")
                c.execute("ALTER TABLE users ADD COLUMN custom_bg VARCHAR(255) DEFAULT NULL;")
                c.execute("ALTER TABLE users ADD COLUMN custom_lock_bg VARCHAR(255) DEFAULT NULL;")
                c.execute("ALTER TABLE users ADD COLUMN lang VARCHAR(10) DEFAULT 'uz';")
                c.execute("ALTER TABLE users MODIFY COLUMN is_verified TINYINT DEFAULT 1;")
                c.execute("UPDATE users SET is_verified = 1 WHERE is_verified = 0 OR is_verified IS NULL;")
        except Exception as e:
            pass
        await restore_deadlines(app)
        await verify_all_chats_on_startup(app)


    app = Application.builder().token(BOT_TOKEN).post_init(post_init).build()



    # 1. Middlewares va Maxsus tekshiruvlar (Manfiy guruhlarda yoziladi, birinchi ishlaydi)
    app.add_handler(TypeHandler(Update, spam_check_middleware), group=-5)
    app.add_handler(CallbackQueryHandler(handle_pin_callback, pattern="^(pin_|forgot_pin|cancel_reset_pin|reset_pin_)"), group=-4)
    app.add_handler(TypeHandler(Update, check_lock_middleware), group=-3)
    app.add_handler(TypeHandler(Update, check_verified_middleware), group=-2)
    app.add_handler(TypeHandler(Update, check_subscription_middleware), group=-1)

    # 2. Asosiy buyruqlar (/start, /cabinet va boshqalar)
    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(CommandHandler("cabinet", cmd_cabinet))
    app.add_handler(CommandHandler("reward_top", cmd_reward_top))
    app.add_handler(CommandHandler(["ommaviy", "ommaviy_tekshiruv"], cmd_ommaviy_tekshiruv))

    # 3. Tugmalar va Matnlar (Messages & Callbacks)
    app.add_handler(CallbackQueryHandler(on_callback))
    app.add_handler(CallbackQueryHandler(on_callback_postprocess), group=1)
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, on_text))
    app.add_handler(MessageHandler(filters.PHOTO | filters.VIDEO | filters.VOICE, on_media))
    app.add_handler(MessageHandler(filters.Document.ALL, on_document))

    # 4. Guruhlar/Kanallar va To'lovlar (Telegram Stars)
    app.add_handler(ChatMemberHandler(on_my_chat_member, ChatMemberHandler.MY_CHAT_MEMBER))
    app.add_handler(ChatMemberHandler(on_chat_member, ChatMemberHandler.CHAT_MEMBER))
    app.add_handler(PreCheckoutQueryHandler(precheckout_callback))
    app.add_handler(MessageHandler(filters.SUCCESSFUL_PAYMENT, successful_payment_callback))

    # Botni ishga tushirish
    print("Bot muvaffaqiyatli ishga tushdi va xabarlarni kutmoqda...")
    app.run_polling(allowed_updates=Update.ALL_TYPES)